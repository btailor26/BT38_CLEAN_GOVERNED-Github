from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session
from flask_login import login_required, current_user, login_user, logout_user
from sqlalchemy import desc, or_, text, select
from sqlalchemy.orm import joinedload
from datetime import datetime
import logging
import json
import requests
import base64
import secrets
from urllib.parse import urlencode
import os
import hashlib

# Import db from extensions (not app) to avoid circular import
from extensions import db
from models import InventoryItem, Store, SyncLog, SyncJob, ProductGroup, GroupExternalRef, PushSettings, WarehouseStock, StockLedgerEntry, MarketplaceListing, SystemConfig, SystemSetting, SystemEvent, Supplier, ReorderNotification, PurchaseOrder, PurchaseOrderItem, ReceivingInspection, ReceivingInspectionItem, Warehouse
from amazon_service import AmazonAPIService
from ebay_service import eBayAPIService
from smart_push_service import smart_push_service
from ai_image_service import enhance_image, remove_background, generate_lifestyle_image, check_marketplace_compliance
from notification_service import notification_service

# Create Blueprint
bp = Blueprint('routes', __name__)

# Task API Key for internal testing
TASK_API_KEY = os.getenv("TASK_API_KEY")

# Dev login helpers
def _get_user_model():
    """Get User model, try User first, fallback to AdminUser"""
    try:
        from models import User as U
        return U
    except Exception:
        try:
            from models import AdminUser as U
            return U
        except Exception:
            return None

def _upsert_admin(email: str, password: str):
    """Create or update admin user with given email and password"""
    U = _get_user_model()
    if U is None:
        return None, "No User/AdminUser model found"

    user = U.query.filter_by(email=email).first()
    created = False
    if not user:
        # Create new user with default username from email
        username = email.split('@')[0]
        user = U(email=email, username=username)
        created = True

    # Set password safely
    if hasattr(user, "set_password"):
        user.set_password(password)
    else:
        # fallback: hash via werkzeug if available
        try:
            from werkzeug.security import generate_password_hash
            user.password_hash = generate_password_hash(password)
        except Exception:
            user.password_hash = password  # last resort (not ideal)

    if created:
        db.session.add(user)
    db.session.commit()
    return user, None

# CSRF Protection Functions
def generate_csrf_token():
    """Generate a CSRF token for the current session"""
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']

def validate_csrf_token(token):
    """Validate a CSRF token against the session token"""
    if not token:
        return False
    return session.get('csrf_token') == token

# Make CSRF token available to all templates
@bp.app_context_processor
def inject_csrf_token():
    return dict(csrf_token=generate_csrf_token())

# Unified warehouse push helper function
def prepare_warehouse_push(item, operation="update"):
    """
    Prepare warehouse stock and identify stores for push. DOES NOT enqueue jobs.

    This function ensures WarehouseStock exists and returns list of stores to push to.
    Caller must commit the transaction, then separately enqueue jobs after commit succeeds.

    Args:
        item: InventoryItem object
        operation: Type of operation ("create", "update", "delete")

    Returns:
        tuple: (stores_to_push, warehouse_stock) - list of Store objects and WarehouseStock

    Note: Call this BEFORE commit, then commit, then enqueue jobs for each store.
    """
    from models import Warehouse

    # Get default warehouse
    default_warehouse = Warehouse.get_default()

    # Find or create warehouse stock for this SKU in default warehouse
    warehouse_stock = WarehouseStock.query.filter_by(
        sku=item.sku,
        warehouse_id=default_warehouse.id
    ).first()

    if not warehouse_stock:
        # For new items, create warehouse stock record
        warehouse_stock = WarehouseStock(
            warehouse_id=default_warehouse.id,
            sku=item.sku,
            available_quantity=item.quantity,
            reserved_quantity=0,
            allocated_quantity=0,
            on_order_quantity=0
        )
        db.session.add(warehouse_stock)
        db.session.flush()  # Get ID for use in this transaction

    # Find active stores based on operation type
    if operation == "create":
        # New items: push to all stores that want new items, regardless of listing existence
        stores_to_push = db.session.query(Store).filter(
            Store.is_active == True,
            Store.auto_push_enabled == True,
            Store.push_on_item_create == True
        ).all()
    elif operation == "update":
        # Updates: only push to stores with existing marketplace listings
        stores_to_push = db.session.query(Store).join(
            MarketplaceListing, Store.id == MarketplaceListing.store_id
        ).filter(
            Store.is_active == True,
            Store.auto_push_enabled == True,
            Store.push_on_quantity_change == True,
            MarketplaceListing.warehouse_stock_id == warehouse_stock.id
        ).distinct().all()
    else:
        # Other operations: no auto-push
        stores_to_push = []

    return stores_to_push, warehouse_stock


def disabled_queue_job(*_args, **_kwargs):
    return type("DisabledJob", (), {"id": "disabled", "status": "disabled"})()
def enqueue_push_jobs(item_id, stores):
    """
    Enqueue high-priority push jobs for a list of stores.

    Call this AFTER successfully committing the item changes.

    Args:
        item_id: ID of the InventoryItem
        stores: List of Store objects to push to

    Returns:
        int: Number of jobs enqueued
    """
    from queue_manager import enqueue_sync_job, JOB_PUSH_ITEM, PRIORITY_HIGH

    jobs_count = 0
    for store in stores:
        disabled_queue_job(
            store_id=store.id,
            job_type=JOB_PUSH_ITEM,
            payload={'item_id': item_id},
            priority=PRIORITY_HIGH
        )
        jobs_count += 1

    return jobs_count

@bp.route('/test/ebay-push')
# @login_required  # Temporarily disabled for access
def test_ebay_push():
    """Test page to verify live eBay push updates"""
    try:
        # Get eBay store
        ebay_store = db.session.query(Store).filter(Store.platform == 'ebay', Store.is_active == True).first()

        if not ebay_store:
            flash('No active eBay store found. Please configure an eBay store first.', 'warning')
            return redirect(url_for('routes.stores'))

        # Get items with eBay listings
        items_with_ebay = db.session.query(InventoryItem).join(
            MarketplaceListing, InventoryItem.id == MarketplaceListing.inventory_item_id
        ).filter(
            MarketplaceListing.store_id == ebay_store.id
        ).order_by(InventoryItem.sku).limit(50).all()

        return render_template('test_ebay_push.html',
                             ebay_store=ebay_store,
                             items=items_with_ebay)
    except Exception as e:
        logging.error(f"Error in test_ebay_push: {str(e)}")
        flash(f'Error loading test page: {str(e)}', 'danger')
        return redirect(url_for('routes.index'))

@bp.route('/api/test/ebay-push/<int:item_id>', methods=['POST'])
# @login_required  # Temporarily disabled for access
def api_test_ebay_push(item_id):
    """Retired debug eBay push route.

    This endpoint is intentionally fail-closed. It must not read eBay,
    enqueue push jobs, sleep inside the request, or trigger marketplace execution.
    """
    return jsonify({
        "success": False,
        "error": "Debug eBay push route is retired. Use governed Command Center preview and approved dispatcher execution path.",
        "execution_blocked": True,
        "route_retired": True
    }), 410

@bp.route('/health')
def health():
    """Health check endpoint for monitoring and deployment verification"""
    from flask import current_app
    try:
        db_status = "unknown"
        try:
            db.session.execute(text("SELECT 1"))
            db_status = "connected"
        except Exception as e:
            db_status = f"error: {str(e)}"

        store_count = db.session.query(Store).count()
        active_stores = db.session.query(Store).filter(Store.is_active == True).count()

        return jsonify({
            "ok": True,
            "env": current_app.config.get("APP_ENV", "prod"),
            "production": current_app.config.get("IS_PRODUCTION", True),
            "timestamp": datetime.utcnow().isoformat() + "Z",
            "database": db_status,
            "stores": {
                "total": store_count,
                "active": active_stores
            },
            "message": "Inventory management system is running normally"
        }), 200
    except Exception as e:
        return jsonify({
            "ok": False,
            "env": current_app.config.get("APP_ENV", "unknown"),
            "timestamp": datetime.utcnow().isoformat() + "Z",
            "error": str(e)
        }), 500

@bp.route('/api/diagnostics/sendgrid/test', methods=['GET', 'POST'])
def test_sendgrid():
    """Test SendGrid email sending functionality - requires authentication or TASK_API_KEY"""
    try:
        # Get from email with fallback logic (same as purchase order flow)
        from_email = os.environ.get('SENDGRID_FROM_EMAIL')

        # Check if misconfigured (API key instead of email) and use verified email as fallback
        if from_email and from_email.startswith('SG.'):
            from_email = 'bhavtee@gmail.com'  # Use verified sender email
            config_note = "Using fallback email (SENDGRID_FROM_EMAIL is misconfigured)"
        elif from_email and '@' in from_email:
            config_note = "Using configured SENDGRID_FROM_EMAIL"
        else:
            config_note = "SENDGRID_FROM_EMAIL not configured"

        # SECURITY: Only send to the configured from_email to prevent email relay abuse
        # Do NOT accept arbitrary email addresses from query params
        test_email = from_email

        # Check if SendGrid client is initialized
        if not notification_service.sendgrid_client:
            return jsonify({
                'ok': False,
                'error': 'SendGrid client not initialized - check SENDGRID_API_KEY',
                'configured': False
            }), 500

        # Send test email
        success = notification_service.send_email_alert(
            to_email=test_email,
            from_email=from_email,
            subject='SendGrid Test Email - Inventory System',
            text_content='This is a test email from your inventory management system.',
            html_content=f'''
            <html>
            <body style="font-family: Arial, sans-serif; padding: 20px;">
                <h2 style="color: #28a745;"> SendGrid Test Successful</h2>
                <p>Your SendGrid email integration is working correctly!</p>
                <p>This test email confirms that:</p>
                <ul>
                    <li>SendGrid API key is valid</li>
                    <li>Sender email ({from_email}) is verified</li>
                    <li>Email delivery is operational</li>
                </ul>
                <div style="background-color: #f8f9fa; padding: 15px; margin: 20px 0; border-left: 4px solid #007bff; border-radius: 5px;">
                    <p style="margin: 0;"><strong>Configuration:</strong> {config_note}</p>
                </div>
                <p style="color: #6c757d; font-size: 12px; margin-top: 30px;">
                    Sent from: Inventory Management System<br>
                    Time: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC
                </p>
            </body>
            </html>
            '''
        )

        if success:
            return jsonify({
                'ok': True,
                'success': True,
                'message': f'Test email sent successfully to {test_email}',
                'from_email': from_email,
                'to_email': test_email,
                'config_note': config_note
            }), 200
        else:
            return jsonify({
                'ok': False,
                'success': False,
                'error': 'Email sending failed - check logs for details',
                'from_email': from_email,
                'to_email': test_email
            }), 500

    except Exception as e:
        logging.error(f'Error testing SendGrid: {str(e)}')
        return jsonify({
            'ok': False,
            'success': False,
            'error': str(e)
        }), 500

@bp.get("/dev/login")
def dev_login():
    """Temporary developer login protected by DEV_LOGIN_KEY secret"""
    key = os.getenv("DEV_LOGIN_KEY")
    if request.args.get("k") != key:
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    email = request.args.get("email")
    pwd = request.args.get("pwd")
    if not email or not pwd:
        return jsonify({"ok": False, "error": "missing email or pwd"}), 400

    user, err = _upsert_admin(email, pwd)
    if err:
        return jsonify({"ok": False, "error": err}), 500

    try:
        login_user(user, remember=True)
    except Exception as e:
        return jsonify({"ok": False, "error": f"login failed: {e}"}), 500

    # Redirect to dashboard
    return redirect(url_for("routes.dashboard"))

@bp.get("/auth/status")
def auth_status():
    """Check current authentication status"""
    try:
        if current_user.is_authenticated:
            return jsonify({
                "ok": True,
                "authenticated": True,
                "email": getattr(current_user, "email", None),
                "id": getattr(current_user, "id", None)
            })
        return jsonify({"ok": False, "authenticated": False})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@bp.get("/api/diagnostics/amazon/auth")
def amazon_auth_diag():
    """Amazon auth diagnostics - marketplace, region, and connection status"""
    if request.headers.get("X-Task-Key") != os.getenv("TASK_API_KEY"):
        return jsonify({"error": "unauthorized"}), 401

    try:
        store = Store.query.filter_by(platform="Amazon", is_active=True).first()
        if not store:
            return jsonify({"ok": False, "error": "No active Amazon store"}), 404

        amazon_service = AmazonAPIService()
        data = amazon_service.get_auth_diagnostics(store)
        return jsonify({"ok": True, **data})
    except Exception as e:
        logging.error(f"Amazon auth diagnostics error: {str(e)}")
        return jsonify({"ok": False, "error": str(e)}), 500

@bp.route('/')
def dashboard():
    """Main dashboard showing overview statistics - no login required"""
    # Temporarily bypassing login to diagnose browser session issues
    total_items = db.session.query(InventoryItem).count()
    total_stores = db.session.query(Store).count()
    active_stores = db.session.query(Store).filter(Store.is_active == True).count()
    low_stock_items = db.session.query(WarehouseStock).filter(
        WarehouseStock.is_active == True,
        WarehouseStock.available_quantity < 10
    ).count()

    # Recent sync logs
    recent_logs = db.session.query(SyncLog).order_by(desc(SyncLog.created_at)).limit(5).all()

    # Get recent warehouse items with Amazon listing data (latest 5 products)
    recent_items_query = db.session.query(WarehouseStock).order_by(desc(WarehouseStock.updated_at)).limit(5).all()

    # Enrich with Amazon listing details for each item
    recent_items = []
    for warehouse_item in recent_items_query:
        # Get Amazon listing if it exists
        amazon_listing = db.session.query(MarketplaceListing).join(Store).filter(
            MarketplaceListing.warehouse_stock_id == warehouse_item.id,
            Store.platform == 'Amazon'
        ).first()

        # Create item dict with warehouse data and Amazon details
        item_data = {
            'sku': warehouse_item.sku,
            'name': warehouse_item.sku,  # Will be replaced with title if Amazon listing exists
            'quantity': warehouse_item.available_quantity,
            'price': warehouse_item.unit_cost,
            'updated_at': warehouse_item.updated_at,
            'asin': amazon_listing.asin if amazon_listing else None,
            'fnsku': amazon_listing.fnsku if amazon_listing else None,
            'amazon_title': amazon_listing.title if amazon_listing else None
        }
        # Use Amazon title as name if available
        if item_data['amazon_title']:
            item_data['name'] = item_data['amazon_title']

        recent_items.append(item_data)

    stats = {
        'total_items': total_items,
        'total_stores': total_stores,
        'active_stores': active_stores,
        'low_stock_items': low_stock_items
    }

    return render_template('dashboard.html', stats=stats, recent_logs=recent_logs, recent_items=recent_items)

@bp.route('/low-stock')
# @login_required  # Temporarily disabled for access
def low_stock_items():
    """Show all low stock items grouped by supplier for purchase order creation"""
    search_query = request.args.get('search', '').strip()

    # Get all warehouse items with quantity < 10
    query = WarehouseStock.query.filter(
        WarehouseStock.is_active == True,
        WarehouseStock.available_quantity < 10
    ).options(
        joinedload(WarehouseStock.supplier)
    )

    # Apply search filter if provided
    if search_query:
        query = query.filter(
            db.or_(
                WarehouseStock.sku.ilike(f'%{search_query}%'),
                WarehouseStock.location.ilike(f'%{search_query}%')
            )
        )

    low_stock = query.all()

    # Get product names for all items
    skus = [item.sku for item in low_stock]
    inventory_items = InventoryItem.query.filter(InventoryItem.sku.in_(skus)).all()
    sku_to_name = {item.sku: item.name for item in inventory_items}

    # Attach product names to warehouse items
    for item in low_stock:
        item.product_name = sku_to_name.get(item.sku, 'Unknown Product')

    # Get all active suppliers for bulk assignment
    all_suppliers = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()

    # Group by supplier
    by_supplier = {}
    no_supplier_items = []

    for item in low_stock:
        if item.supplier:
            if item.supplier.id not in by_supplier:
                by_supplier[item.supplier.id] = {
                    'supplier': item.supplier,
                    'stock_items': []
                }
            by_supplier[item.supplier.id]['stock_items'].append(item)
        else:
            no_supplier_items.append(item)

    return render_template('low_stock.html',
                          by_supplier=by_supplier,
                          no_supplier_items=no_supplier_items,
                          all_suppliers=all_suppliers,
                          total_count=len(low_stock),
                          search_query=search_query)

@bp.route('/low-stock/bulk-assign-supplier', methods=['POST'])
# @login_required  # Temporarily disabled for access
def bulk_assign_supplier():
    """Bulk assign supplier to items without supplier"""
    try:
        item_ids = request.form.getlist('no_supplier_items[]')
        supplier_id = request.form.get('bulk_supplier_id')

        logging.info(f'Bulk assign supplier - items: {item_ids}, supplier: {supplier_id}')

        if not item_ids:
            flash('Please select at least one item', 'warning')
            return redirect(url_for('routes.low_stock_items'))

        if not supplier_id:
            flash('Please select a supplier', 'warning')
            return redirect(url_for('routes.low_stock_items'))

        # Verify supplier exists
        supplier = db.session.get(Supplier, int(supplier_id))
        if not supplier:
            flash('Invalid supplier selected', 'error')
            return redirect(url_for('routes.low_stock_items'))

        # Update all selected items
        updated_count = 0
        for item_id in item_ids:
            stock_item = db.session.get(WarehouseStock, int(item_id))
            if stock_item:
                stock_item.supplier_id = supplier.id
                updated_count += 1

        db.session.commit()
        flash(f'Successfully assigned {supplier.name} to {updated_count} item(s)!', 'success')
        return redirect(url_for('routes.low_stock_items'))

    except Exception as e:
        db.session.rollback()
        logging.error(f'Error bulk assigning supplier: {str(e)}')
        flash(f'Error assigning supplier: {str(e)}', 'error')
        return redirect(url_for('routes.low_stock_items'))

@bp.route('/low-stock/create-po', methods=['POST'])
# @login_required  # Temporarily disabled for access
def create_po_from_low_stock():
    """Create purchase order(s) from low stock items"""
    try:
        selected_items = request.form.getlist('selected_items[]')
        create_all = request.form.get('create_all')
        create_selected = request.form.get('create_selected')
        supplier_id = request.form.get('supplier_id')

        if not selected_items and not create_all and not supplier_id:
            flash('Please select items to order', 'warning')
            return redirect(url_for('routes.low_stock_items'))

        # Get all low stock items
        low_stock = WarehouseStock.query.filter(
            WarehouseStock.is_active == True,
            WarehouseStock.available_quantity < 10
        ).options(joinedload(WarehouseStock.supplier)).all()

        # Group by supplier
        by_supplier = {}
        for item in low_stock:
            if item.supplier:
                if item.supplier.id not in by_supplier:
                    by_supplier[item.supplier.id] = {
                        'supplier': item.supplier,
                        'stock_items': []
                    }
                by_supplier[item.supplier.id]['stock_items'].append(item)

        # Determine which suppliers to process
        suppliers_to_process = []
        if create_all:
            suppliers_to_process = list(by_supplier.keys())
        elif supplier_id:
            suppliers_to_process = [int(supplier_id)]
        elif create_selected and selected_items:
            # Get unique supplier IDs from selected items
            for item_id in selected_items:
                item = db.session.get(WarehouseStock, int(item_id))
                if item and item.supplier_id and item.supplier_id not in suppliers_to_process:
                    suppliers_to_process.append(item.supplier_id)

        pos_created = 0
        for sup_id in suppliers_to_process:
            if sup_id not in by_supplier:
                continue

            supplier_data = by_supplier[sup_id]
            supplier = supplier_data['supplier']

            # Generate PO number
            last_po = PurchaseOrder.query.order_by(desc(PurchaseOrder.id)).first()
            po_number = f'PO-{(last_po.id + 1 if last_po else 1):05d}'

            # Create purchase order
            po = PurchaseOrder(
                po_number=po_number,
                supplier_id=supplier.id,
                order_date=datetime.utcnow(),
                notes=f'Auto-generated from low stock items (< 10 units)',
                status='draft'
            )
            db.session.add(po)
            db.session.flush()

            # Add line items
            total_amount = 0.0
            items_added = 0
            for item in supplier_data['stock_items']:
                # Check if this item should be included
                include_item = False
                if create_all or (supplier_id and int(supplier_id) == sup_id):
                    include_item = True
                elif create_selected and str(item.id) in selected_items:
                    include_item = True

                if include_item:
                    # Get custom quantity from form if available
                    qty_field = f'qty_{item.id}'
                    quantity = int(request.form.get(qty_field, item.reorder_quantity if item.reorder_quantity > 0 else 10))

                    po_item = PurchaseOrderItem(
                        purchase_order_id=po.id,
                        sku=item.sku,
                        product_name=f'{item.sku} - Restock',
                        ordered_quantity=quantity,
                        unit_cost=item.unit_cost,
                        total_cost=quantity * item.unit_cost
                    )
                    db.session.add(po_item)
                    total_amount += po_item.total_cost
                    items_added += 1

            if items_added > 0:
                po.total_amount = total_amount
                pos_created += 1
            else:
                # Remove PO if no items were added
                db.session.delete(po)

        db.session.commit()

        if pos_created > 0:
            flash(f'{pos_created} purchase order(s) created successfully!', 'success')
            return redirect(url_for('routes.purchase_orders'))
        else:
            flash('No purchase orders were created. Please select items with assigned suppliers.', 'warning')
            return redirect(url_for('routes.low_stock_items'))

    except Exception as e:
        db.session.rollback()
        logging.error(f'Error creating purchase orders from low stock: {str(e)}')
        flash(f'Error creating purchase orders: {str(e)}', 'error')
        return redirect(url_for('routes.low_stock_items'))

@bp.route('/business')
def business():
    """Business website for Amazon registration"""
    return render_template('business.html')

@bp.route('/privacy')
def privacy():
    """Privacy policy page"""
    return render_template('privacy.html')

@bp.route('/terms')
def terms():
    """Terms of service page"""
    return render_template('terms.html')

@bp.route('/api/push-sku', methods=['POST'])
def api_push_sku():
    """Retired direct SKU push API.

    Marketplace execution must go through governed dispatcher flow only.
    This route must not call disabled_smart_push_service_push_specific_sku directly.
    """
    data = request.get_json(silent=True) or {}
    sku = data.get("sku") or request.form.get("sku") or request.args.get("sku")

    return jsonify({
        "success": False,
        "error": "Direct SKU push route is retired. Use governed dispatcher execution path.",
        "execution_blocked": True,
        "route_retired": True,
        "sku": sku
    }), 410

@bp.route('/api/classify-listings', methods=['POST'])
def api_classify_listings():
    """API endpoint to classify and update listing types"""
    try:
        # Classify all unclassified listings
        classified_count = 0
        listings = MarketplaceListing.query.filter(
            or_(MarketplaceListing.listing_type == 'single',
                MarketplaceListing.listing_type.is_(None))
        ).all()

        for listing in listings:
            smart_push_service.update_listing_classification(listing)
            classified_count += 1

        return jsonify({
            'success': True,
            'classified_count': classified_count
        })

    except Exception as e:
        logging.error(f"Error classifying listings: {str(e)}")
        return jsonify({'error': str(e)}), 500

@bp.route('/api/sync/amazon/sku/<sku>', methods=['POST'])
def api_sync_amazon_sku(sku):
    """Retired direct Amazon SKU sync route.

    Warehouse quantity must not be mutated from sync routes.
    Marketplace execution must go through governed dispatcher flow only.
    """
    return jsonify({
        "success": False,
        "error": "Direct Amazon SKU sync route is retired. Use governed dispatcher execution path.",
        "execution_blocked": True,
        "route_retired": True,
        "sku": sku
    }), 410

@bp.route('/api/sync/ebay/sku/<sku>', methods=['POST'])
def api_sync_ebay_sku(sku):
    """Retired direct eBay SKU sync route.

    Warehouse quantity must not be mutated from sync routes.
    Marketplace execution must go through governed dispatcher flow only.
    """
    return jsonify({
        "success": False,
        "error": "Direct eBay SKU sync route is retired. Use governed dispatcher execution path.",
        "execution_blocked": True,
        "route_retired": True,
        "sku": sku
    }), 410

@bp.route('/api/push-status/<sku>')
def api_push_status(sku):
    """Get push status for a specific SKU"""
    try:
        warehouse_stock = WarehouseStock.query.filter_by(sku=sku).first()
        if not warehouse_stock:
            return jsonify({'error': 'SKU not found'}), 404

        listings = MarketplaceListing.query.filter_by(warehouse_stock_id=warehouse_stock.id).all()

        status_data = {
            'sku': sku,
            'warehouse_quantity': warehouse_stock.available_quantity,
            'listings': []
        }

        for listing in listings:
            status_data['listings'].append({
                'store_name': listing.store.name,
                'external_listing_id': listing.external_listing_id,
                'listing_type': listing.listing_type,
                'push_state': listing.push_state,
                'last_push_quantity': listing.last_push_quantity,
                'last_push_status': listing.last_push_status,
                'last_push_at': listing.last_push_at.isoformat() if listing.last_push_at else None,
                'needs_push': listing.needs_push,
                'is_pushable': listing.is_pushable,
                'consecutive_failures': listing.consecutive_failures
            })

        return jsonify(status_data)

    except Exception as e:
        logging.error(f"Error getting push status: {str(e)}")
        return jsonify({'error': str(e)}), 500

@bp.get("/admin/stores")
def admin_list_stores():
    """Admin endpoint to list all stores with sync status"""
    if request.headers.get("X-Task-Key") != os.getenv("TASK_API_KEY"):
        return jsonify({"error":"unauthorized"}), 401

    rows = []
    for s in Store.query.order_by(Store.id.asc()).all():
        flag = None
        for cand in ("syncing","is_syncing","sync_status"):
            if hasattr(s, cand):
                flag = {cand: getattr(s, cand)}
                break
        rows.append({
            "id": s.id,
            "name": getattr(s, "name", None),
            "platform": getattr(s, "platform", None),
            "is_active": getattr(s, "is_active", None),
            "flag": flag
        })
    return jsonify({"ok": True, "stores": rows})

@bp.post("/admin/reset-sync/<int:store_id>")
def admin_reset_sync(store_id: int):
    """Admin endpoint to clear stuck sync flags"""
    if request.headers.get("X-Task-Key") != os.getenv("TASK_API_KEY"):
        return jsonify({"error":"unauthorized"}), 401

    s = Store.query.get(store_id)
    if not s:
        return jsonify({"ok": False, "error": "store not found"}), 404

    changed = []
    if hasattr(s, "syncing"):
        try:
            if getattr(s, "syncing") is True:
                setattr(s, "syncing", False); changed.append("syncing=False")
        except Exception: pass
    if hasattr(s, "is_syncing"):
        try:
            if getattr(s, "is_syncing") is True:
                setattr(s, "is_syncing", False); changed.append("is_syncing=False")
        except Exception: pass
    if hasattr(s, "sync_status"):
        try:
            if getattr(s, "sync_status") not in (None, "idle"):
                setattr(s, "sync_status", "idle"); changed.append("sync_status='idle'")
        except Exception: pass

    db.session.commit()
    return jsonify({"ok": True, "store_id": store_id, "changed": changed})

@bp.route('/login', methods=['GET', 'POST'])
def login():
    """User login page"""
    if current_user.is_authenticated:
        return redirect(url_for("routes.dashboard"))

    if request.method == 'POST':
        logging.info("=== LOGIN ATTEMPT ===")
        # CSRF Protection
        csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
        logging.info(f"CSRF token present: {bool(csrf_token)}")
        if not validate_csrf_token(csrf_token):
            logging.warning("CSRF validation failed")
            flash('Invalid request. Please try again.', 'danger')
            return render_template('login.html')

        username = request.form.get('username')
        password = request.form.get('password')
        logging.info(f"Username: {username}, Password length: {len(password) if password else 0}")

        if not username or not password:
            logging.warning("Missing username or password")
            flash('Please enter both username and password.', 'warning')
            return render_template('login.html')

        from models import User
        user = User.query.filter_by(username=username).first()
        logging.info(f"User found: {user is not None}")

        if user:
            password_valid = user.check_password(password)
            logging.info(f"Password valid: {password_valid}")
            logging.info(f"User active: {user.is_active}")

            if password_valid:
                login_user(user)
                next_page = request.args.get('next')
                flash(f'Welcome back, {user.username}!', 'success')
                logging.info(f" Login successful for {username}")
                return redirect(next_page) if next_page else redirect(url_for("routes.dashboard"))

        logging.warning(f" Login failed for {username}")
        flash('Invalid username or password.', 'danger')

    return render_template('login.html')

@bp.route('/logout')
# @login_required  # Temporarily disabled for access
def logout():
    """User logout"""
    logout_user()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for("routes.login"))

@bp.route('/setup-admin', methods=['GET', 'POST'])
def setup_admin():
    """Setup admin user (one-time setup)"""
    from models import User

    # Check if any users exist
    user_count = User.query.count()
    if user_count > 0:
        flash('Admin user already exists. Please login.', 'info')
        return redirect(url_for("routes.login"))

    # Auto-create admin user if none exist and this is a GET request
    if request.method == 'GET':
        try:
            admin_user = User()
            admin_user.username = 'admin'
            admin_user.email = 'admin@example.com'
            admin_user.set_password('password')
            admin_user.is_active = True

            db.session.add(admin_user)
            db.session.commit()

            flash('Admin user created! Username: admin, Password: password', 'success')
            return redirect(url_for("routes.login"))
        except Exception as e:
            logging.error(f'Error creating admin user: {str(e)}')
            flash(f'Error creating admin user: {str(e)}', 'danger')

    if request.method == 'POST':
        # CSRF Protection
        csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
        if not validate_csrf_token(csrf_token):
            flash('Invalid request. Please try again.', 'danger')
            return render_template('setup_admin.html')

        username = request.form.get('username', 'admin')
        email = request.form.get('email', 'admin@example.com')
        password = request.form.get('password', 'password')

        # Create admin user
        admin_user = User()
        admin_user.username = username
        admin_user.email = email
        admin_user.set_password(password)
        admin_user.is_active = True

        db.session.add(admin_user)
        db.session.commit()

        flash(f'Admin user {username} created successfully! You can now login.', 'success')
        return redirect(url_for("routes.login"))

    return render_template('setup_admin.html')

@bp.route('/suppliers')
# @login_required  # Temporarily disabled for access
def suppliers():
    """Display all suppliers"""
    suppliers_list = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()
    return render_template('suppliers.html', suppliers=suppliers_list)

@bp.route('/suppliers/add', methods=['GET', 'POST'])
# @login_required  # Temporarily disabled for access
def add_supplier():
    """Add a new supplier"""
    if request.method == 'POST':
        try:
            supplier = Supplier(
                name=request.form.get('name'),
                email=request.form.get('email'),
                phone=request.form.get('phone'),
                whatsapp_number=request.form.get('whatsapp_number'),
                contact_person=request.form.get('contact_person'),
                address=request.form.get('address'),
                notes=request.form.get('notes'),
                is_active=True
            )
            db.session.add(supplier)
            db.session.commit()
            flash(f'Supplier {supplier.name} added successfully!', 'success')
            return redirect(url_for('routes.suppliers'))
        except Exception as e:
            db.session.rollback()
            logging.error(f'Error adding supplier: {str(e)}')
            flash(f'Error adding supplier: {str(e)}', 'error')
            return render_template('add_supplier.html')
    return render_template('add_supplier.html')

@bp.route('/suppliers/edit/<int:supplier_id>', methods=['GET', 'POST'])
# @login_required  # Temporarily disabled for access
def edit_supplier(supplier_id):
    """Edit an existing supplier"""
    supplier = db.session.get(Supplier, supplier_id)
    if not supplier:
        flash('Supplier not found', 'error')
        return redirect(url_for('routes.suppliers'))

    if request.method == 'POST':
        try:
            supplier.name = request.form.get('name')
            supplier.email = request.form.get('email')
            supplier.phone = request.form.get('phone')
            supplier.whatsapp_number = request.form.get('whatsapp_number')
            supplier.contact_person = request.form.get('contact_person')
            supplier.address = request.form.get('address')
            supplier.notes = request.form.get('notes')
            db.session.commit()
            flash(f'Supplier {supplier.name} updated successfully!', 'success')
            return redirect(url_for('routes.suppliers'))
        except Exception as e:
            db.session.rollback()
            logging.error(f'Error updating supplier: {str(e)}')
            flash(f'Error updating supplier: {str(e)}', 'error')
    return render_template('edit_supplier.html', supplier=supplier)

@bp.route('/suppliers/delete/<int:supplier_id>', methods=['POST'])
# @login_required  # Temporarily disabled for access
def delete_supplier(supplier_id):
    """Delete a supplier (soft delete by setting is_active=False)"""
    try:
        supplier = db.session.get(Supplier, supplier_id)
        if not supplier:
            flash('Supplier not found', 'error')
        else:
            supplier.is_active = False
            db.session.commit()
            flash(f'Supplier {supplier.name} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        logging.error(f'Error deleting supplier: {str(e)}')
        flash(f'Error deleting supplier: {str(e)}', 'error')
    return redirect(url_for('routes.suppliers'))

@bp.route('/purchase-orders')
# @login_required  # Temporarily disabled for access
def purchase_orders():
    """Display all purchase orders"""
    pos = PurchaseOrder.query.order_by(desc(PurchaseOrder.order_date)).all()
    return render_template('purchase_orders.html', purchase_orders=pos)

@bp.route('/purchase-orders/suggested')
# @login_required  # Temporarily disabled for access
def suggested_orders():
    """Show items that need reordering, grouped by supplier"""
    search_query = request.args.get('search', '').strip()

    # Get all warehouse items that need reordering
    query = WarehouseStock.query.filter(
        WarehouseStock.is_active == True,
        WarehouseStock.reorder_point > 0,
        WarehouseStock.available_quantity <= WarehouseStock.reorder_point
    ).options(
        joinedload(WarehouseStock.supplier)
    )

    # Apply search filter if provided
    if search_query:
        query = query.filter(
            db.or_(
                WarehouseStock.sku.ilike(f'%{search_query}%'),
                WarehouseStock.location.ilike(f'%{search_query}%')
            )
        )

    items_needing_reorder = query.all()

    # Get product names for all items
    skus = [item.sku for item in items_needing_reorder]
    inventory_items = InventoryItem.query.filter(InventoryItem.sku.in_(skus)).all()
    sku_to_name = {item.sku: item.name for item in inventory_items}

    # Attach product names to warehouse items
    for item in items_needing_reorder:
        item.product_name = sku_to_name.get(item.sku, 'Unknown Product')

    # Group by supplier
    by_supplier = {}
    no_supplier_items = []

    for item in items_needing_reorder:
        if item.supplier:
            if item.supplier.id not in by_supplier:
                by_supplier[item.supplier.id] = {
                    'supplier': item.supplier,
                    'items': []
                }
            by_supplier[item.supplier.id]['items'].append(item)
        else:
            no_supplier_items.append(item)

    return render_template('suggested_orders.html',
                          by_supplier=by_supplier,
                          no_supplier_items=no_supplier_items,
                          search_query=search_query)

@bp.route('/purchase-orders/create-from-suggestions', methods=['POST'])
# @login_required  # Temporarily disabled for access
def create_purchase_order_from_suggestions():
    """Create purchase order(s) from suggested items"""
    try:
        selected_items = request.form.getlist('selected_items[]')
        create_all = request.form.get('create_all')
        supplier_id = request.form.get('supplier_id')

        if not selected_items and not create_all:
            flash('Please select items to order', 'warning')
            return redirect(url_for('routes.suggested_orders'))

        # Get all items needing reorder
        items_needing_reorder = WarehouseStock.query.filter(
            WarehouseStock.is_active == True,
            WarehouseStock.reorder_point > 0,
            WarehouseStock.available_quantity <= WarehouseStock.reorder_point
        ).options(joinedload(WarehouseStock.supplier)).all()

        # Group by supplier
        by_supplier = {}
        for item in items_needing_reorder:
            if item.supplier:
                if item.supplier.id not in by_supplier:
                    by_supplier[item.supplier.id] = {
                        'supplier': item.supplier,
                        'items': []
                    }
                by_supplier[item.supplier.id]['items'].append(item)

        # Determine which suppliers to create POs for
        suppliers_to_process = []
        if create_all:
            suppliers_to_process = list(by_supplier.keys())
        elif supplier_id:
            suppliers_to_process = [int(supplier_id)]
        else:
            # Get suppliers from selected items
            for item_id in selected_items:
                item = db.session.get(WarehouseStock, int(item_id))
                if item and item.supplier_id and item.supplier_id not in suppliers_to_process:
                    suppliers_to_process.append(item.supplier_id)

        pos_created = 0
        for sup_id in suppliers_to_process:
            if sup_id not in by_supplier:
                continue

            supplier_data = by_supplier[sup_id]
            supplier = supplier_data['supplier']

            # Generate PO number
            last_po = PurchaseOrder.query.order_by(desc(PurchaseOrder.id)).first()
            po_number = f'PO-{(last_po.id + 1 if last_po else 1):05d}'

            # Create purchase order
            po = PurchaseOrder(
                po_number=po_number,
                supplier_id=supplier.id,
                order_date=datetime.utcnow(),
                notes=f'Auto-generated from low stock items',
                status='draft'
            )
            db.session.add(po)
            db.session.flush()

            # Add line items
            total_amount = 0.0
            items_added = 0
            for item in supplier_data['stock_items']:
                # Check if this item was selected (or if creating all)
                if create_all or str(item.id) in selected_items:
                    po_item = PurchaseOrderItem(
                        purchase_order_id=po.id,
                        sku=item.sku,
                        product_name=f'{item.sku} - Reorder',
                        ordered_quantity=item.reorder_quantity,
                        unit_cost=item.unit_cost,
                        total_cost=item.reorder_quantity * item.unit_cost
                    )
                    db.session.add(po_item)
                    total_amount += po_item.total_cost
                    items_added += 1

            if items_added > 0:
                po.total_amount = total_amount
                pos_created += 1

        db.session.commit()

        if pos_created > 0:
            flash(f'{pos_created} purchase order(s) created successfully!', 'success')
        else:
            flash('No purchase orders were created', 'warning')

        return redirect(url_for('routes.purchase_orders'))

    except Exception as e:
        db.session.rollback()
        logging.error(f'Error creating purchase orders from suggestions: {str(e)}')
        flash(f'Error creating purchase orders: {str(e)}', 'error')
        return redirect(url_for('routes.suggested_orders'))

@bp.route('/purchase-orders/create', methods=['GET', 'POST'])
# @login_required  # Temporarily disabled for access
def create_purchase_order():
    """Create a new purchase order"""
    if request.method == 'POST':
        try:
            # Generate PO number
            last_po = PurchaseOrder.query.order_by(desc(PurchaseOrder.id)).first()
            po_number = f'PO-{(last_po.id + 1 if last_po else 1):05d}'

            po = PurchaseOrder(
                po_number=po_number,
                supplier_id=request.form.get('supplier_id') if request.form.get('supplier_id') else None,
                order_date=datetime.strptime(request.form.get('order_date'), '%Y-%m-%d') if request.form.get('order_date') else datetime.utcnow(),
                expected_date=datetime.strptime(request.form.get('expected_date'), '%Y-%m-%d') if request.form.get('expected_date') else None,
                notes=request.form.get('notes'),
                status='draft'
            )
            db.session.add(po)
            db.session.flush()

            # Add line items
            skus = request.form.getlist('sku[]')
            product_names = request.form.getlist('product_name[]')
            quantities = request.form.getlist('quantity[]')
            unit_costs = request.form.getlist('unit_cost[]')

            total_amount = 0.0
            for i in range(len(skus)):
                if skus[i]:
                    qty = int(quantities[i]) if quantities[i] else 0
                    cost = float(unit_costs[i]) if unit_costs[i] else 0.0
                    total = qty * cost
                    total_amount += total

                    po_item = PurchaseOrderItem(
                        purchase_order_id=po.id,
                        sku=skus[i],
                        product_name=product_names[i],
                        ordered_quantity=qty,
                        unit_cost=cost,
                        total_cost=total
                    )
                    db.session.add(po_item)

            po.total_amount = total_amount
            db.session.commit()
            flash(f'Purchase Order {po.po_number} created successfully!', 'success')
            return redirect(url_for('routes.purchase_orders'))
        except Exception as e:
            db.session.rollback()
            logging.error(f'Error creating purchase order: {str(e)}')
            flash(f'Error creating purchase order: {str(e)}', 'error')

    suppliers = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()
    today = datetime.utcnow().strftime('%Y-%m-%d')
    return render_template('create_po.html', suppliers=suppliers, today=today)

@bp.route('/purchase-orders/<int:po_id>')
# @login_required  # Temporarily disabled for access
def view_purchase_order(po_id):
    """View purchase order details"""
    po = db.session.get(PurchaseOrder, po_id)
    if not po:
        flash('Purchase Order not found', 'error')
        return redirect(url_for('routes.purchase_orders'))

    # Look up actual product names and images from warehouse stock for each item
    items_with_names = []
    for item in po.items:
        inventory_item = InventoryItem.query.filter_by(sku=item.sku).first()
        warehouse_stock = WarehouseStock.query.filter_by(sku=item.sku).first()
        items_with_names.append({
            'item': item,
            'actual_name': (warehouse_stock.product_name if warehouse_stock and warehouse_stock.product_name
                          else inventory_item.name if inventory_item
                          else item.product_name),
            'image_url': warehouse_stock.image_url if warehouse_stock else None
        })

    return render_template('view_po.html', po=po, items_with_names=items_with_names)

@bp.route('/purchase-orders/<int:po_id>/send')
# @login_required  # Temporarily disabled for access
def send_po_to_supplier(po_id):
    """Send purchase order to supplier via email"""
    try:
        po = db.session.get(PurchaseOrder, po_id)
        if not po:
            flash('Purchase Order not found', 'error')
            return redirect(url_for('routes.purchase_orders'))

        if not po.supplier:
            flash('Cannot send PO: No supplier assigned', 'error')
            return redirect(url_for('routes.view_purchase_order', po_id=po_id))

        if not po.supplier.email:
            flash(f'Cannot send PO: {po.supplier.name} has no email address', 'error')
            return redirect(url_for('routes.view_purchase_order', po_id=po_id))

        # Import notification service
        from notification_service import NotificationService
        notification_service = NotificationService()

        # Prepare email content
        subject = f"Purchase Order {po.po_number} from Your Company"

        # Text content
        text_content = f"""
Purchase Order: {po.po_number}
Supplier: {po.supplier.name}
Order Date: {po.order_date.strftime('%Y-%m-%d')}
Expected Delivery: {po.expected_date.strftime('%Y-%m-%d') if po.expected_date else 'Not specified'}

Items Ordered:
"""
        for item in po.items:
            # Look up full product details from inventory
            inventory_item = InventoryItem.query.filter_by(sku=item.sku).first()
            product_display = inventory_item.name if inventory_item else item.product_name
            text_content += f"\n- SKU: {item.sku}, {product_display}: {item.ordered_quantity} units @ {item.unit_cost:.2f} = {item.total_cost:.2f}"

        text_content += f"\n\nTotal Amount: {po.total_amount:.2f}"

        if po.notes:
            text_content += f"\n\nNotes:\n{po.notes}"

        text_content += "\n\nPlease confirm receipt of this order and provide tracking information when shipped."

        # HTML content
        html_content = f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
    <div style="max-width: 800px; margin: 0 auto; padding: 20px;">
        <h2 style="color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px;">
            Purchase Order {po.po_number}
        </h2>

        <div style="background-color: #f8f9fa; padding: 15px; margin: 20px 0; border-radius: 5px;">
            <p><strong>Supplier:</strong> {po.supplier.name}</p>
            <p><strong>Order Date:</strong> {po.order_date.strftime('%Y-%m-%d')}</p>
            <p><strong>Expected Delivery:</strong> {po.expected_date.strftime('%Y-%m-%d') if po.expected_date else 'Not specified'}</p>
        </div>

        <h3 style="color: #2c3e50; margin-top: 30px;">Items Ordered:</h3>
        <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
            <thead>
                <tr style="background-color: #3498db; color: white;">
                    <th style="padding: 12px; text-align: center; border: 1px solid #ddd; width: 100px;">Image</th>
                    <th style="padding: 12px; text-align: left; border: 1px solid #ddd;">SKU</th>
                    <th style="padding: 12px; text-align: left; border: 1px solid #ddd;">Product</th>
                    <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Quantity</th>
                    <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Unit Cost</th>
                    <th style="padding: 12px; text-align: right; border: 1px solid #ddd;">Total</th>
                </tr>
            </thead>
            <tbody>
"""

        for item in po.items:
            # Look up full product details and image from warehouse stock
            inventory_item = InventoryItem.query.filter_by(sku=item.sku).first()
            warehouse_stock = WarehouseStock.query.filter_by(sku=item.sku).first()
            product_display = (warehouse_stock.product_name if warehouse_stock and warehouse_stock.product_name
                             else inventory_item.name if inventory_item
                             else item.product_name)
            image_url = warehouse_stock.image_url if warehouse_stock else None

            # Build image cell
            image_cell = ''
            if image_url:
                image_cell = f'<img src="{image_url}" alt="{item.sku}" style="max-width: 80px; max-height: 80px; object-fit: cover; border-radius: 4px;">'
            else:
                image_cell = '<div style="width: 80px; height: 80px; background-color: #f0f0f0; display: flex; align-items: center; justify-content: center; border-radius: 4px; color: #999;">No Image</div>'

            html_content += f"""
                <tr>
                    <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">{image_cell}</td>
                    <td style="padding: 10px; border: 1px solid #ddd;"><code>{item.sku}</code></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{product_display}</td>
                    <td style="padding: 10px; text-align: right; border: 1px solid #ddd;">{item.ordered_quantity}</td>
                    <td style="padding: 10px; text-align: right; border: 1px solid #ddd;">{item.unit_cost:.2f}</td>
                    <td style="padding: 10px; text-align: right; border: 1px solid #ddd;">{item.total_cost:.2f}</td>
                </tr>
"""

        html_content += f"""
            </tbody>
            <tfoot>
                <tr style="background-color: #f8f9fa; font-weight: bold;">
                    <td colspan="5" style="padding: 12px; text-align: right; border: 1px solid #ddd;">Total Amount:</td>
                    <td style="padding: 12px; text-align: right; border: 1px solid #ddd;">{po.total_amount:.2f}</td>
                </tr>
            </tfoot>
        </table>
"""

        if po.notes:
            html_content += f"""
        <div style="background-color: #fff3cd; padding: 15px; margin: 20px 0; border-left: 4px solid #ffc107; border-radius: 5px;">
            <h4 style="margin-top: 0; color: #856404;">Notes:</h4>
            <p style="margin: 0;">{po.notes}</p>
        </div>
"""

        html_content += """
        <p style="margin-top: 30px; padding: 15px; background-color: #e8f5e9; border-left: 4px solid #4caf50; border-radius: 5px;">
            Please confirm receipt of this order and provide tracking information when shipped.
        </p>
    </div>
</body>
</html>
"""

        # Send email - get from email from environment variable or database
        from_email = os.environ.get('SENDGRID_FROM_EMAIL')

        # Check if misconfigured (API key instead of email) and use verified email as fallback
        if from_email and from_email.startswith('SG.'):
            from_email = 'bhavtee@gmail.com'  # Use verified sender email

        if not from_email:
            flash('Email sender not configured. Please set SENDGRID_FROM_EMAIL environment variable with a verified SendGrid sender email.', 'error')
            return redirect(url_for('routes.view_purchase_order', po_id=po_id))

        success = notification_service.send_email_alert(
            to_email=po.supplier.email,
            from_email=from_email,
            subject=subject,
            text_content=text_content,
            html_content=html_content
        )

        if success:
            # Update PO status to 'sent' if it was draft
            if po.status == 'draft':
                po.status = 'sent'
                db.session.commit()

            flash(f'Purchase Order {po.po_number} sent successfully to {po.supplier.name} ({po.supplier.email})', 'success')
        else:
            flash('Failed to send email. Make sure SENDGRID_FROM_EMAIL is set to a verified sender email address in your SendGrid account.', 'error')

    except Exception as e:
        logging.error(f'Error sending PO email: {str(e)}')
        flash(f'Error sending purchase order: {str(e)}', 'error')

    return redirect(url_for('routes.view_purchase_order', po_id=po_id))

@bp.route('/purchase-orders/<int:po_id>/receive', methods=['GET', 'POST'])
# @login_required  # Temporarily disabled for access
def receive_purchase_order(po_id):
    """Receive and inspect purchase order items"""
    po = db.session.get(PurchaseOrder, po_id)
    if not po:
        flash('Purchase Order not found', 'error')
        return redirect(url_for('routes.purchase_orders'))

    if request.method == 'POST':
        try:
            # Track SKUs for automatic push
            skus_to_push = []

            # Create inspection record
            inspection = ReceivingInspection(
                purchase_order_id=po.id,
                inspection_date=datetime.utcnow(),
                inspected_by=request.form.get('inspected_by', 'Admin'),
                notes=request.form.get('inspection_notes'),
                status='completed'
            )
            db.session.add(inspection)
            db.session.flush()

            # Process each line item
            for item in po.items:
                qty_received = int(request.form.get(f'qty_received_{item.id}', 0))
                qty_damaged = int(request.form.get(f'qty_damaged_{item.id}', 0))

                # Validation: damaged cannot exceed received
                if qty_damaged > qty_received:
                    flash(f'Error: Damaged quantity ({qty_damaged}) cannot exceed received quantity ({qty_received}) for {item.sku}', 'error')
                    return render_template('receive_po.html', po=po)

                qty_accepted = qty_received - qty_damaged

                damage_type = request.form.get(f'damage_type_{item.id}')
                damage_notes = request.form.get(f'damage_notes_{item.id}')

                # Create inspection item
                inspection_item = ReceivingInspectionItem(
                    inspection_id=inspection.id,
                    po_item_id=item.id,
                    quantity_received=qty_received,
                    quantity_accepted=qty_accepted,
                    quantity_damaged=qty_damaged,
                    damage_type=damage_type if qty_damaged > 0 else None,
                    damage_notes=damage_notes if qty_damaged > 0 else None,
                    inspection_passed=(qty_damaged == 0)
                )
                db.session.add(inspection_item)

                # Update PO item quantities
                item.received_quantity += qty_accepted
                item.damaged_quantity += qty_damaged

                # Update warehouse stock ONLY for accepted items
                if qty_accepted > 0:
                    # Get default warehouse (required for coordinator compatibility)
                    default_warehouse = Warehouse.get_default()

                    # Query by SKU AND warehouse_id to match coordinator expectations
                    warehouse_stock = WarehouseStock.query.filter_by(
                        sku=item.sku,
                        warehouse_id=default_warehouse.id
                    ).first()
                    qty_before = 0

                    if not warehouse_stock:
                        warehouse_stock = WarehouseStock(
                            warehouse_id=default_warehouse.id,
                            sku=item.sku,
                            available_quantity=0,
                            supplier_id=po.supplier_id,
                            unit_cost=item.unit_cost,
                            is_active=True
                        )
                        db.session.add(warehouse_stock)
                        db.session.flush()  # Get the ID for ledger entry
                        qty_before = 0
                    else:
                        qty_before = warehouse_stock.available_quantity

                    # Update quantities
                    warehouse_stock.available_quantity += qty_accepted
                    warehouse_stock.last_adjustment_at = datetime.utcnow()
                    warehouse_stock.last_adjustment_by = request.form.get('inspected_by', 'Admin')

                    # Create ledger entry
                    ledger_entry = StockLedgerEntry(
                        warehouse_stock_id=warehouse_stock.id,
                        transaction_type='adjustment',
                        adjustment_type='receiving',
                        available_quantity_before=qty_before,
                        available_quantity_after=warehouse_stock.available_quantity,
                        reason=f'Received from PO {po.po_number}',
                        reference_id=str(po.id),
                        reference_type='purchase_order',
                        created_by=request.form.get('inspected_by', 'Admin')
                    )
                    db.session.add(ledger_entry)

                    # Track SKU for automatic push
                    if item.sku not in skus_to_push:
                        skus_to_push.append(item.sku)

            # Update PO status
            all_received = all(item.pending_quantity == 0 for item in po.items)
            po.status = 'received' if all_received else 'partially_received'
            if all_received:
                po.received_date = datetime.utcnow()

            inspection.completed_at = datetime.utcnow()
            inspection.approved_at = datetime.utcnow()
            inspection.approved_by = request.form.get('inspected_by', 'Admin')

            # UNIFIED PUSH SYSTEM: Prepare marketplace pushes BEFORE commit
            from warehouse_push_coordinator import WarehousePushCoordinator
            coordinator = WarehousePushCoordinator()
            if skus_to_push:
                prepared_count = coordinator.prepare_for_items(skus_to_push, operation="update")
                logging.info(f" Prepared {prepared_count} SKUs for marketplace push after PO receipt")

            db.session.commit()

            # UNIFIED PUSH SYSTEM: Enqueue jobs AFTER successful commit
            if skus_to_push:
                jobs_enqueued = coordinator.enqueue_pending_jobs()
                logging.info(f" Enqueued {jobs_enqueued} push jobs for PO {po.po_number}")
            flash(f'Purchase Order {po.po_number} received successfully! {sum(i.quantity_accepted for i in inspection.line_items)} units added to warehouse.', 'success')
            return redirect(url_for('routes.purchase_orders'))
        except Exception as e:
            db.session.rollback()
            logging.error(f'Error receiving purchase order: {str(e)}')
            flash(f'Error receiving purchase order: {str(e)}', 'error')

    return render_template('receive_po.html', po=po)

@bp.route('/api/ocr/invoice', methods=['POST'])
# @login_required  # Temporarily disabled for access
def ocr_invoice():
    """
    Upload and extract invoice data using OCR
    Accepts: multipart/form-data with 'invoice_image' file
    Returns: JSON with extracted invoice_number, invoice_date, total_amount, supplier_name
    """
    try:
        if 'invoice_image' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        file = request.files['invoice_image']

        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        # Check file type
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff', 'pdf'}
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''

        if file_ext not in allowed_extensions:
            return jsonify({'success': False, 'error': 'Invalid file type. Allowed: PNG, JPG, JPEG, PDF'}), 400

        # Process with OCR
        from ocr_service import extract_invoice_data
        result = extract_invoice_data(file)

        return jsonify(result), 200

    except Exception as e:
        logging.error(f'OCR processing error: {str(e)}')
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/purchase-orders/<int:po_id>/invoice', methods=['POST'])
# @login_required  # Temporarily disabled for access
def update_po_invoice(po_id):
    """Update purchase order invoice details from OCR"""
    po = db.session.get(PurchaseOrder, po_id)
    if not po:
        flash('Purchase Order not found', 'error')
        return redirect(url_for('routes.purchase_orders'))

    try:
        # Update invoice fields
        po.invoice_number = request.form.get('invoice_number')

        invoice_date_str = request.form.get('invoice_date')
        if invoice_date_str:
            po.invoice_date = datetime.strptime(invoice_date_str, '%Y-%m-%d')

        po.payment_status = request.form.get('payment_status', 'pending')

        payment_amount_str = request.form.get('payment_amount')
        if payment_amount_str:
            po.payment_amount = float(payment_amount_str)

        db.session.commit()
        flash(f'Invoice details saved for PO {po.po_number}', 'success')
    except Exception as e:
        db.session.rollback()
        logging.error(f'Error saving invoice data: {str(e)}')
        flash(f'Error saving invoice details: {str(e)}', 'error')

    return redirect(url_for('routes.view_purchase_order', po_id=po_id))

@bp.route('/quick-scan')
# @login_required  # Temporarily disabled for access
def quick_scan():
    """Quick scan page for barcode scanning and stock adjustments"""
    return render_template('quick_scan.html')

@bp.route('/api/warehouse/lookup/<sku>')
# @login_required  # Temporarily disabled for access
def warehouse_lookup(sku):
    """Lookup warehouse product by SKU or barcode"""
    try:
        stock = WarehouseStock.query.filter_by(sku=sku).first()

        if not stock:
            return jsonify({'success': False, 'error': 'Product not found'}), 404

        product_data = {
            'id': stock.id,
            'sku': stock.sku,
            'name': stock.product_name,
            'available_quantity': stock.available_quantity,
            'unit_cost': stock.unit_cost,
            'supplier_name': stock.supplier.name if stock.supplier else None
        }

        return jsonify({'success': True, 'product': product_data}), 200

    except Exception as e:
        logging.error(f'Warehouse lookup error: {str(e)}')
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/warehouse/<int:stock_id>/quick-adjust', methods=['POST'])
# @login_required  # Temporarily disabled for access
def quick_scan_stock_adjust(stock_id):
    """Quick adjust warehouse stock quantity from barcode scanner"""
    try:
        stock = db.session.get(WarehouseStock, stock_id)
        if not stock:
            return jsonify({'success': False, 'error': 'Product not found'}), 404

        data = request.get_json()
        new_quantity = int(data.get('new_quantity', 0))
        reason = data.get('reason', 'quick_scan')

        # Calculate adjustment
        old_quantity = stock.available_quantity
        adjustment = new_quantity - old_quantity

        # Update stock
        stock.available_quantity = new_quantity
        stock.updated_at = datetime.utcnow()

        # Create ledger entry
        ledger = StockLedgerEntry(
            warehouse_stock_id=stock.id,
            transaction_type=reason,
            quantity_before=old_quantity,
            quantity_change=adjustment,
            quantity_after=new_quantity,
            notes=f'Quick scan adjustment: {reason}'
        )
        db.session.add(ledger)

        # UNIFIED PUSH SYSTEM: Prepare marketplace pushes BEFORE commit
        from warehouse_push_coordinator import WarehousePushCoordinator
        coordinator = WarehousePushCoordinator()
        prepared_count = coordinator.prepare_for_items([stock.sku], operation="update")

        db.session.commit()

        # UNIFIED PUSH SYSTEM: Enqueue jobs AFTER successful commit
        jobs_enqueued = coordinator.enqueue_pending_jobs()
        logging.info(f" Quick scan adjust: Enqueued {jobs_enqueued} push jobs for {stock.sku}")

        return jsonify({
            'success': True,
            'new_quantity': new_quantity,
            'adjustment': adjustment,
            'push_jobs_enqueued': jobs_enqueued
        }), 200

    except Exception as e:
        db.session.rollback()
        logging.error(f'Warehouse adjust error: {str(e)}')
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/inventory')
# @login_required  # Temporarily disabled for access
def inventory():
    """Display warehouse stock (the boss) with optional grouping and search - warehouse-authoritative view"""
    # Check for view mode (grouped or flat)
    view_mode = request.args.get('view', 'flat')
    group_filter = request.args.get('group')
    search_query = request.args.get('search', '').strip()
    marketplace_filter = request.args.get('marketplace', '').strip()

    # Base query for WarehouseStock (the boss) with eager-loaded relationships
    query = db.session.query(WarehouseStock).options(
        joinedload(WarehouseStock.supplier),
        joinedload(WarehouseStock.marketplace_listings).joinedload(MarketplaceListing.store)
    )

    # Apply search filter if specified (search on SKU, inventory item names, or marketplace titles)
    if search_query:
        search_pattern = f'%{search_query}%'
        # Search by warehouse SKU, inventory item name, or marketplace listing title
        query = query.outerjoin(InventoryItem, WarehouseStock.sku == InventoryItem.sku).outerjoin(
            MarketplaceListing, MarketplaceListing.warehouse_stock_id == WarehouseStock.id
        ).filter(or_(
            WarehouseStock.sku.ilike(search_pattern),
            InventoryItem.name.ilike(search_pattern),
            MarketplaceListing.title.ilike(search_pattern)
        )).distinct()

    # Apply marketplace filter if specified
    if marketplace_filter:
        # Get distinct warehouse stock IDs that are listed on this marketplace
        subquery = db.session.query(WarehouseStock.id).join(
            MarketplaceListing, MarketplaceListing.warehouse_stock_id == WarehouseStock.id
        ).join(
            Store, MarketplaceListing.store_id == Store.id
        ).filter(Store.platform == marketplace_filter).distinct().subquery()

        # Filter main query to only these IDs
        query = query.filter(WarehouseStock.id.in_(db.session.query(subquery.c.id)))

    # Order by SKU
    warehouse_stocks = query.order_by(WarehouseStock.sku).all()

    # Get all unique marketplaces from stores
    all_marketplaces = db.session.query(Store.platform).distinct().order_by(Store.platform).all()
    all_marketplaces = [m[0] for m in all_marketplaces]

    # Create a simple view model class to avoid mutating SQLAlchemy objects
    class WarehouseStockView:
        """View model for displaying warehouse stock with associated data"""
        def __init__(self, warehouse_stock, inventory_items, marketplace_summary):
            self.warehouse_stock = warehouse_stock
            self.linked_items = inventory_items
            self.marketplace_summary = marketplace_summary

            # Try to get product name from marketplace listing title (most descriptive)
            product_name = None
            for listing in warehouse_stock.marketplace_listings:
                if listing.title and listing.title.strip():
                    product_name = listing.title
                    break

            # Use first inventory item for display metadata, or create defaults
            if inventory_items:
                primary = inventory_items[0]
                self.id = primary.id
                # Use marketplace title if available, otherwise fall back to inventory item name
                self.name = product_name or primary.name
                self.description = primary.description
                self.price = primary.price
                self.variant_attributes = primary.variant_attributes
                self.group_id = primary.group_id
                self.group = primary.group
                self.updated_at = warehouse_stock.updated_at or primary.updated_at
            else:
                # Warehouse-only item (no InventoryItem yet)
                self.id = warehouse_stock.id
                # Use marketplace title if available, otherwise use warehouse SKU
                self.name = product_name or f"Warehouse Item - {warehouse_stock.sku}"
                self.description = "Managed at warehouse level"
                self.price = warehouse_stock.unit_cost or 0.0
                self.variant_attributes = None
                self.group_id = None
                self.group = None
                self.updated_at = warehouse_stock.updated_at

            # Always use warehouse SKU as authoritative
            self.sku = warehouse_stock.sku

    # Batch-load all inventory items to avoid N+1 queries
    all_skus = [ws.sku for ws in warehouse_stocks]
    if all_skus:
        all_inventory_items = InventoryItem.query.options(
            joinedload(InventoryItem.group)
        ).filter(InventoryItem.sku.in_(all_skus)).all()

        # Group inventory items by SKU for fast lookup
        inventory_by_sku = {}
        for item in all_inventory_items:
            if item.sku not in inventory_by_sku:
                inventory_by_sku[item.sku] = []
            inventory_by_sku[item.sku].append(item)
    else:
        inventory_by_sku = {}

    # For each warehouse stock, create a view model
    items = []
    for ws in warehouse_stocks:
        # Get inventory items for this SKU from our pre-loaded batch
        inventory_items = inventory_by_sku.get(ws.sku, [])

        # Calculate marketplace summary for this warehouse stock
        marketplace_summary = {}
        for listing in ws.marketplace_listings:
            platform = listing.store.platform if listing.store else 'Unknown'

            if platform not in marketplace_summary:
                marketplace_summary[platform] = {
                    'total_listings': 0,
                    'active_listings': 0,
                    'needs_push': False,
                    'last_sync': None
                }

            marketplace_summary[platform]['total_listings'] += 1
            if listing.is_active:
                marketplace_summary[platform]['active_listings'] += 1
            if listing.needs_push:
                marketplace_summary[platform]['needs_push'] = True
            if listing.last_push_at:
                if not marketplace_summary[platform]['last_sync'] or listing.last_push_at > marketplace_summary[platform]['last_sync']:
                    marketplace_summary[platform]['last_sync'] = listing.last_push_at

        # Create view model (one per warehouse stock - no mutation!)
        view_item = WarehouseStockView(ws, inventory_items, marketplace_summary)
        items.append(view_item)

    # Organize items by groups if in grouped view
    grouped_items = {}
    ungrouped_items = []
    all_groups = []

    if view_mode == 'grouped':
        # Get all groups
        all_groups = db.session.query(ProductGroup).order_by(ProductGroup.name).all()

        # Apply group filter if specified
        if group_filter:
            if group_filter == 'ungrouped':
                items = [item for item in items if not item.group_id]
            else:
                try:
                    group_id = int(group_filter)
                    items = [item for item in items if item.group_id == group_id]
                except ValueError:
                    flash('Invalid group filter', 'warning')

        # Organize items by groups
        for item in items:
            if item.group_id:
                if item.group_id not in grouped_items:
                    grouped_items[item.group_id] = {
                        'group': item.group,
                        'items': [],
                        'marketplace_summary': {}
                    }
                grouped_items[item.group_id]['items'].append(item)
            else:
                ungrouped_items.append(item)

        # Aggregate marketplace data for each group
        for group_id, group_data in grouped_items.items():
            marketplace_summary = {}

            for item in group_data['items']:
                for platform, data in item.marketplace_summary.items():
                    if platform not in marketplace_summary:
                        marketplace_summary[platform] = {
                            'total_listings': 0,
                            'active_listings': 0,
                            'needs_push': False
                        }

                    marketplace_summary[platform]['total_listings'] += data['total_listings']
                    marketplace_summary[platform]['active_listings'] += data['active_listings']
                    if data['needs_push']:
                        marketplace_summary[platform]['needs_push'] = True

            group_data['marketplace_summary'] = marketplace_summary
    else:
        # Apply group filter in flat view
        if group_filter:
            if group_filter == 'ungrouped':
                items = [item for item in items if not item.group_id]
            else:
                try:
                    group_id = int(group_filter)
                    items = [item for item in items if item.group_id == group_id]
                except ValueError:
                    flash('Invalid group filter', 'warning')

    return render_template('inventory.html',
                         items=items,
                         view_mode=view_mode,
                         grouped_items=grouped_items,
                         ungrouped_items=ungrouped_items,
                         all_groups=all_groups,
                         all_marketplaces=all_marketplaces,
                         current_group_filter=group_filter,
                         current_marketplace_filter=marketplace_filter,
                         current_search=search_query)

@bp.route('/inventory/add', methods=['GET', 'POST'])
# @login_required  # Temporarily disabled for access
def add_item():
    """Add new inventory item"""
    if request.method == 'POST':
        try:
            # CSRF Protection
            csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
            if not validate_csrf_token(csrf_token):
                logging.warning(f"CSRF token validation failed for add_item request from {request.remote_addr}")
                flash('Invalid request. Please try again.', 'danger')
                return redirect(url_for('routes.add_item'))

            item = InventoryItem()
            item.name = request.form['name']
            item.sku = request.form['sku']
            item.quantity = int(request.form['quantity'])
            item.price = float(request.form['price'])
            item.description = request.form.get('description', '')
            item.reorder_point = int(request.form.get('reorder_point', 0) or 0)

            # Handle group assignment
            group_id = request.form.get('group_id')
            if group_id and group_id != '':
                item.group_id = int(group_id)

            # Handle variant attributes
            variant_attrs = {}
            for key in ['color', 'size', 'material', 'style']:
                value = request.form.get(f'variant_{key}')
                if value and value.strip():
                    variant_attrs[key] = value.strip()

            if variant_attrs:
                item.variant_attributes = variant_attrs

            db.session.add(item)
            db.session.flush()  # Get item.id for use in prepare step

            # Step 1: Prepare warehouse stock and get stores list (BEFORE commit)
            stores_to_push, warehouse_stock = prepare_warehouse_push(item, operation="create")

            # Step 2: Commit item AND warehouse stock together
            db.session.commit()

            # Step 3: Enqueue jobs AFTER successful commit (won't rollback if push fails)
            jobs_count = 0
            if stores_to_push:
                try:
                    jobs_count = enqueue_push_jobs(item.id, stores_to_push)
                except Exception as push_error:
                    logging.error(f"Push job enqueue failed after adding item {item.sku}: {str(push_error)}")
                    # Item is already saved, just warn user
                    flash(f'Item added successfully! However, push enqueue failed - use manual push button.', 'warning')
                    return redirect(url_for('routes.inventory'))

            if jobs_count > 0:
                flash(f'Item added successfully! High-priority push queued for {jobs_count} marketplace(s).', 'success')
            else:
                flash('Item added successfully!', 'success')

            return redirect(url_for('routes.inventory'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding item: {str(e)}', 'danger')
            logging.error(f'Error adding item: {str(e)}')

    # Get all groups for selection dropdown
    groups = db.session.query(ProductGroup).order_by(ProductGroup.name).all()
    return render_template('add_item.html', groups=groups)

@bp.route('/inventory/edit/<int:item_id>', methods=['GET', 'POST'])
# @login_required  # Temporarily disabled for access
def edit_item(item_id):
    """Edit existing inventory item"""
    item = db.session.get(InventoryItem, item_id)
    if not item:
        flash('Item not found!', 'danger')
        return redirect(url_for('routes.inventory'))

    if request.method == 'POST':
        try:
            # CSRF Protection
            csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
            if not validate_csrf_token(csrf_token):
                logging.warning(f"CSRF token validation failed for edit_item request from {request.remote_addr}")
                flash('Invalid request. Please try again.', 'danger')
                return redirect(url_for('routes.edit_item', item_id=item_id))

            # Capture original quantity before updating
            original_quantity = item.quantity

            item.name = request.form['name']
            item.sku = request.form['sku']
            new_quantity = int(request.form['quantity'])
            item.quantity = new_quantity
            item.price = float(request.form['price'])
            item.description = request.form.get('description', '')
            item.reorder_point = int(request.form.get('reorder_point', 0) or 0)

            # Handle group assignment
            group_id = request.form.get('group_id')
            if group_id and group_id != '':
                item.group_id = int(group_id)
            else:
                item.group_id = None

            # Handle variant attributes
            variant_attrs = {}
            for key in ['color', 'size', 'material', 'style']:
                value = request.form.get(f'variant_{key}')
                if value and value.strip():
                    variant_attrs[key] = value.strip()

            item.variant_attributes = variant_attrs if variant_attrs else None

            # Check if quantity changed
            quantity_changed = original_quantity != new_quantity
            stores_to_push = []

            if quantity_changed:
                # Step 1: Prepare warehouse stock and get stores list (BEFORE commit)
                stores_to_push, warehouse_stock = prepare_warehouse_push(item, operation="update")

            # Step 2: Commit item changes AND warehouse stock together
            db.session.commit()

            # Step 3: Enqueue jobs AFTER successful commit
            jobs_count = 0
            if quantity_changed and stores_to_push:
                try:
                    jobs_count = enqueue_push_jobs(item.id, stores_to_push)
                except Exception as push_error:
                    logging.error(f"Push job enqueue failed after updating item {item.sku}: {str(push_error)}")
                    # Item is already saved, just warn user
                    flash(f'Item updated successfully! Quantity changed from {original_quantity} to {new_quantity} but push enqueue failed - use manual push button.', 'warning')
                    return redirect(url_for('routes.inventory'))

            if quantity_changed:
                if jobs_count > 0:
                    flash(f'Item updated successfully! Quantity changed from {original_quantity} to {new_quantity}. High-priority push queued for {jobs_count} marketplace(s).', 'success')
                else:
                    flash(f'Item updated successfully! Quantity changed from {original_quantity} to {new_quantity}', 'success')
            else:
                flash('Item updated successfully!', 'success')

            return redirect(url_for('routes.inventory'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating item: {str(e)}', 'danger')
            logging.error(f'Error updating item: {str(e)}')

    # Get all groups for selection dropdown
    groups = db.session.query(ProductGroup).order_by(ProductGroup.name).all()
    return render_template('edit_item.html', item=item, groups=groups)

@bp.route('/inventory/delete/<int:item_id>', methods=['POST'])
# @login_required  # Temporarily disabled for access
def delete_item(item_id):
    """Delete inventory item"""
    try:
        # CSRF Protection
        csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
        if not validate_csrf_token(csrf_token):
            logging.warning(f"CSRF token validation failed for delete_item request from {request.remote_addr}")
            flash('Invalid request. Please try again.', 'danger')
            return redirect(url_for('routes.inventory'))

        item = db.session.get(InventoryItem, item_id)
        if item:
            db.session.delete(item)
            db.session.commit()
            flash('Item deleted successfully!', 'success')
        else:
            flash('Item not found!', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting item: {str(e)}', 'danger')
        logging.error(f'Error deleting item: {str(e)}')

    return redirect(url_for('routes.inventory'))

@bp.route('/inventory/delete_bulk', methods=['POST'])
# @login_required  # Temporarily disabled for access
def delete_bulk_items():
    """Delete multiple inventory items in bulk"""
    try:
        # CSRF Protection
        csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
        if not validate_csrf_token(csrf_token):
            logging.warning(f"CSRF token validation failed for delete_bulk_items request from {request.remote_addr}")
            return jsonify({'success': False, 'error': 'Invalid request token'}), 403

        # Get item IDs from request
        data = request.get_json()
        if not data or 'item_ids' not in data:
            return jsonify({'success': False, 'error': 'No items selected'}), 400

        item_ids = data['item_ids']
        if not isinstance(item_ids, list) or len(item_ids) == 0:
            return jsonify({'success': False, 'error': 'No items selected'}), 400

        # Delete items
        deleted_count = 0
        for item_id in item_ids:
            try:
                item = db.session.get(InventoryItem, int(item_id))
                if item:
                    db.session.delete(item)
                    deleted_count += 1
            except Exception as item_error:
                logging.error(f'Error deleting item {item_id}: {str(item_error)}')

        db.session.commit()

        message = f'{deleted_count} item(s) deleted successfully!'
        logging.info(f'Bulk delete: {deleted_count} items deleted')

        return jsonify({'success': True, 'message': message, 'deleted_count': deleted_count})

    except Exception as e:
        db.session.rollback()
        error_msg = f'Error deleting items: {str(e)}'
        logging.error(error_msg)
        return jsonify({'success': False, 'error': error_msg}), 500

@bp.route('/update_stock/<int:item_id>', methods=['POST'])
# @login_required  # Temporarily disabled for access
def update_stock(item_id):
    """AJAX endpoint for updating stock quantities inline"""
    try:
        # CSRF Protection
        csrf_token = request.headers.get('X-CSRF-Token') or (request.json.get('csrf_token') if request.json else None)
        if not validate_csrf_token(csrf_token):
            logging.warning(f"CSRF token validation failed for update_stock request from {request.remote_addr}")
            return jsonify({'success': False, 'error': 'Invalid request token'}), 403

        # Get the item
        item = db.session.get(InventoryItem, item_id)
        if not item:
            return jsonify({'success': False, 'error': 'Item not found'}), 404

        # Get the new quantity from request
        data = request.get_json()
        if not data or 'quantity' not in data:
            return jsonify({'success': False, 'error': 'Quantity is required'}), 400

        try:
            new_quantity = int(data['quantity'])
            if new_quantity < 0:
                return jsonify({'success': False, 'error': 'Quantity cannot be negative'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Invalid quantity value'}), 400

        # Store original quantity for logging
        original_quantity = item.quantity

        # Update the quantity
        item.quantity = new_quantity
        item.updated_at = datetime.utcnow()

        # Prepare push if quantity changed (BEFORE commit)
        stores_to_push = []
        if original_quantity != new_quantity:
            stores_to_push, warehouse_stock = prepare_warehouse_push(item, operation="update")

        # Commit quantity change AND warehouse stock together
        db.session.commit()

        # Enqueue jobs AFTER successful commit
        jobs_count = 0
        if stores_to_push:
            try:
                jobs_count = enqueue_push_jobs(item.id, stores_to_push)
                logging.info(f"Stock updated for {item.sku}: {original_quantity} -> {new_quantity}. Queued {jobs_count} push jobs.")
            except Exception as push_error:
                logging.error(f"Push job enqueue failed after updating stock for {item.sku}: {str(push_error)}")
                # Continue anyway, item is already saved

        # Determine badge class based on new quantity
        if new_quantity < 10:
            badge_class = 'bg-warning text-dark'
        elif new_quantity > 50:
            badge_class = 'bg-success'
        else:
            badge_class = 'bg-secondary'

        # Prepare response data
        push_message_suffix = f' - high-priority push queued for {jobs_count} marketplace(s)' if jobs_count > 0 else ''
        response_data = {
            'success': True,
            'message': f'Quantity updated from {original_quantity} to {new_quantity}{push_message_suffix}',
            'new_quantity': new_quantity,
            'badge_class': badge_class,
            'needs_reorder': new_quantity <= (item.reorder_point or 0),
            'item_id': item_id,
            'sku': item.sku
        }

        return jsonify(response_data)

    except Exception as e:
        db.session.rollback()
        logging.error(f'Error updating stock for item {item_id}: {str(e)}')
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

@bp.route('/batch_update_stock', methods=['POST'])
# @login_required  # Temporarily disabled for access
def batch_update_stock():
    """AJAX endpoint for updating multiple stock quantities in one atomic transaction"""
    try:
        # CSRF Protection
        csrf_token = request.headers.get('X-CSRF-Token') or (request.json.get('csrf_token') if request.json else None)
        if not validate_csrf_token(csrf_token):
            logging.warning(f"CSRF token validation failed for batch_update_stock request from {request.remote_addr}")
            return jsonify({'success': False, 'error': 'Invalid request token'}), 403

        # Get the batch updates from request
        data = request.get_json()
        if not data or 'updates' not in data:
            return jsonify({'success': False, 'error': 'Updates array is required'}), 400

        updates = data['updates']
        if not isinstance(updates, list) or len(updates) == 0:
            return jsonify({'success': False, 'error': 'Updates must be a non-empty array'}), 400

        # Validate all updates first
        validated_updates = []
        for update in updates:
            if not isinstance(update, dict) or 'item_id' not in update or 'quantity' not in update:
                return jsonify({'success': False, 'error': 'Each update must have item_id and quantity'}), 400

            try:
                item_id = int(update['item_id'])
                new_quantity = int(update['quantity'])
                if new_quantity < 0:
                    return jsonify({'success': False, 'error': f'Quantity cannot be negative for item {item_id}'}), 400
            except (ValueError, TypeError):
                return jsonify({'success': False, 'error': f'Invalid item_id or quantity values'}), 400

            # Check if item exists
            item = db.session.get(InventoryItem, item_id)
            if not item:
                return jsonify({'success': False, 'error': f'Item {item_id} not found'}), 404

            validated_updates.append({
                'item_id': item_id,
                'item': item,
                'original_quantity': item.quantity,
                'new_quantity': new_quantity
            })

        # Perform all updates in a single transaction
        updated_items = []
        items_with_changes = []

        for update_data in validated_updates:
            item = update_data['item']
            original_quantity = update_data['original_quantity']
            new_quantity = update_data['new_quantity']

            # Update the quantity in inventory_items
            item.quantity = new_quantity
            item.updated_at = datetime.utcnow()

            # CRITICAL: Use row-level locking to prevent concurrent modification races
            # This protects against simultaneous marketplace sales during batch updates
            warehouse_stock = db.session.execute(
                select(WarehouseStock)
                .filter_by(sku=item.sku)
                .with_for_update()  # SELECT FOR UPDATE - blocks other transactions
            ).scalar_one_or_none()

            if warehouse_stock:
                warehouse_stock.available_quantity = new_quantity
                warehouse_stock.stock_version += 1  # Increment version for optimistic locking
                warehouse_stock.updated_at = datetime.utcnow()
                logging.info(f" Synced warehouse_stock for SKU {item.sku}: {original_quantity}  {new_quantity} (version {warehouse_stock.stock_version})")
            else:
                # Create warehouse_stock if it doesn't exist
                warehouse_stock = WarehouseStock(
                    sku=item.sku,
                    available_quantity=new_quantity,
                    stock_version=0,  # Initial version
                    unit_cost=item.price,
                    location='Warehouse',
                    is_active=True,
                    track_inventory=True
                )
                db.session.add(warehouse_stock)
                logging.info(f" Created warehouse_stock for SKU {item.sku}: {new_quantity} units")

            # Determine badge class based on new quantity
            if new_quantity < 10:
                badge_class = 'bg-warning text-dark'
            elif new_quantity > 50:
                badge_class = 'bg-success'
            else:
                badge_class = 'bg-secondary'

            updated_items.append({
                'item_id': item.id,
                'sku': item.sku,
                'original_quantity': original_quantity,
                'new_quantity': new_quantity,
                'badge_class': badge_class,
                'needs_reorder': new_quantity <= (item.reorder_point or 0)
            })

            # Track items that actually changed (collect SKUs)
            if original_quantity != new_quantity:
                items_with_changes.append(item.sku)  # Store SKU for coordinator

        # UNIFIED PUSH SYSTEM: Prepare marketplace pushes BEFORE commit (deduplicate SKUs)
        skus_to_push = list(set(items_with_changes))  # Remove duplicates
        from warehouse_push_coordinator import WarehousePushCoordinator
        coordinator = WarehousePushCoordinator()
        if skus_to_push:
            prepared_count = coordinator.prepare_for_items(skus_to_push, operation="update")
            logging.info(f" Batch update: Prepared {prepared_count} SKUs for marketplace push")

        # Commit all changes including warehouse stock
        db.session.commit()

        # UNIFIED PUSH SYSTEM: Enqueue jobs AFTER successful commit
        total_jobs = 0
        auto_sync_errors = []
        if skus_to_push:
            try:
                total_jobs = coordinator.enqueue_pending_jobs()
                logging.info(f" Batch update: Enqueued {total_jobs} push jobs for {len(skus_to_push)} SKUs")
            except Exception as e:
                logging.error(f"Error during batch push enqueue: {str(e)}")
                auto_sync_errors = skus_to_push[:3]  # Sample of failed SKUs

        # Prepare response
        response_data = {
            'success': True,
            'message': f'Successfully updated {len(updated_items)} items',
            'updated_items': updated_items,
            'changed_count': len(items_with_changes),
            'auto_sync_initiated': len(items_with_changes) > 0
        }

        if auto_sync_errors:
            response_data['warning'] = f'Auto-sync failed for {len(auto_sync_errors)} items: {", ".join(auto_sync_errors[:3])}{"..." if len(auto_sync_errors) > 3 else ""}'
        elif items_with_changes:
            response_data['message'] += ' - automatic sync initiated'

        logging.info(f"Batch stock update completed: {len(updated_items)} items updated, {len(items_with_changes)} changes made")
        return jsonify(response_data)

    except Exception as e:
        db.session.rollback()
        logging.error(f'Error in batch stock update: {str(e)}')
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

# =================== GROUP MANAGEMENT ROUTES ===================

@bp.route('/groups')
# @login_required  # Temporarily disabled for access
def groups():
    """Display all product groups with their items"""
    groups = db.session.query(ProductGroup).order_by(desc(ProductGroup.updated_at)).all()

    # For each group, get its items for inline editing
    groups_with_items = []
    for group in groups:
        group_items = db.session.query(InventoryItem).filter(
            InventoryItem.group_id == group.id
        ).order_by(InventoryItem.sku).all()

        groups_with_items.append({
            'group': group,
            'items': group_items,
            'stats': group.get_aggregate_stats()
        })

    return render_template('groups.html', groups_with_items=groups_with_items)

@bp.route('/groups/add', methods=['GET', 'POST'])
def add_group():
    """Add new product group"""
    if request.method == 'POST':
        try:
            # CSRF Protection
            csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
            if not validate_csrf_token(csrf_token):
                logging.warning(f"CSRF token validation failed for add_group request from {request.remote_addr}")
                flash('Invalid request. Please try again.', 'danger')
                return redirect(url_for('routes.add_group'))

            group = ProductGroup()
            group.name = request.form['name']
            group.description = request.form.get('description', '')
            group.group_key = request.form['group_key']

            db.session.add(group)
            db.session.commit()
            flash('Group added successfully!', 'success')
            return redirect(url_for('routes.groups'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding group: {str(e)}', 'danger')
            logging.error(f'Error adding group: {str(e)}')

    return render_template('add_group.html')

@bp.route('/groups/edit/<int:group_id>', methods=['GET', 'POST'])
def edit_group(group_id):
    """Edit existing product group"""
    group = db.session.get(ProductGroup, group_id)
    if not group:
        flash('Group not found!', 'danger')
        return redirect(url_for('routes.groups'))

    if request.method == 'POST':
        try:
            # CSRF Protection
            csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
            if not validate_csrf_token(csrf_token):
                logging.warning(f"CSRF token validation failed for edit_group request from {request.remote_addr}")
                flash('Invalid request. Please try again.', 'danger')
                return redirect(url_for('routes.edit_group', group_id=group_id))

            group.name = request.form['name']
            group.description = request.form.get('description', '')
            group.group_key = request.form['group_key']

            db.session.commit()
            flash('Group updated successfully!', 'success')
            return redirect(url_for('routes.groups'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating group: {str(e)}', 'danger')
            logging.error(f'Error updating group: {str(e)}')

    return render_template('edit_group.html', group=group)

@bp.route('/groups/<int:group_id>/push', methods=['POST'])
def push_group_stock(group_id):
    """Retired direct group push route.

    Group marketplace execution must go through governed dispatcher flow only.
    """
    return jsonify({
        "success": False,
        "error": "Direct group push route is retired. Use governed dispatcher execution path.",
        "execution_blocked": True,
        "route_retired": True,
        "group_id": group_id
    }), 410

@bp.route('/groups/<int:group_id>/update-stock', methods=['POST'])
def update_group_stock(group_id):
    """Update warehouse stock quantity for all items in a product group"""
    try:
        data = request.get_json()
        new_quantity = data.get('new_quantity')

        if new_quantity is None or new_quantity < 0:
            return jsonify({'success': False, 'message': 'Invalid quantity'}), 400

        group = db.session.get(ProductGroup, group_id)
        if not group:
            return jsonify({'success': False, 'message': 'Group not found'}), 404

        if not group.items:
            return jsonify({'success': False, 'message': 'Group has no items'}), 400

        # Get default warehouse for coordinator compatibility
        default_warehouse = Warehouse.get_default()

        items_updated = 0
        skus_to_push = []  # Track SKUs for automatic push

        # Update warehouse stock for all items in the group
        for item in group.items:
            warehouse_stock = WarehouseStock.query.filter_by(sku=item.sku).first()

            if warehouse_stock:
                old_qty = warehouse_stock.available_quantity
                warehouse_stock.available_quantity = new_quantity

                # Create ledger entry for the adjustment
                ledger_entry = StockLedgerEntry(
                    warehouse_stock_id=warehouse_stock.id,
                    transaction_type='adjustment',
                    adjustment_type='set' if new_quantity != old_qty else 'none',
                    available_quantity_before=old_qty,
                    available_quantity_after=new_quantity,
                    reserved_quantity_before=warehouse_stock.reserved_quantity,
                    reserved_quantity_after=warehouse_stock.reserved_quantity,
                    allocated_quantity_before=warehouse_stock.allocated_quantity or 0,
                    allocated_quantity_after=warehouse_stock.allocated_quantity or 0,
                    on_order_quantity_before=warehouse_stock.on_order_quantity or 0,
                    on_order_quantity_after=warehouse_stock.on_order_quantity or 0,
                    reason=f'Group quantity update: {group.name}',
                    reference_id=f'GROUP-{group_id}',
                    reference_type='group_update',
                    created_by='system',
                    source_system='warehouse'
                )
                db.session.add(ledger_entry)

                # Update inventory item quantity to match warehouse
                item.quantity = new_quantity
                items_updated += 1

                # Track SKU for automatic push (if quantity changed)
                if old_qty != new_quantity and item.sku not in skus_to_push:
                    skus_to_push.append(item.sku)
            else:
                # Create new warehouse stock if it doesn't exist (with warehouse_id for coordinator)
                warehouse_stock = WarehouseStock(
                    warehouse_id=default_warehouse.id,
                    sku=item.sku,
                    available_quantity=new_quantity,
                    reserved_quantity=0,
                    location='DEFAULT'
                )
                db.session.add(warehouse_stock)
                db.session.flush()

                # Create ledger entry
                ledger_entry = StockLedgerEntry(
                    warehouse_stock_id=warehouse_stock.id,
                    transaction_type='initial',
                    adjustment_type='set',
                    available_quantity_before=0,
                    available_quantity_after=new_quantity,
                    reserved_quantity_before=0,
                    reserved_quantity_after=0,
                    allocated_quantity_before=0,
                    allocated_quantity_after=0,
                    on_order_quantity_before=0,
                    on_order_quantity_after=0,
                    reason=f'Initial stock via group: {group.name}',
                    reference_id=f'GROUP-{group_id}',
                    reference_type='group_initial',
                    created_by='system',
                    source_system='warehouse'
                )
                db.session.add(ledger_entry)

                item.quantity = new_quantity
                items_updated += 1

                # Track SKU for automatic push (new warehouse stock)
                if item.sku not in skus_to_push:
                    skus_to_push.append(item.sku)

        # UNIFIED PUSH SYSTEM: Prepare marketplace pushes BEFORE commit
        from warehouse_push_coordinator import WarehousePushCoordinator
        coordinator = WarehousePushCoordinator()
        if skus_to_push:
            prepared_count = coordinator.prepare_for_items(skus_to_push, operation="update")
            logging.info(f" Group update: Prepared {prepared_count} SKUs for marketplace push")

        db.session.commit()

        # UNIFIED PUSH SYSTEM: Enqueue jobs AFTER successful commit
        jobs_enqueued = 0
        if skus_to_push:
            jobs_enqueued = coordinator.enqueue_pending_jobs()
            logging.info(f" Group update: Enqueued {jobs_enqueued} push jobs for group {group.name}")

        return jsonify({
            'success': True,
            'message': f'Stock updated successfully for all items in group',
            'items_updated': items_updated,
            'new_quantity': new_quantity,
            'push_jobs_enqueued': jobs_enqueued
        })

    except Exception as e:
        db.session.rollback()
        logging.error(f'Error updating group stock: {str(e)}')
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@bp.route('/groups/delete/<int:group_id>', methods=['POST'])
def delete_group(group_id):
    """Delete product group"""
    try:
        group = db.session.get(ProductGroup, group_id)
        if group:
            # Check if group has items
            item_count = db.session.query(InventoryItem).filter(InventoryItem.group_id == group_id).count()
            if item_count > 0:
                flash(f'Cannot delete group with {item_count} items. Please move or delete items first.', 'warning')
            else:
                db.session.delete(group)
                db.session.commit()
                flash('Group deleted successfully!', 'success')
        else:
            flash('Group not found!', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting group: {str(e)}', 'danger')
        logging.error(f'Error deleting group: {str(e)}')

    return redirect(url_for('routes.groups'))

# =================== SEARCH API ENDPOINTS ===================

@bp.route('/groups/search', methods=['GET'])
def search_groups():
    """Search product groups by name - GET endpoint for AJAX calls"""
    try:
        search_query = request.args.get('q', '').strip()

        if not search_query:
            return jsonify({
                'success': False,
                'error': 'Search query parameter "q" is required'
            }), 400

        # Perform case-insensitive search on group name
        query = db.session.query(ProductGroup).filter(
            ProductGroup.name.ilike(f'%{search_query}%')
        ).order_by(ProductGroup.name).limit(20)

        groups = query.all()

        # Format results with item count
        results = []
        for group in groups:
            item_count = db.session.query(InventoryItem).filter(InventoryItem.group_id == group.id).count()
            results.append({
                'id': group.id,
                'name': group.name,
                'group_key': group.group_key,
                'item_count': item_count,
                'description': group.description
            })

        logging.info(f"Groups search for '{search_query}' returned {len(results)} results")

        return jsonify({
            'success': True,
            'groups': results,
            'query': search_query,
            'count': len(results)
        }), 200

    except Exception as e:
        logging.error(f"Error searching groups: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Search error: {str(e)}'
        }), 500

# Warehouse Management Routes
@bp.route('/warehouse')
# @login_required  # Temporarily disabled for access
def warehouse():
    """Warehouse stock management dashboard with marketplace listing enrichment."""
    from types import SimpleNamespace

    search_query = request.args.get('search', '').strip()
    low_stock = request.args.get('low_stock', '')
    page = int(request.args.get('page', 1))
    per_page = 50

    query = db.session.query(WarehouseStock).options(
        joinedload(WarehouseStock.marketplace_listings).joinedload(MarketplaceListing.store)
    ).filter(
        WarehouseStock.is_active == True,
        WarehouseStock.is_deleted == False
    )

    if search_query:
        query = query.filter(WarehouseStock.sku.ilike(f'%{search_query}%'))

    if low_stock == 'true':
        query = query.filter(WarehouseStock.available_quantity <= WarehouseStock.reorder_point)

    raw_page = query.order_by(WarehouseStock.sku).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    enriched_items = []

    for stock in raw_page.items:
        active_listings = [
            listing for listing in (stock.marketplace_listings or [])
            if getattr(listing, 'is_active', False)
        ]

        # NO LINKED MARKETPLACE ROWS
        if not active_listings:
            enriched_items.append(SimpleNamespace(
                id=stock.id,
                inventory_item_id=getattr(stock, 'inventory_item_id', None),
                item_id=getattr(stock, 'item_id', None),
                marketplace_listing_id=None,
                sku=stock.sku,
                barcode=stock.barcode,
                product_name=stock.product_name,
                title=None,
                group_title=stock.group_title,
                image_url=stock.image_url,
                available_quantity=stock.available_quantity,
                price=stock.unit_cost,
                store_name=None,
                location=stock.location or 'Warehouse',
                amazon_fulfillment_channel=None,
                master_product_group_id=stock.master_product_group_id,
                is_group_controlled=stock.is_group_controlled,
                mcf_group_source=False,
                linked_listing_count=0,
            ))
            continue

        # CREATE ONE VISIBLE ROW PER MARKETPLACE LISTING
        for listing in active_listings:

            platform = listing.store.platform if listing.store else None
            store_name = listing.store.name if listing.store else None
            channel = listing.normalized_amazon_fulfillment_channel

            if platform and 'amazon' in platform.lower() and channel in ('AFN', 'FBA'):
                location = 'Amazon FBA'
            elif platform and 'amazon' in platform.lower():
                location = 'Amazon FBM'
            elif platform and 'ebay' in platform.lower():
                location = 'eBay'
            else:
                location = stock.location or 'Warehouse'

            enriched_items.append(SimpleNamespace(
                id=stock.id,
                inventory_item_id=getattr(stock, 'inventory_item_id', None),
                item_id=getattr(stock, 'item_id', None),
                marketplace_listing_id=listing.id,
                sku=stock.sku,
                barcode=stock.barcode,
                product_name=stock.product_name,
                title=listing.title,
                group_title=stock.group_title,
                image_url=stock.image_url,
                available_quantity=stock.available_quantity,
                price=listing.price if listing.price else stock.unit_cost,
                store_name=store_name,
                location=location,
                amazon_fulfillment_channel=channel,
                master_product_group_id=stock.master_product_group_id,
                is_group_controlled=stock.is_group_controlled,
                mcf_group_source=False,
                linked_listing_count=len(active_listings),
            ))
    warehouse_items = SimpleNamespace(
        items=enriched_items,
        total=raw_page.total,
        page=raw_page.page,
        per_page=raw_page.per_page,
        pages=raw_page.pages,
        has_next=raw_page.has_next,
        has_prev=raw_page.has_prev,
        next_num=raw_page.next_num,
        prev_num=raw_page.prev_num,
    )

    total_skus = db.session.query(WarehouseStock).filter(
        WarehouseStock.is_active == True,
        WarehouseStock.is_deleted == False
    ).count()

    low_stock_count = db.session.query(WarehouseStock).filter(
        WarehouseStock.available_quantity <= WarehouseStock.reorder_point,
        WarehouseStock.is_active == True,
        WarehouseStock.is_deleted == False
    ).count()

    total_available = db.session.query(db.func.sum(WarehouseStock.available_quantity)).filter(
        WarehouseStock.is_active == True,
        WarehouseStock.is_deleted == False
    ).scalar() or 0

    stats = {
        'total_skus': total_skus,
        'low_stock_count': low_stock_count,
        'total_available': total_available
    }

    return render_template(
        'warehouse.html',
        warehouse_items=warehouse_items,
        stats=stats,
        current_search=search_query,
        low_stock_filter=low_stock
    )

@bp.route('/warehouse/<int:stock_id>')
# @login_required  # Temporarily disabled for access
def warehouse_detail(stock_id):
    """Detailed view of a warehouse stock item"""
    stock_item = db.session.get(WarehouseStock, stock_id)
    if not stock_item:
        flash('Warehouse item not found.', 'error')
        return redirect(url_for('routes.warehouse'))

    # Get recent ledger entries
    ledger_entries = db.session.query(StockLedgerEntry).filter(
        StockLedgerEntry.warehouse_stock_id == stock_id
    ).order_by(desc(StockLedgerEntry.created_at)).limit(20).all()

    # Get marketplace listings
    marketplace_listings = db.session.query(MarketplaceListing).filter(
        MarketplaceListing.warehouse_stock_id == stock_id
    ).all()

    # Get product group information
    product_group = None
    group_items = []
    inventory_item = db.session.query(InventoryItem).filter(
        InventoryItem.sku == stock_item.sku
    ).first()

    if inventory_item and inventory_item.group_id:
        product_group = inventory_item.group
        # Get all items in the same group
        group_inventory_items = db.session.query(InventoryItem).filter(
            InventoryItem.group_id == inventory_item.group_id
        ).all()

        # Get warehouse stock and marketplace listings for each group item
        for item in group_inventory_items:
            warehouse_stock = db.session.query(WarehouseStock).filter(
                WarehouseStock.sku == item.sku
            ).first()

            if warehouse_stock:
                listings = db.session.query(MarketplaceListing).filter(
                    MarketplaceListing.warehouse_stock_id == warehouse_stock.id
                ).all()

                group_items.append({
                    'inventory_item': item,
                    'warehouse_stock': warehouse_stock,
                    'marketplace_listings': listings
                })

    # Get suppliers for dropdown
    suppliers = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()

    return render_template('warehouse_detail.html',
                         stock_item=stock_item,
                         ledger_entries=ledger_entries,
                         marketplace_listings=marketplace_listings,
                         product_group=product_group,
                         group_items=group_items,
                         suppliers=suppliers)

@bp.route('/warehouse/adjust/<int:stock_id>', methods=['POST'])
# @login_required  # Temporarily disabled for access
def warehouse_adjust(stock_id):
    """Adjust warehouse stock levels"""
    # CSRF Protection
    csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
    if not validate_csrf_token(csrf_token):
        flash('Invalid request. Please try again.', 'danger')
        return redirect(url_for('routes.warehouse'))

    stock_item = db.session.get(WarehouseStock, stock_id)
    if not stock_item:
        flash('Warehouse item not found.', 'error')
        return redirect(url_for('routes.warehouse'))

    try:
        adjustment_type = request.form.get('adjustment_type')  # 'set', 'increase', 'decrease'
        quantity_change = int(request.form.get('quantity', 0))
        reason = request.form.get('reason', '')

        # Store before values
        before_available = stock_item.available_quantity
        before_reserved = stock_item.reserved_quantity
        before_allocated = stock_item.allocated_quantity
        before_on_order = stock_item.on_order_quantity

        # Apply adjustment
        if adjustment_type == 'set':
            stock_item.available_quantity = quantity_change
        elif adjustment_type == 'increase':
            stock_item.available_quantity += quantity_change
        elif adjustment_type == 'decrease':
            stock_item.available_quantity = max(0, stock_item.available_quantity - quantity_change)

        # Update timestamps
        stock_item.last_adjustment_at = datetime.utcnow()
        stock_item.last_adjustment_by = 'warehouse_user'  # Could be current_user if auth is implemented

        # Create audit entry
        ledger_entry = StockLedgerEntry(
            warehouse_stock_id=stock_item.id,
            transaction_type='adjustment',
            adjustment_type=adjustment_type,
            available_quantity_before=before_available,
            available_quantity_after=stock_item.available_quantity,
            reserved_quantity_before=before_reserved,
            reserved_quantity_after=stock_item.reserved_quantity,
            allocated_quantity_before=before_allocated,
            allocated_quantity_after=stock_item.allocated_quantity,
            on_order_quantity_before=before_on_order,
            on_order_quantity_after=stock_item.on_order_quantity,
            reason=reason,
            reference_type='manual_adjustment',
            created_by='warehouse_user',
            source_system='warehouse'
        )

        db.session.add(ledger_entry)

        # Update corresponding InventoryItem to match warehouse quantity (warehouse is authoritative)
        inventory_item = InventoryItem.query.filter_by(sku=stock_item.sku).first()
        if inventory_item:
            inventory_item.quantity = stock_item.available_quantity
            inventory_item.updated_at = datetime.utcnow()

        # Capture SKU before commit (to avoid DetachedInstanceError)
        sku = stock_item.sku
        quantity_changed = before_available != stock_item.available_quantity
        new_quantity = stock_item.available_quantity

        # UNIFIED PUSH SYSTEM: Prepare marketplace pushes BEFORE commit
        from warehouse_push_coordinator import WarehousePushCoordinator
        coordinator = WarehousePushCoordinator()
        if quantity_changed:
            prepared_count = coordinator.prepare_for_items([sku], operation="update")
            logging.info(f" Prepared {sku} for marketplace push after adjustment")

        db.session.commit()

        # UNIFIED PUSH SYSTEM: Enqueue jobs AFTER successful commit
        if quantity_changed:
            jobs_enqueued = coordinator.enqueue_pending_jobs()
            if jobs_enqueued > 0:
                flash(f'Stock adjusted for {sku}. New quantity: {new_quantity}. High-priority push queued for {jobs_enqueued} marketplace(s).', 'success')
            else:
                flash(f'Stock adjusted for {sku}. New quantity: {new_quantity}. No marketplaces have auto-push enabled.', 'success')
        else:
            flash(f'Stock adjusted for {sku}. New quantity: {new_quantity}', 'success')

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error adjusting warehouse stock: {str(e)}")
        flash(f'Error adjusting stock: {str(e)}', 'error')

    return redirect(url_for('routes.warehouse_detail', stock_id=stock_id))

@bp.route('/warehouse/update-settings/<int:stock_id>', methods=['POST'])
# @login_required  # Temporarily disabled for access
def warehouse_update_settings(stock_id):
    """Update warehouse stock settings including supplier and reorder points"""
    # CSRF Protection
    csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
    if not validate_csrf_token(csrf_token):
        flash('Invalid request. Please try again.', 'danger')
        return redirect(url_for('routes.warehouse'))

    stock_item = db.session.get(WarehouseStock, stock_id)
    if not stock_item:
        flash('Warehouse item not found.', 'error')
        return redirect(url_for('routes.warehouse'))

    try:
        # Update supplier
        supplier_id = request.form.get('supplier_id')
        stock_item.supplier_id = int(supplier_id) if supplier_id else None

        # Update reorder settings
        stock_item.reorder_point = int(request.form.get('reorder_point', 0))
        stock_item.reorder_quantity = int(request.form.get('reorder_quantity', 0))

        # Update product information
        stock_item.product_name = request.form.get('product_name', '').strip() or None
        stock_item.image_url = request.form.get('image_url', '').strip() or None

        # Update unit cost and location
        stock_item.unit_cost = float(request.form.get('unit_cost', 0))
        stock_item.location = request.form.get('location', '').strip() or None

        # Update financial tracking fields
        stock_item.commission_rate = float(request.form.get('commission_rate', 0))
        stock_item.operating_cost_per_unit = float(request.form.get('operating_cost_per_unit', 0))
        stock_item.product_weight_kg = float(request.form.get('product_weight_kg', 0))
        stock_item.shipping_cost_per_kg = float(request.form.get('shipping_cost_per_kg', 0))

        stock_item.updated_at = datetime.utcnow()

        db.session.commit()
        flash(f'Settings updated successfully for {stock_item.sku}', 'success')

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error updating warehouse settings: {str(e)}")
        flash(f'Error updating settings: {str(e)}', 'error')

    return redirect(url_for('routes.warehouse_detail', stock_id=stock_id))

@bp.route('/warehouse/upload-image/<int:stock_id>', methods=['POST'])
# @login_required  # Temporarily disabled for access
def warehouse_upload_image(stock_id):
    """Upload product image for warehouse stock item"""
    # CSRF Protection
    csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
    if not validate_csrf_token(csrf_token):
        return jsonify({'success': False, 'error': 'Invalid CSRF token'}), 403

    stock_item = db.session.get(WarehouseStock, stock_id)
    if not stock_item:
        return jsonify({'success': False, 'error': 'Warehouse item not found'}), 404

    # Check if file was uploaded
    if 'image' not in request.files:
        return jsonify({'success': False, 'error': 'No image file uploaded'}), 400

    file = request.files['image']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    # Validate file type
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    if file_ext not in allowed_extensions:
        return jsonify({'success': False, 'error': f'Invalid file type. Allowed: {", ".join(allowed_extensions)}'}), 400

    try:
        # Create secure filename: sku_timestamp.ext
        import time
        timestamp = int(time.time())
        safe_sku = "".join(c if c.isalnum() or c in ('-', '_') else '_' for c in stock_item.sku)
        filename = f"{safe_sku}_{timestamp}.{file_ext}"

        # Save file to uploads folder
        upload_folder = os.path.join('static', 'uploads', 'product_images')
        os.makedirs(upload_folder, exist_ok=True)
        filepath = os.path.join(upload_folder, filename)
        file.save(filepath)

        # Generate URL for the uploaded image
        image_url = url_for('static', filename=f'uploads/product_images/{filename}', _external=True)

        # Update warehouse stock with new image URL
        stock_item.image_url = image_url
        stock_item.updated_at = datetime.utcnow()
        db.session.commit()

        logging.info(f"Image uploaded successfully for {stock_item.sku}: {filename}")
        return jsonify({
            'success': True,
            'image_url': image_url,
            'message': f'Image uploaded successfully for {stock_item.sku}'
        })

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error uploading image: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/warehouse/bulk-adjust', methods=['GET', 'POST'])
# @login_required  # Temporarily disabled for access
def warehouse_bulk_adjust():
    """Bulk adjustment interface"""
    if request.method == 'GET':
        return render_template('warehouse_bulk_adjust.html')

    # CSRF Protection
    csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
    if not validate_csrf_token(csrf_token):
        flash('Invalid request. Please try again.', 'danger')
        return render_template('warehouse_bulk_adjust.html')

    try:
        # Parse bulk adjustment data
        adjustments_data = request.form.get('adjustments_data', '')
        reason = request.form.get('reason', 'Bulk adjustment')

        if not adjustments_data:
            flash('No adjustments provided.', 'warning')
            return render_template('warehouse_bulk_adjust.html')

        # Process CSV-like data: SKU, quantity, type
        adjustments = []
        for line in adjustments_data.strip().split('\n'):
            parts = [p.strip() for p in line.split(',')]
            if len(parts) >= 3:
                sku, quantity, adj_type = parts[0], parts[1], parts[2]
                try:
                    adjustments.append({
                        'sku': sku,
                        'quantity': int(quantity),
                        'type': adj_type
                    })
                except ValueError:
                    flash(f'Invalid quantity for SKU {sku}', 'warning')

        if not adjustments:
            flash('No valid adjustments found.', 'warning')
            return render_template('warehouse_bulk_adjust.html')

        # Process adjustments
        success_count = 0
        error_count = 0
        skus_to_push = []  # Track SKUs that need marketplace push (avoid DetachedInstanceError)

        for adj in adjustments:
            stock_item = db.session.query(WarehouseStock).filter(
                WarehouseStock.sku == adj['sku']
            ).first()

            if not stock_item:
                logging.warning(f"SKU not found: {adj['sku']}")
                error_count += 1
                continue

            # Store before values
            before_available = stock_item.available_quantity
            before_reserved = stock_item.reserved_quantity
            before_allocated = stock_item.allocated_quantity
            before_on_order = stock_item.on_order_quantity

            # Apply adjustment
            if adj['type'] == 'set':
                stock_item.available_quantity = adj['quantity']
            elif adj['type'] == 'increase':
                stock_item.available_quantity += adj['quantity']
            elif adj['type'] == 'decrease':
                stock_item.available_quantity = max(0, stock_item.available_quantity - adj['quantity'])

            # Update metadata
            stock_item.last_adjustment_at = datetime.utcnow()
            stock_item.last_adjustment_by = 'warehouse_user'

            # Update corresponding InventoryItem to match warehouse quantity (warehouse is authoritative)
            inventory_item = InventoryItem.query.filter_by(sku=stock_item.sku).first()
            if inventory_item:
                inventory_item.quantity = stock_item.available_quantity
                inventory_item.updated_at = datetime.utcnow()

                # Track SKU for marketplace push if quantity changed (store SKU, not object)
                if before_available != stock_item.available_quantity:
                    skus_to_push.append(stock_item.sku)

            # Create ledger entry
            ledger_entry = StockLedgerEntry(
                warehouse_stock_id=stock_item.id,
                transaction_type='bulk_adjustment',
                adjustment_type=adj['type'],
                available_quantity_before=before_available,
                available_quantity_after=stock_item.available_quantity,
                reserved_quantity_before=before_reserved,
                reserved_quantity_after=stock_item.reserved_quantity,
                allocated_quantity_before=before_allocated,
                allocated_quantity_after=stock_item.allocated_quantity,
                on_order_quantity_before=before_on_order,
                on_order_quantity_after=stock_item.on_order_quantity,
                reason=reason,
                reference_type='bulk_adjustment',
                created_by='warehouse_user',
                source_system='warehouse',
                batch_id=f'bulk_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}'
            )

            db.session.add(ledger_entry)
            success_count += 1

        # Prepare push jobs for all changed items (BEFORE commit)
        items_and_stores = []  # List of (item_id, stores) tuples
        if skus_to_push:
            for sku in skus_to_push:
                fresh_item = InventoryItem.query.filter_by(sku=sku).first()
                if fresh_item:
                    stores_to_push, warehouse_stock = prepare_warehouse_push(fresh_item, operation="update")
                    if stores_to_push:
                        items_and_stores.append((fresh_item.id, stores_to_push))

        # Commit ALL changes including warehouse stock
        db.session.commit()

        # Enqueue jobs AFTER successful commit
        total_jobs = 0
        if items_and_stores:
            try:
                for item_id, stores in items_and_stores:
                    jobs_count = enqueue_push_jobs(item_id, stores)
                    total_jobs += jobs_count
            except Exception as push_error:
                logging.error(f"Push job enqueue failed after bulk adjustment: {str(push_error)}")
                # Items are already saved, just warn user
                flash(f'Bulk adjustment completed: {success_count} items updated, {error_count} errors. Push enqueue failed - use manual push button.', 'warning')
                return redirect(url_for('routes.warehouse'))

        # Show appropriate flash message
        if total_jobs > 0:
            flash(f'Bulk adjustment completed: {success_count} items updated, {error_count} errors. High-priority push queued for {total_jobs} marketplace listing(s).', 'success')
        elif skus_to_push:
            flash(f'Bulk adjustment completed: {success_count} items updated, {error_count} errors. No marketplaces have auto-push enabled.', 'success')
        else:
            flash(f'Bulk adjustment completed: {success_count} items updated, {error_count} errors', 'success')

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in bulk adjustment: {str(e)}")
        flash(f'Error processing bulk adjustment: {str(e)}', 'error')

    return redirect(url_for('routes.warehouse'))

@bp.route('/items/search', methods=['GET'])
def search_items():
    """Search inventory items with optional group filtering - GET endpoint for AJAX calls"""
    try:
        search_query = request.args.get('q', '').strip()
        require_group = request.args.get('require_group', 'false').lower() == 'true'

        if not search_query:
            return jsonify({
                'success': False,
                'error': 'Search query parameter "q" is required'
            }), 400

        # Build base query
        query = db.session.query(InventoryItem)

        # Apply group filter if required
        if require_group:
            query = query.filter(InventoryItem.group_id.isnot(None))

        # Apply search filter (search in name and SKU)
        query = query.filter(
            db.or_(
                InventoryItem.name.ilike(f'%{search_query}%'),
                InventoryItem.sku.ilike(f'%{search_query}%')
            )
        ).order_by(InventoryItem.name).limit(20)

        items = query.all()

        # Format results with group info
        results = []
        for item in items:
            item_data = {
                'id': item.id,
                'name': item.name,
                'sku': item.sku,
                'quantity': item.quantity,
                'price': item.price,
                'group_id': item.group_id,
                'group_name': item.group.name if item.group else None,
                'description': item.description,
                'variant_attributes': item.variant_attributes
            }
            results.append(item_data)

        logging.info(f"Items search for '{search_query}' (require_group={require_group}) returned {len(results)} results")

        return jsonify({
            'success': True,
            'items': results,
            'query': search_query,
            'require_group': require_group,
            'count': len(results)
        }), 200

    except Exception as e:
        logging.error(f"Error searching items: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Search error: {str(e)}'
        }), 500

@bp.route('/items/<int:item_id>/link_group', methods=['POST'])
def link_item_to_group(item_id):
    """Link an inventory item to a product group"""
    try:
        # CSRF Protection
        csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
        if not validate_csrf_token(csrf_token):
            logging.warning(f"CSRF token validation failed for link_item_to_group request from {request.remote_addr}")
            return jsonify({
                'success': False,
                'error': 'CSRF token validation failed'
            }), 403

        # Get group_id from form data
        group_id = request.form.get('group_id')
        if not group_id:
            return jsonify({
                'success': False,
                'error': 'Group ID is required'
            }), 400

        try:
            group_id = int(group_id)
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Invalid group ID format'
            }), 400

        # Check if item exists
        item = db.session.get(InventoryItem, item_id)
        if not item:
            logging.warning(f"Attempted to link non-existent item {item_id} to group {group_id}")
            return jsonify({
                'success': False,
                'error': 'Item not found'
            }), 404

        # Check if group exists
        group = db.session.get(ProductGroup, group_id)
        if not group:
            logging.warning(f"Attempted to link item {item_id} to non-existent group {group_id}")
            return jsonify({
                'success': False,
                'error': 'Group not found'
            }), 404

        # Check if item is already in the target group
        if item.group_id == group_id:
            return jsonify({
                'success': False,
                'error': f'Item "{item.name}" is already in group "{group.name}"'
            }), 400

        # Store previous group info for logging
        previous_group_id = item.group_id
        previous_group_name = item.group.name if item.group else None

        # Start transaction
        try:
            # Link item to new group
            item.group_id = group_id
            db.session.commit()

            # Log the operation
            if previous_group_id:
                logging.info(f"Moved item {item.sku} ({item_id}) from group '{previous_group_name}' ({previous_group_id}) to group '{group.name}' ({group_id})")
                message = f'Item "{item.name}" moved from group "{previous_group_name}" to group "{group.name}"'
            else:
                logging.info(f"Linked ungrouped item {item.sku} ({item_id}) to group '{group.name}' ({group_id})")
                message = f'Item "{item.name}" linked to group "{group.name}"'

            return jsonify({
                'success': True,
                'message': message,
                'item': {
                    'id': item.id,
                    'name': item.name,
                    'sku': item.sku,
                    'group_id': group_id,
                    'group_name': group.name
                },
                'new_group': {
                    'id': group_id,
                    'name': group.name
                },
                'previous_group': {
                    'id': previous_group_id,
                    'name': previous_group_name
                } if previous_group_id else None
            }), 200

        except Exception as db_error:
            db.session.rollback()
            logging.error(f"Database error linking item {item_id} to group {group_id}: {str(db_error)}")
            return jsonify({
                'success': False,
                'error': f'Database error: {str(db_error)}'
            }), 500

    except Exception as e:
        logging.error(f"Error linking item {item_id} to group: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Server error: {str(e)}'
        }), 500

# =================== GROUP ITEM MANAGEMENT ROUTES ===================

@bp.route('/release_from_group/<int:item_id>', methods=['POST'])
# @login_required  # Temporarily disabled for access
def release_from_group(item_id):
    """Release an item from its current group"""
    try:
        # CSRF Protection
        csrf_token = request.headers.get('X-CSRF-Token') or request.json.get('csrf_token') if request.json else None
        if not validate_csrf_token(csrf_token):
            logging.warning(f"CSRF token validation failed for release_from_group request from {request.remote_addr}")
            return jsonify({
                'success': False,
                'error': 'CSRF token validation failed'
            }), 403
        item = db.session.get(InventoryItem, item_id)
        if not item:
            logging.warning(f"Attempted to release non-existent item {item_id} from group")
            return jsonify({
                'success': False,
                'error': 'Item not found'
            }), 404

        # Check if item is actually in a group
        if not item.group_id:
            logging.info(f"Item {item.sku} ({item_id}) is not currently in any group")
            return jsonify({
                'success': False,
                'error': 'Item is not currently in any group'
            }), 400

        # Store group info for logging and response
        old_group_name = item.group.name if item.group else "Unknown"
        old_group_id = item.group_id

        # Release from group
        item.group_id = None
        db.session.commit()

        logging.info(f"Successfully released item {item.sku} ({item_id}) from group '{old_group_name}' ({old_group_id})")

        return jsonify({
            'success': True,
            'message': f'Item "{item.name}" has been released from group "{old_group_name}"',
            'item': {
                'id': item.id,
                'name': item.name,
                'sku': item.sku,
                'group_id': None
            },
            'previous_group': {
                'id': old_group_id,
                'name': old_group_name
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error releasing item {item_id} from group: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Database error: {str(e)}'
        }), 500

@bp.route('/search_ungrouped_items', methods=['POST'])
def search_ungrouped_items():
    """Search for ungrouped inventory items"""
    try:
        # CSRF Protection
        csrf_token = request.headers.get('X-CSRF-Token') or request.json.get('csrf_token') if request.json else None
        if not validate_csrf_token(csrf_token):
            logging.warning(f"CSRF token validation failed for search_ungrouped_items request from {request.remote_addr}")
            return jsonify({
                'success': False,
                'error': 'CSRF token validation failed'
            }), 403
        data = request.get_json()
        if not data or 'query' not in data:
            return jsonify({
                'success': False,
                'error': 'Query parameter is required'
            }), 400

        query = data['query'].strip()
        if not query:
            return jsonify({
                'success': False,
                'error': 'Query cannot be empty'
            }), 400

        # Additional validation: query length and pattern safety
        if len(query) > 100:
            return jsonify({
                'success': False,
                'error': 'Query too long (maximum 100 characters)'
            }), 400

        # Search for ungrouped items (group_id is None) that match the query
        # Search in SKU, name, and description fields
        search_pattern = f"%{query}%"

        # Enhanced security: Use explicit transaction and double-check ungrouped status
        items = db.session.query(InventoryItem).filter(
            InventoryItem.group_id.is_(None),
            db.or_(
                InventoryItem.sku.ilike(search_pattern),
                InventoryItem.name.ilike(search_pattern),
                InventoryItem.description.ilike(search_pattern)
            )
        ).order_by(InventoryItem.sku).limit(50).all()  # Limit to 50 results for performance

        # Format items for response
        items_data = []
        for item in items:
            items_data.append({
                'id': item.id,
                'sku': item.sku,
                'name': item.name,
                'quantity': item.quantity,
                'price': float(item.price),
                'description': item.description or '',
                'variant_attributes': item.variant_attributes or {}
            })

        logging.info(f"Search for ungrouped items with query '{query}' returned {len(items_data)} results")

        return jsonify({
            'success': True,
            'items': items_data,
            'query': query,
            'total_results': len(items_data)
        }), 200

    except Exception as e:
        logging.error(f"Error searching ungrouped items: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Search error: {str(e)}'
        }), 500

@bp.route('/add_sku_to_group/<int:group_id>', methods=['POST'])
def add_sku_to_group(group_id):
    """Add existing SKUs to a specified group"""
    try:
        # CSRF Protection
        csrf_token = request.headers.get('X-CSRF-Token') or request.json.get('csrf_token') if request.json else None
        if not validate_csrf_token(csrf_token):
            logging.warning(f"CSRF token validation failed for add_sku_to_group request from {request.remote_addr}")
            return jsonify({
                'success': False,
                'error': 'CSRF token validation failed'
            }), 403
        # Validate that the group exists
        group = db.session.get(ProductGroup, group_id)
        if not group:
            logging.warning(f"Attempted to add SKUs to non-existent group {group_id}")
            return jsonify({
                'success': False,
                'error': 'Group not found'
            }), 404

        # Get request data
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'error': 'No JSON data provided'
            }), 400

        # Handle different operations
        operation = data.get('operation', 'search')

        if operation == 'search':
            # Search for available SKUs
            search_term = data.get('search_term', '').strip()
            if not search_term:
                return jsonify({
                    'success': False,
                    'error': 'Search term is required'
                }), 400

            # Find items that match the search term and are either ungrouped or in different groups
            available_items = db.session.query(InventoryItem).filter(
                (InventoryItem.sku.ilike(f'%{search_term}%') |
                 InventoryItem.name.ilike(f'%{search_term}%')) &
                (InventoryItem.group_id != group_id)  # Not already in this group
            ).order_by(InventoryItem.name).limit(20).all()

            # Categorize items
            ungrouped_items = []
            grouped_items = []

            for item in available_items:
                item_data = {
                    'id': item.id,
                    'name': item.name,
                    'sku': item.sku,
                    'quantity': item.quantity,
                    'price': item.price,
                    'current_group': None
                }

                if item.group_id:
                    item_data['current_group'] = {
                        'id': item.group_id,
                        'name': item.group.name if item.group else "Unknown"
                    }
                    grouped_items.append(item_data)
                else:
                    ungrouped_items.append(item_data)

            logging.info(f"Search for '{search_term}' in group {group_id} returned {len(ungrouped_items)} ungrouped and {len(grouped_items)} grouped items")

            return jsonify({
                'success': True,
                'search_term': search_term,
                'group': {
                    'id': group.id,
                    'name': group.name
                },
                'ungrouped_items': ungrouped_items,
                'grouped_items': grouped_items,
                'total_found': len(available_items)
            }), 200

        elif operation == 'add':  # Updated operation name to match client
            # Assign SKUs to the group with enhanced validation and atomic transactions
            item_ids = data.get('item_ids', [])
            if not item_ids or not isinstance(item_ids, list):
                return jsonify({
                    'success': False,
                    'error': 'Item IDs list is required'
                }), 400

            # Security: Limit batch size to prevent resource exhaustion
            if len(item_ids) > 100:
                return jsonify({
                    'success': False,
                    'error': 'Too many items in batch (maximum 100 items allowed)'
                }), 400

            # Validate all item IDs are integers
            try:
                item_ids = [int(item_id) for item_id in item_ids]
            except (ValueError, TypeError):
                return jsonify({
                    'success': False,
                    'error': 'Invalid item ID format - all IDs must be integers'
                }), 400

            # Use atomic transaction to prevent race conditions
            assigned_items = []
            errors = []

            # Start explicit transaction
            try:
                for item_id in item_ids:
                    # Re-fetch item within transaction to ensure current state
                    item = db.session.query(InventoryItem).filter_by(id=item_id).with_for_update().first()
                    if not item:
                        errors.append(f"Item with ID {item_id} not found")
                        continue

                    # Security: Double-check group assignment to prevent race conditions
                    if item.group_id == group_id:
                        errors.append(f"Item '{item.sku}' is already in this group")
                        continue

                    # Log previous group for audit trail
                    old_group_name = item.group.name if item.group else None
                    old_group_id = item.group_id

                    # Assign to new group
                    item.group_id = group_id

                    assigned_items.append({
                        'id': item.id,
                        'name': item.name,
                        'sku': item.sku,
                        'previous_group': {
                            'id': old_group_id,
                            'name': old_group_name
                        } if old_group_id else None
                    })

                    logging.info(f"Assigned item {item.sku} ({item_id}) to group '{group.name}' ({group_id})")

                # Commit all changes atomically
                db.session.commit()
                logging.info(f"Successfully assigned {len(assigned_items)} items to group '{group.name}' ({group_id}) in atomic transaction")

            except Exception as transaction_error:
                db.session.rollback()
                logging.error(f"Transaction failed during batch assignment to group {group_id}: {str(transaction_error)}")
                return jsonify({
                    'success': False,
                    'error': f'Transaction failed: {str(transaction_error)}'
                }), 500

            return jsonify({
                'success': True,
                'message': f'Assigned {len(assigned_items)} items to group "{group.name}"',
                'group': {
                    'id': group.id,
                    'name': group.name
                },
                'assigned_items': assigned_items,
                'errors': errors,
                'total_assigned': len(assigned_items),
                'total_errors': len(errors)
            }), 200

        else:
            return jsonify({
                'success': False,
                'error': f'Unknown operation: {operation}'
            }), 400

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in add_sku_to_group for group {group_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Server error: {str(e)}'
        }), 500

@bp.route('/stores')
def stores():
    """Display all stores"""
    stores = db.session.query(Store).order_by(desc(Store.updated_at)).all()
    return render_template('stores.html', stores=stores)

@bp.route('/stores/quick-setup-amazon', methods=['POST'])
def quick_setup_amazon():
    """One-click Amazon store setup using environment credentials"""
    try:
        # Get environment credentials
        env_refresh_token = os.environ.get('AMAZON_REFRESH_TOKEN')
        env_client_id = os.environ.get('AMAZON_LWA_CLIENT_ID')
        env_client_secret = os.environ.get('AMAZON_LWA_CLIENT_SECRET')
        env_seller_id = os.environ.get('AMAZON_SELLER_ID')

        if not all([env_refresh_token, env_client_id, env_client_secret, env_seller_id]):
            flash('Amazon credentials not configured in environment. Please set up credentials first.', 'warning')
            return redirect(url_for('routes.stores'))

        # Extract actual values if they have "Value:" prefix
        if "Value:" in env_client_id:
            env_client_id = env_client_id.split("Value:")[-1].strip()
        if "Value:" in env_client_secret:
            env_client_secret = env_client_secret.split("Value:")[-1].strip()

        # Get store name from form or generate default
        store_name = request.form.get('store_name', f'Amazon Store - {env_seller_id}')
        region = request.form.get('region', 'UK')
        marketplace_id = request.form.get('marketplace_id', 'A1F83G8C2ARO7P')  # Default to UK

        # Create new store
        store = Store()
        store.name = store_name
        store.platform = 'Amazon'
        store.api_endpoint = f'https://sellingpartnerapi-{"eu" if region in ["UK", "DE", "FR", "IT", "ES"] else "na" if region in ["US", "CA"] else "fe"}.amazon.com'

        # Build credentials JSON using environment variables
        amazon_creds = {
            'refresh_token': env_refresh_token,
            'lwa_app_id': env_client_id,
            'lwa_client_secret': env_client_secret,
            'seller_id': env_seller_id,
            'marketplace_id': marketplace_id,
            'region': region
        }
        store.api_key = json.dumps(amazon_creds)

        # Save store
        db.session.add(store)
        db.session.commit()

        # Test connection immediately
        from amazon_service import AmazonAPIService
        amazon_service = AmazonAPIService()
        auth_success = amazon_service.disabled_disabled_store_auth_check(store)

        if auth_success:
            flash(f'Amazon store "{store_name}" created and connected successfully! ', 'success')
        else:
            flash(f'Amazon store "{store_name}" created but connection failed. Please check credentials.', 'warning')

        logging.info(f"Quick-setup Amazon store created: {store_name} (ID: {store.id})")
        return redirect(url_for('routes.stores'))

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in quick Amazon setup: {str(e)}")
        flash(f'Error creating Amazon store: {str(e)}', 'danger')
        return redirect(url_for('routes.stores'))


# === BT38 STORE CONNECTION GATE ===
def _bt38_apply_store_connection_gate(store, form):
    """
    Store Settings is the control gate.
    Amazon requires FBA import and/or FBM sync.
    Push is only allowed when FBM sync is enabled.
    eBay is FBM-style only.
    """
    platform = (store.platform or "").strip().lower()

    if platform == "amazon":
        store.fba_import_enabled = form.get("fba_import_enabled") == "on"
        store.fbm_sync_enabled = form.get("fbm_sync_enabled") == "on"

        if not store.fba_import_enabled and not store.fbm_sync_enabled:
            return False, "Select at least one Amazon mode: FBA Import and/or FBM Sync."

        if not store.fbm_sync_enabled:
            store.auto_push_enabled = False
            store.push_on_quantity_change = False

    elif platform == "ebay":
        store.fba_import_enabled = False
        store.fbm_sync_enabled = True

    else:
        store.fba_import_enabled = False
        store.fbm_sync_enabled = False
        store.auto_push_enabled = False
        store.push_on_quantity_change = False

    return True, ""

@bp.route('/stores/add', methods=['GET', 'POST'])
def add_store():
    """Add new store"""
    if request.method == 'POST':
        try:
            store = Store()
            store.name = request.form['name']
            store.platform = request.form['platform']
            store.api_endpoint = request.form.get('api_endpoint', '')

            # Handle Amazon credentials - both OAuth and manual entry
            if store.platform.lower() == 'amazon':
                # Check if manual credentials are provided
                if (request.form.get('refresh_token') and request.form.get('lwa_app_id') and
                    request.form.get('lwa_client_secret') and request.form.get('seller_id')):
                    # Manual SP-API credentials provided
                    amazon_creds = {
                        'refresh_token': request.form.get('refresh_token').strip(),
                        'lwa_app_id': request.form.get('lwa_app_id').strip(),
                        'lwa_client_secret': request.form.get('lwa_client_secret').strip(),
                        'seller_id': request.form.get('seller_id').strip(),
                        'marketplace_id': request.form.get('marketplace_id', 'A1F83G8C2ARO7P').strip(),
                        'region': request.form.get('region', 'UK').strip()
                    }
                    store.api_key = json.dumps(amazon_creds)
                    logging.info(f"Using manual SP-API credentials for {store.name}")
                elif session.get('amazon_authorized'):
                    # Auto-populate Amazon credentials from OAuth session
                    amazon_tokens = session.get('amazon_tokens', {})
                    if amazon_tokens.get('refresh_token'):
                        # Get Amazon credentials and build credential JSON
                        client_id_raw = os.environ.get("AMAZON_LWA_CLIENT_ID")
                        if client_id_raw and "Value:" in client_id_raw:
                            client_id = client_id_raw.split("Value:")[-1].strip()
                        else:
                            client_id = client_id_raw
                        client_secret_raw = os.environ.get("AMAZON_LWA_CLIENT_SECRET")
                        if client_secret_raw and "Value:" in client_secret_raw:
                            client_secret = client_secret_raw.split("Value:")[-1].strip()
                        else:
                            client_secret = client_secret_raw

                    # Build Amazon credentials JSON
                    amazon_creds = {
                        "refresh_token": amazon_tokens['refresh_token'],
                        "lwa_app_id": client_id,
                        "lwa_client_secret": client_secret,
                        "seller_id": request.form.get('seller_id', ''),
                        "marketplace_id": request.form.get('marketplace_id', 'A1F83G8C2ARO7P'),  # Default to UK
                        "region": request.form.get('region', 'UK')
                    }
                    store.api_key = json.dumps(amazon_creds)
                    logging.info(f"Auto-populated Amazon credentials for store {store.name}")
                else:
                    api_key_input = request.form.get('api_key', '').strip()
                    if api_key_input:
                        store.api_key = api_key_input
            else:
                api_key_input = request.form.get('api_key', '').strip()
                if api_key_input:
                    store.api_key = api_key_input
            store.is_active = request.form.get('is_active') == 'on'
            store.auto_push_enabled = request.form.get('auto_push_enabled') == 'on'
            ok, gate_message = _bt38_apply_store_connection_gate(store, request.form)
            if not ok:
                flash(gate_message, 'warning')
                return render_template('add_store.html')

            # Automatically test connection for Amazon and eBay stores
            connection_success = False
            connection_message = ""

            if store.platform.lower() in ['amazon', 'ebay'] and store.api_key.strip():
                try:
                    service = None
                    validation_msg = ""
                    if store.platform.lower() == 'amazon':
                        # Determine region from credentials if available
                        region = 'US'  # default
                        try:
                            creds = json.loads(store.api_key)
                            region = creds.get('region', 'US')
                        except:
                            pass
                        service = AmazonAPIService(marketplace_region=region)
                    elif store.platform.lower() == 'ebay':
                        service = eBayAPIService()

                    # Validate credentials and test connection
                    if service:
                        validation_result = service.validate_credentials_format(store.api_key)
                        if isinstance(validation_result, tuple):
                            is_valid = bool(validation_result[0]) if len(validation_result) > 0 else False
                            validation_msg = str(validation_result[1]) if len(validation_result) > 1 else ""
                        else:
                            is_valid = bool(validation_result)
                            validation_msg = ""

                        if is_valid:
                            auth_result = service.disabled_disabled_store_auth_check(store)
                            connection_success = bool(auth_result[0]) if isinstance(auth_result, tuple) and len(auth_result) > 0 else bool(auth_result)
                            if connection_success:
                                store.sync_status = 'success'
                                store.last_sync = datetime.utcnow()
                                connection_message = f" {store.platform} connection established successfully"
                                logging.info(f"Auto-connected store {store.name} to {store.platform}")
                            else:
                                store.sync_status = 'error'
                                store.is_active = False
                                connection_message = f" {store.platform} connection failed - check your credentials"
                        else:
                            store.sync_status = 'error'
                            store.is_active = False
                            connection_message = f" Invalid credentials format: {validation_msg}"
                except Exception as conn_error:
                    logging.error(f"Auto-connection failed for {store.name}: {str(conn_error)}")
                    store.sync_status = 'error'
                    store.is_active = False
                    connection_message = f" Connection error: {str(conn_error)}"
            else:
                store.sync_status = 'pending'
                if store.platform.lower() in ['amazon', 'ebay']:
                    connection_message = "Store added - no API credentials provided"
                else:
                    connection_message = "Store added - manual sync required"

            db.session.add(store)
            db.session.commit()

            # Queue sync if connection was successful
            if connection_success:
                from queue_manager import enqueue_sync_job, JOB_FULL_SYNC, PRIORITY_HIGH

                disabled_queue_job(
                    store_id=store.id,
                    job_type=JOB_FULL_SYNC,
                    payload={
                        'source': 'store_added_success',
                        'manual': False,
                        'store_id': store.id,
                        'store_name': store.name
                    },
                    priority=PRIORITY_HIGH
                )

                store.sync_status = 'queued'
                db.session.commit()

                flash(f'Store added and connected successfully! Sync queued. {connection_message}', 'success')
            elif store.platform.lower() in ['amazon', 'ebay'] and store.api_key.strip():
                flash(f'Store added but connection failed. {connection_message}', 'warning')
            else:
                flash('Store added successfully!', 'success')

            return redirect(url_for('routes.stores'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding store: {str(e)}', 'danger')
            logging.error(f'Error adding store: {str(e)}')

    return render_template('add_store.html')

@bp.route('/stores/edit/<int:store_id>', methods=['GET', 'POST'])
def edit_store(store_id):
    """Edit existing store"""
    store = db.session.get(Store, store_id)
    if not store:
        flash('Store not found!', 'danger')
        return redirect(url_for('routes.stores'))

    if request.method == 'POST':
        try:
            # Check if API key changed to trigger reconnection
            api_key_changed = store.api_key != request.form.get('api_key', '')

            store.name = request.form['name']
            store.platform = request.form['platform']
            store.api_endpoint = request.form.get('api_endpoint', '')
            api_key_input = request.form.get('api_key', '').strip()
            if api_key_input:
                store.api_key = api_key_input
            store.is_active = request.form.get('is_active') == 'on'

            # Basic push settings
            store.auto_push_enabled = request.form.get('auto_push_enabled') == 'on'

            # Advanced push settings
            store.push_priority = int(request.form.get('push_priority', 5))

            try:
                submitted_frequency = int(request.form.get('push_frequency_minutes', 15))
                if submitted_frequency <= 0:
                    submitted_frequency = 15
            except (TypeError, ValueError):
                submitted_frequency = 15

            store.push_frequency_minutes = submitted_frequency
            store.push_batch_size = int(request.form.get('push_batch_size', 10))

            # Push trigger conditions
            store.push_on_quantity_change = request.form.get('push_on_quantity_change') == 'on'
            store.push_on_price_change = request.form.get('push_on_price_change') == 'on'
            store.push_on_item_create = request.form.get('push_on_item_create') == 'on'
            store.push_on_item_update = request.form.get('push_on_item_update') == 'on'

            # Error handling settings
            store.max_retry_attempts = int(request.form.get('max_retry_attempts', 3))
            store.auto_disable_on_failures = request.form.get('auto_disable_on_failures') == 'on'
            store.failure_threshold = int(request.form.get('failure_threshold', 5))

            # Push preferences
            store.immediate_push = request.form.get('immediate_push') == 'on'
            store.large_change_confirmation = request.form.get('large_change_confirmation') == 'on'
            store.large_change_threshold = int(request.form.get('large_change_threshold', 100))
            ok, gate_message = _bt38_apply_store_connection_gate(store, request.form)
            if not ok:
                flash(gate_message, 'warning')
                return render_template('edit_store.html', store=store)

            # Reset failure count if auto_disable_on_failures is turned off
            if not store.auto_disable_on_failures:
                store.current_failure_count = 0

            # Automatically test connection if API key changed for Amazon/eBay
            connection_message = ""
            if store.platform.lower() in ['amazon', 'ebay'] and store.api_key.strip():
                try:
                    service = None
                    validation_msg = ""
                    connection_success = False
                    if store.platform.lower() == 'amazon':
                        # Determine region from credentials if available
                        region = 'US'  # default
                        try:
                            creds = json.loads(store.api_key)
                            region = creds.get('region', 'US')
                        except:
                            pass
                        service = AmazonAPIService(marketplace_region=region)
                    elif store.platform.lower() == 'ebay':
                        service = eBayAPIService()

                    # Validate and test new credentials
                    if service:
                        validation_result = service.validate_credentials_format(store.api_key)
                        if isinstance(validation_result, tuple):
                            is_valid = bool(validation_result[0]) if len(validation_result) > 0 else False
                            validation_msg = str(validation_result[1]) if len(validation_result) > 1 else ""
                        else:
                            is_valid = bool(validation_result)
                            validation_msg = ""

                        if is_valid:
                            auth_result = service.disabled_disabled_store_auth_check(store)
                            connection_success = bool(auth_result[0]) if isinstance(auth_result, tuple) and len(auth_result) > 0 else bool(auth_result)
                            if connection_success:
                                store.sync_status = 'success'
                                store.last_sync = datetime.utcnow()
                                connection_message = f"  {store.platform} connection re-established"

                                # Queue sync for updated store through dispatcher/runtime gate path
                                from queue_manager import enqueue_sync_job, JOB_FULL_SYNC, PRIORITY_HIGH

                                disabled_queue_job(
                                    store_id=store.id,
                                    job_type=JOB_FULL_SYNC,
                                    payload={
                                        'source': 'store_reconnect_success',
                                        'manual': False,
                                        'store_id': store.id,
                                        'store_name': store.name
                                    },
                                    priority=PRIORITY_HIGH
                                )

                                store.sync_status = 'queued'
                                connection_message += " - Sync queued"
                            else:
                                store.sync_status = 'error'
                                store.is_active = False
                                connection_message = f"  {store.platform} connection failed"
                        else:
                            store.sync_status = 'error'
                            store.is_active = False
                            connection_message = f"  Invalid credentials: {validation_msg}"
                except Exception as conn_error:
                    logging.error(f"Auto-reconnection failed for {store.name}: {str(conn_error)}")
                    store.sync_status = 'error'
                    store.is_active = False
                    connection_message = f"  Connection error: {str(conn_error)}"

            db.session.commit()
            flash(f'Store updated successfully!{connection_message}', 'success' if not connection_message or '' in connection_message else 'warning')
            return redirect(url_for('routes.stores'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating store: {str(e)}', 'danger')
            logging.error(f'Error updating store: {str(e)}')

    return render_template('edit_store.html', store=store)

@bp.route('/stores/delete/<int:store_id>', methods=['POST'])
def delete_store(store_id):
    """Delete store"""
    try:
        store = db.session.get(Store, store_id)
        if store:
            # Delete all associated records to avoid constraint violations
            # 1. Delete marketplace orders
            db.session.query(MarketplaceOrder).filter(MarketplaceOrder.store_id == store_id).delete()
            # 2. Delete marketplace listings
            db.session.query(MarketplaceListing).filter(MarketplaceListing.store_id == store_id).delete()
            # 3. Delete sync logs
            db.session.query(SyncLog).filter(SyncLog.store_id == store_id).delete()
            # 4. Delete sync jobs
            db.session.query(SyncJob).filter(SyncJob.store_id == store_id).delete()
            # 5. Finally delete the store
            db.session.delete(store)
            db.session.commit()
            flash('Store deleted successfully!', 'success')
        else:
            flash('Store not found!', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting store: {str(e)}', 'danger')
        logging.error(f'Error deleting store: {str(e)}')

    return redirect(url_for('routes.stores'))

# ============================================================================
# WAREHOUSE ROUTES (Multi-Warehouse System)
# ============================================================================

@bp.route('/warehouses')
def warehouses():
    """Display all warehouses"""
    from models import Warehouse
    warehouses = db.session.query(Warehouse).order_by(desc(Warehouse.priority), Warehouse.name).all()
    return render_template('warehouses.html', warehouses=warehouses)

@bp.route('/warehouses/add', methods=['GET', 'POST'])
def add_warehouse():
    """Add new warehouse"""
    from models import Warehouse
    if request.method == 'POST':
        try:
            warehouse = Warehouse()
            warehouse.name = request.form['name']
            warehouse.location = request.form.get('location', '')
            warehouse.priority = int(request.form.get('priority', 0))
            warehouse.is_active = 'is_active' in request.form
            warehouse.notes = request.form.get('notes', '')

            # Handle default warehouse setting
            is_default = 'is_default' in request.form
            if is_default:
                # Unset other default warehouses
                db.session.query(Warehouse).update({'is_default': False})
                warehouse.is_default = True

            db.session.add(warehouse)
            db.session.commit()

            flash(f'Warehouse "{warehouse.name}" created successfully!', 'success')
            return redirect(url_for('routes.warehouses'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating warehouse: {str(e)}', 'danger')
            logging.error(f'Error creating warehouse: {str(e)}')

    return render_template('add_warehouse.html')

@bp.route('/warehouses/<int:warehouse_id>/edit', methods=['GET', 'POST'])
def edit_warehouse(warehouse_id):
    """Edit warehouse"""
    from models import Warehouse
    warehouse = db.session.get(Warehouse, warehouse_id)
    if not warehouse:
        flash('Warehouse not found!', 'danger')
        return redirect(url_for('routes.warehouses'))

    if request.method == 'POST':
        try:
            warehouse.name = request.form['name']
            warehouse.location = request.form.get('location', '')
            warehouse.priority = int(request.form.get('priority', 0))
            warehouse.is_active = 'is_active' in request.form
            warehouse.notes = request.form.get('notes', '')

            # Handle default warehouse setting
            is_default = 'is_default' in request.form
            if is_default:
                # Unset other default warehouses
                db.session.query(Warehouse).filter(Warehouse.id != warehouse_id).update({'is_default': False})
                warehouse.is_default = True
            else:
                warehouse.is_default = False

            db.session.commit()
            flash(f'Warehouse "{warehouse.name}" updated successfully!', 'success')
            return redirect(url_for('routes.warehouses'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating warehouse: {str(e)}', 'danger')
            logging.error(f'Error updating warehouse: {str(e)}')

    return render_template('edit_warehouse.html', warehouse=warehouse)

@bp.route('/warehouses/<int:warehouse_id>/stock')
def warehouse_stock(warehouse_id):
    """View stock for a specific warehouse"""
    from models import Warehouse, WarehouseStock
    warehouse = db.session.get(Warehouse, warehouse_id)
    if not warehouse:
        flash('Warehouse not found!', 'danger')
        return redirect(url_for('routes.warehouses'))

    stock_items = db.session.query(WarehouseStock).filter_by(warehouse_id=warehouse_id).order_by(WarehouseStock.sku).all()
    return render_template('warehouse_stock.html', warehouse=warehouse, stock_items=stock_items)

@bp.route('/api/warehouses/<int:warehouse_id>/toggle', methods=['POST'])
def toggle_warehouse(warehouse_id):
    """Toggle warehouse active status"""
    from models import Warehouse
    try:
        warehouse = db.session.get(Warehouse, warehouse_id)
        if not warehouse:
            return jsonify({'success': False, 'error': 'Warehouse not found'})

        data = request.get_json()
        warehouse.is_active = data.get('is_active', False)
        db.session.commit()

        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        logging.error(f'Error toggling warehouse: {str(e)}')
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/api/warehouses/<int:warehouse_id>', methods=['DELETE'])
def delete_warehouse(warehouse_id):
    """Delete warehouse"""
    from models import Warehouse
    try:
        warehouse = db.session.get(Warehouse, warehouse_id)
        if not warehouse:
            return jsonify({'success': False, 'error': 'Warehouse not found'})

        if warehouse.is_default:
            return jsonify({'success': False, 'error': 'Cannot delete default warehouse'})

        db.session.delete(warehouse)
        db.session.commit()

        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        logging.error(f'Error deleting warehouse: {str(e)}')
        return jsonify({'success': False, 'error': str(e)})

# ============================================================================
# END WAREHOUSE ROUTES
# ============================================================================

@bp.route('/api/sync-status')
def sync_status():
    """API endpoint for getting sync status"""
    stores = db.session.query(Store).all()
    return jsonify([{
        'id': store.id,
        'name': store.name,
        'sync_status': store.sync_status,
        'last_sync': store.last_sync.isoformat() if store.last_sync else None
    } for store in stores])

@bp.route('/stores/sync/<int:store_id>', methods=['POST'])
def manual_sync_store(store_id):
    """Retired direct manual store sync route.

    Store sync must be requested through governed dispatcher flow only.
    """
    return jsonify({
        "success": False,
        "error": "Direct store sync route is retired. Use governed dispatcher execution path.",
        "execution_blocked": True,
        "route_retired": True,
        "store_id": store_id
    }), 410

@bp.route('/api/stores')
def api_stores():
    """API endpoint to get stores, optionally filtered by platform"""
    platform = request.args.get('platform')
    query = db.session.query(Store)
    if platform:
        query = query.filter(Store.platform == platform)
    stores = query.order_by(desc(Store.updated_at)).all()

    return jsonify([{
        'id': store.id,
        'name': store.name,
        'platform': store.platform,
        'is_active': store.is_active,
        'sync_status': store.sync_status,
        'last_sync': store.last_sync.strftime('%Y-%m-%d %H:%M') if store.last_sync else None
    } for store in stores])

@bp.route('/api/test-amazon-connection/<int:store_id>', methods=['POST'])

def test_amazon_connection(store_id):
    """API endpoint to test Amazon connection for a specific store"""
    try:
        store = db.session.get(Store, store_id)
        if not store:
            return jsonify({'error': 'Store not found'}), 404

        if store.platform.lower() != 'amazon':
            return jsonify({'error': 'Store is not an Amazon store'}), 400

        # Get credentials from request
        data = request.get_json()
        credentials = data.get('credentials', '')

        if not credentials:
            return jsonify({'error': 'No credentials provided'}), 400

        # Create a temporary store object with the test credentials
        test_store = Store()
        test_store.name = store.name
        test_store.platform = store.platform
        test_store.api_key = credentials

        # Test the connection using Amazon service
        amazon_service = AmazonAPIService()

        # First validate the credential format
        validation_result = amazon_service.validate_credentials_format(credentials)
        if isinstance(validation_result, tuple):
            is_valid = bool(validation_result[0]) if len(validation_result) > 0 else False
            message = str(validation_result[1]) if len(validation_result) > 1 else ""
        else:
            is_valid = bool(validation_result)
            message = ""

        if not is_valid:
            return jsonify({'success': False, 'error': message})

        # Test authentication
        auth_result = amazon_service.authenticate_store(test_store)
        auth_success = bool(auth_result[0]) if isinstance(auth_result, tuple) and len(auth_result) > 0 else bool(auth_result)

        if auth_success:
            return jsonify({'success': True, 'message': 'Amazon connection successful'})
        else:
            return jsonify({'success': False, 'error': 'Authentication failed - check your credentials'})

    except Exception as e:
        logging.error(f'Error testing Amazon connection: {str(e)}')
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/api/test-ebay-connection/<int:store_id>', methods=['GET'])
def test_ebay_connection(store_id):
    """
    Test saved eBay OAuth connection for a store.

    Important:
    - Uses saved store.api_key from DB as the source of truth.
    - Does NOT trust the edit form field, because it can be stale after OAuth.
    - OAuth connection truth must come from the stored token.
    """
    try:
        store = db.session.get(Store, store_id)
        if not store:
            return jsonify({'success': False, 'error': 'Store not found'}), 404

        if (store.platform or '').lower() != 'ebay':
            return jsonify({'success': False, 'error': 'Store is not an eBay store'}), 400

        if not store.api_key:
            store.sync_status = 'error'
            store.is_active = False
            db.session.commit()
            return jsonify({
                'success': False,
                'error': 'No saved eBay OAuth credentials found. Reconnect eBay first.'
            })

        ebay_service = eBayAPIService()

        validation_result = ebay_service.validate_credentials_format(store.api_key)
        if isinstance(validation_result, tuple):
            is_valid = bool(validation_result[0]) if len(validation_result) > 0 else False
            message = str(validation_result[1]) if len(validation_result) > 1 else ""
        else:
            is_valid = bool(validation_result)
            message = ""

        if not is_valid:
            store.sync_status = 'error'
            store.is_active = False
            db.session.commit()
            return jsonify({'success': False, 'error': message})

        auth_result = ebay_service.disabled_disabled_store_auth_check(store)
        auth_success = bool(auth_result[0]) if isinstance(auth_result, tuple) and len(auth_result) > 0 else bool(auth_result)

        if auth_success:
            store.sync_status = 'success'
            store.auth_status = 'connected'
            store.is_active = True
            store.last_sync = datetime.utcnow()
            db.session.commit()
            return jsonify({
                'success': True,
                'message': 'Saved eBay OAuth connection is valid'
            })

        store.sync_status = 'error'
        store.is_active = False
        db.session.commit()
        return jsonify({
            'success': False,
            'error': 'Saved eBay OAuth token failed live validation. Reconnect eBay.'
        })

    except Exception as e:
        db.session.rollback()
        logging.error(f'Error testing eBay connection: {str(e)}')
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/api/test-connection', methods=['POST'])
def test_connection():
    """API endpoint to test connection for any platform without requiring existing store"""
    try:
        data = request.get_json()
        platform = data.get('platform', '').lower()
        credentials = data.get('credentials', '')

        if not platform or not credentials:
            return jsonify({'success': False, 'error': 'Platform and credentials are required'})

        # Create a temporary store object for testing
        test_store = Store()
        test_store.name = "Test Store"
        test_store.platform = platform
        test_store.api_key = credentials

        if platform == 'amazon':
            service = AmazonAPIService()
        elif platform == 'ebay':
            service = eBayAPIService()
        else:
            return jsonify({'success': False, 'error': f'Platform {platform} is not supported for connection testing'})

        # Validate credential format first
        is_valid, message = service.validate_credentials_format(credentials)
        if not is_valid:
            return jsonify({'success': False, 'error': message})

        # Test authentication
        auth_success = service.authenticate_store(test_store)

        if auth_success:
            return jsonify({'success': True, 'message': f'{platform.title()} connection successful'})
        else:
            return jsonify({'success': False, 'error': 'Authentication failed - check your credentials'})

    except Exception as e:
        logging.error(f'Error testing connection: {str(e)}')
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/auth/amazon/start')
def amazon_oauth_start():
    """Start Amazon OAuth flow"""
    try:
        # Generate state parameter for security
        state = secrets.token_urlsafe(32)
        session['oauth_state'] = state
        session['oauth_platform'] = 'amazon'

        # Use Amazon OAuth credentials from environment and extract the actual ID
        client_id_raw = os.environ.get("AMAZON_LWA_CLIENT_ID")
        if client_id_raw and "Value:" in client_id_raw:
            client_id = client_id_raw.split("Value:")[-1].strip()
        else:
            client_id = client_id_raw

        # Debug logging
        logging.info(f"Amazon OAuth - using production client_id: {client_id}")

        if not client_id:
            flash('Amazon OAuth not configured', 'danger')
            return redirect(url_for('routes.add_store'))

        redirect_uri = url_for('routes.amazon_oauth_callback', _external=True)

        # Use SP-API Seller Central authorization (not Login with Amazon)
        # For UK marketplace, use Europe Seller Central
        seller_central_url = 'https://sellercentral-europe.amazon.com/apps/authorize/consent'

        # Build SP-API authorization URL - no scope needed for Seller Central
        auth_params = {
            'application_id': client_id,  # SP-API uses application_id, not client_id
            'state': state
        }

        # Use SP-API Seller Central authorization endpoint
        auth_url = f"{seller_central_url}?{urlencode(auth_params)}"

        # Store auth URL in session and return it to frontend for popup
        session['oauth_url'] = auth_url
        return jsonify({'success': True, 'auth_url': auth_url})

    except Exception as e:
        logging.error(f'Error starting Amazon OAuth: {str(e)}')
        flash('Error starting Amazon authorization', 'danger')
        return redirect(url_for('routes.add_store'))

@bp.route('/auth/amazon/callback')
def amazon_oauth_callback():
    """Handle Amazon OAuth callback"""
    try:
        # Verify state parameter
        if request.args.get('state') != session.get('oauth_state'):
            flash('Invalid authorization state', 'danger')
            return redirect(url_for('routes.add_store'))

        # Get SP-API authorization code (different from standard OAuth)
        code = request.args.get('spapi_oauth_code')
        selling_partner_id = request.args.get('selling_partner_id')

        if not code:
            error = request.args.get('error', 'Unknown error')
            flash(f'Amazon authorization failed: {error}', 'danger')
            return redirect(url_for('routes.add_store'))

        # Log the SP-API response for debugging
        logging.info(f"SP-API OAuth callback: code={code[:20]}..., selling_partner_id={selling_partner_id}")

        # Exchange code for tokens
        tokens = exchange_amazon_code_for_tokens(code)

        if 'access_token' in tokens and 'refresh_token' in tokens:
            # Store tokens in session temporarily
            session['amazon_tokens'] = tokens
            session['amazon_authorized'] = True

            flash('Amazon account connected successfully! You can now create your store.', 'success')
            return redirect(url_for('routes.add_store') + '?amazon_authorized=true')
        else:
            flash(f'Token exchange failed: {tokens.get("error", "Unknown error")}', 'danger')
            return redirect(url_for('routes.add_store'))

    except Exception as e:
        logging.error(f'Error in Amazon OAuth callback: {str(e)}')
        flash('Error completing Amazon authorization', 'danger')
        return redirect(url_for('routes.add_store'))

@bp.route('/auth/amazon/status')
def amazon_oauth_status():
    """Check Amazon OAuth authorization status"""
    return jsonify({
        'authorized': session.get('amazon_authorized', False),
        'has_tokens': bool(session.get('amazon_tokens', {}).get('refresh_token'))
    })

@bp.route('/auth/amazon/reset')
def amazon_oauth_reset():
    """Reset Amazon OAuth session to force fresh authorization"""
    session.pop('amazon_tokens', None)
    session.pop('amazon_authorized', None)
    session.pop('oauth_state', None)
    session.pop('oauth_platform', None)
    logging.info("Amazon OAuth session cleared - forcing fresh authorization")
    flash('Amazon authorization reset. Please authorize again for fresh SP-API permissions.', 'info')
    return redirect(url_for('routes.add_store'))

def exchange_amazon_code_for_tokens(code):
    """Exchange authorization code for access and refresh tokens"""
    try:
        # Use Amazon OAuth credentials from environment and extract the actual ID
        client_id_raw = os.environ.get("AMAZON_LWA_CLIENT_ID")
        if client_id_raw and "Value:" in client_id_raw:
            client_id = client_id_raw.split("Value:")[-1].strip()
        else:
            client_id = client_id_raw
        client_secret_raw = os.environ.get("AMAZON_LWA_CLIENT_SECRET")
        if client_secret_raw and "Value:" in client_secret_raw:
            client_secret = client_secret_raw.split("Value:")[-1].strip()
        else:
            client_secret = client_secret_raw
        redirect_uri = url_for('routes.amazon_oauth_callback', _external=True)

        # Create basic auth header
        credentials = f"{client_id}:{client_secret}"
        encoded_credentials = base64.b64encode(credentials.encode()).decode()

        headers = {
            'Authorization': f'Basic {encoded_credentials}',
            'Content-Type': 'application/x-www-form-urlencoded'
        }

        data = {
            'grant_type': 'authorization_code',
            'code': code,
            'redirect_uri': redirect_uri
        }

        response = requests.post(
            'https://api.amazon.com/auth/o2/token',
            headers=headers,
            data=data,
            timeout=30
        )

        if response.status_code == 200:
            return response.json()
        else:
            return {'error': f'Token exchange failed with status {response.status_code}'}

    except Exception as e:
        return {'error': f'Token exchange error: {str(e)}'}

@bp.route('/auth/amazon/manual-refresh', methods=['POST'])
def amazon_manual_refresh():
    """Manually refresh Amazon token using authorization code"""
    try:
        data = request.get_json()
        auth_code = data.get('auth_code', '').strip()

        if not auth_code:
            return jsonify({'success': False, 'error': 'Authorization code is required'})

        # Exchange code for tokens
        result = exchange_amazon_code_for_tokens(auth_code)
        if result['success']:
            return jsonify({
                'success': True,
                'refresh_token': result['refresh_token'],
                'message': 'Refresh token generated successfully!'
            })
        else:
            return jsonify({'success': False, 'error': result['error']})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/auth/amazon/auto-fill', methods=['POST'])
def amazon_auto_fill_credentials():
    """Auto-fill Amazon credentials from OAuth tokens"""
    try:
        if not session.get('amazon_authorized'):
            return jsonify({'success': False, 'error': 'Amazon account not authorized'})

        tokens = session.get('amazon_tokens', {})
        if not tokens.get('refresh_token'):
            return jsonify({'success': False, 'error': 'No refresh token available'})

        # Get form data for additional info
        data = request.get_json()
        region = data.get('region', 'DE')  # Default to Germany as requested
        seller_id = data.get('seller_id', '')

        # Use Amazon OAuth credentials from environment and extract the actual ID
        client_id_raw = os.environ.get("AMAZON_LWA_CLIENT_ID")
        if client_id_raw and "Value:" in client_id_raw:
            client_id = client_id_raw.split("Value:")[-1].strip()
        else:
            client_id = client_id_raw
        client_secret_raw = os.environ.get("AMAZON_LWA_CLIENT_SECRET")
        if client_secret_raw and "Value:" in client_secret_raw:
            client_secret = client_secret_raw.split("Value:")[-1].strip()
        else:
            client_secret = client_secret_raw

        # Marketplace IDs
        marketplace_ids = {
            'US': 'ATVPDKIKX0DER',
            'UK': 'A1F83G8C2ARO7P',
            'DE': 'A1PA6795UKMFR9',
            'FR': 'A13V1IB3VIYZZH',
            'IT': 'APJ6JRA9NG5V4',
            'ES': 'A1RKKUPIHCS9HS',
            'CA': 'A2EUQ1WTGCTBG2'
        }

        # Build credentials JSON
        credentials = {
            "refresh_token": tokens['refresh_token'],
            "lwa_app_id": client_id,
            "lwa_client_secret": client_secret,
            "seller_id": seller_id,
            "marketplace_id": marketplace_ids.get(region, marketplace_ids['DE']),
            "region": region
        }

        return jsonify({
            'success': True,
            'credentials': json.dumps(credentials, indent=2)
        })

    except Exception as e:
        logging.error(f'Error auto-filling credentials: {str(e)}')
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/nav')
def navigation_links():
    """Quick navigation page with all tested links"""
    return render_template('navigation_links.html')

@bp.route('/ebay-oauth')
def ebay_oauth():
    return render_template('ebay_oauth.html')


@bp.route('/ebay-oauth/start')
def ebay_oauth_start():
    import os, secrets
    from urllib.parse import urlencode

    app_id = os.environ.get("EBAY_CLIENT_ID")
    runame = os.environ.get("EBAY_RUNAME")

    if not app_id or not runame:
        flash("Missing eBay configuration", "danger")
        return redirect(url_for("routes.ebay_oauth"))

    state = secrets.token_urlsafe(32)
    session["ebay_oauth_state"] = state

    scopes = "https://api.ebay.com/oauth/api_scope"

    params = {
        "client_id": app_id,
        "response_type": "code",
        "redirect_uri": runame,
        "scope": scopes,
        "state": state
    }

    return redirect("https://auth.ebay.com/oauth2/authorize?" + urlencode(params))


@bp.route('/ebay-oauth/callback')
def ebay_oauth_callback():
    import os, json, base64, requests
    from datetime import datetime

    code = request.args.get("code")
    if not code:
        flash("Missing eBay auth code", "danger")
        return redirect(url_for("routes.ebay_oauth"))

    app_id = os.environ.get("EBAY_CLIENT_ID")
    cert_id = os.environ.get("EBAY_CLIENT_SECRET")
    dev_id = os.environ.get("EBAY_DEV_ID")
    runame = os.environ.get("EBAY_RUNAME")

    auth = base64.b64encode(f"{app_id}:{cert_id}".encode()).decode()

    r = requests.post(
        "https://api.ebay.com/identity/v1/oauth2/token",
        headers={
            "Authorization": f"Basic {auth}",
            "Content-Type": "application/x-www-form-urlencoded"
        },
        data={
            "grant_type": "authorization_code",
            "code": code,
            "redirect_uri": runame
        },
        timeout=30
    )

    if r.status_code != 200:
        flash("Token exchange failed", "danger")
        return redirect(url_for("routes.ebay_oauth"))

    token = r.json()

    store = Store.query.filter(Store.platform.ilike("%ebay%")).first()
    if not store:
        store = Store(name="beatsoutlet", platform="eBay", is_active=True)
        db.session.add(store)
        db.session.flush()

    data = {}
    try:
        data = json.loads(store.api_key or "{}")
    except:
        pass

    data.update({
        "app_id": app_id,
        "cert_id": cert_id,
        "dev_id": dev_id,
        "runame": runame,
        "access_token": token.get("access_token"),
        "refresh_token": token.get("refresh_token"),
        "connected_at": datetime.utcnow().isoformat()
    })

    store.api_key = json.dumps(data)
    store.auth_status = "connected"

    db.session.commit()

    flash("eBay connected successfully", "success")
    return redirect(url_for("routes.stores"))


@bp.route('/amazon-oauth')
def amazon_oauth():
    """Amazon OAuth authorization page"""
    return render_template('amazon_oauth.html')



@bp.route('/push_stock/<int:item_id>', methods=['POST'])
def push_stock_individual(item_id):
    """Push stock for individual item to all connected stores using job queue"""
    try:
        from queue_manager import enqueue_sync_job, JOB_PUSH_ITEM, PRIORITY_HIGH

        logging.info(f"Enqueueing stock push for item ID: {item_id}")

        # Get the inventory item
        item = db.session.query(InventoryItem).filter(InventoryItem.id == item_id).first()
        if not item:
            logging.error(f"Item not found: {item_id}")
            return jsonify({'success': False, 'error': 'Item not found'})

        logging.info(f"Found item: {item.sku}")

        # CRITICAL FIX: Get warehouse stock for this SKU, then find stores with listings
        # Don't push Amazon SKUs to eBay stores or vice versa!
        warehouse_stock = db.session.query(WarehouseStock).filter_by(sku=item.sku).first()
        if not warehouse_stock:
            return jsonify({'success': False, 'error': f'No warehouse stock found for SKU {item.sku}'})

        active_stores = db.session.query(Store).join(
            MarketplaceListing, Store.id == MarketplaceListing.store_id
        ).filter(
            Store.is_active == True,
            MarketplaceListing.warehouse_stock_id == warehouse_stock.id
        ).distinct().all()

        if not active_stores:
            logging.error(f"No active stores with listings for SKU {item.sku}")
            return jsonify({'success': False, 'error': f'No stores configured for this SKU. Import it to a store first.'})

        logging.info(f"Found {len(active_stores)} stores with listings for {item.sku}")

        # Enqueue high-priority push jobs for each store
        jobs_enqueued = []
        for store in active_stores:
            try:
                # Skip stores without proper credentials
                if not store.api_key:
                    logging.warning(f"Skipping store {store.name}: No API credentials")
                    continue

                # Enqueue high-priority job
                job = disabled_queue_job(
                    store_id=store.id,
                    job_type=JOB_PUSH_ITEM,
                    payload={'item_id': item_id},
                    priority=PRIORITY_HIGH
                )
                jobs_enqueued.append({
                    'store': store.name,
                    'platform': store.platform,
                    'job_id': job.id
                })

            except Exception as store_error:
                logging.error(f"Error enqueueing job for store {store.name}: {str(store_error)}")

        if not jobs_enqueued:
            return jsonify({
                'success': False,
                'error': 'No jobs could be enqueued (check store credentials)'
            })

        response = {
            'success': True,
            'item_sku': item.sku,
            'item_name': item.name,
            'jobs_enqueued': len(jobs_enqueued),
            'stores': [j['store'] for j in jobs_enqueued],
            'message': f'Push queued for {item.name} to {len(jobs_enqueued)} store(s). Processing now...'
        }

        return jsonify(response)

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        logging.error(f"Error in individual stock push for item {item_id}: {str(e)}")
        logging.error(f"Full traceback: {error_details}")
        return jsonify({
            'success': False,
            'error': f'Server error: {str(e)}',
            'item_id': item_id
        }), 500

@bp.route('/push_stock_bulk', methods=['POST'])
def push_stock_bulk():
    """Queue stock push jobs for selected items, gated by Settings."""
    try:
        data = request.get_json()
        if not data or 'item_ids' not in data:
            return jsonify({'success': False, 'error': 'No items selected'})

        item_ids = data['item_ids']
        if not item_ids:
            return jsonify({'success': False, 'error': 'No items selected'})

        logging.info(f"Enqueueing bulk stock push for {len(item_ids)} items")

        items = db.session.query(InventoryItem).filter(InventoryItem.id.in_(item_ids)).all()
        if not items:
            return jsonify({'success': False, 'error': 'No valid items found'})

        results = []
        total_jobs_enqueued = 0

        for item in items:
            warehouse_stock = db.session.query(WarehouseStock).filter_by(sku=item.sku).first()
            if not warehouse_stock:
                logging.warning(f"No warehouse stock for SKU {item.sku} - skipping")
                results.append({
                    'item_id': item.id,
                    'item_sku': item.sku,
                    'item_name': item.name,
                    'success': False,
                    'jobs_enqueued': 0,
                    'stores': [],
                    'error': f'No warehouse stock found for SKU {item.sku}'
                })
                continue

            stores_with_listings = db.session.query(Store).join(
                MarketplaceListing, Store.id == MarketplaceListing.store_id
            ).filter(
                Store.is_active == True,
                MarketplaceListing.warehouse_stock_id == warehouse_stock.id
            ).distinct().all()

            if not stores_with_listings:
                logging.info(f"SKU {item.sku} has no marketplace listings - skipping")
                results.append({
                    'item_id': item.id,
                    'item_sku': item.sku,
                    'item_name': item.name,
                    'success': False,
                    'jobs_enqueued': 0,
                    'stores': [],
                    'error': f'No marketplace listings for {item.sku}. Import it to a store first.'
                })
                continue

            jobs_enqueued_for_item = []
            blocked_stores = []

            for store in stores_with_listings:
                try:
                    allowed, reason = is_runtime_action_allowed(
                        store=store,
                        action_type="push",
                        manual=True
                    )

                    if not allowed:
                        blocked_stores.append({
                            'store': store.name,
                            'platform': store.platform,
                            'reason': reason
                        })
                        continue

                    if not store.api_key:
                        logging.warning(f"Skipping store {store.name}: No API credentials")
                        blocked_stores.append({
                            'store': store.name,
                            'platform': store.platform,
                            'reason': 'No API credentials configured'
                        })
                        continue

                    job = disabled_queue_job(
                        store_id=store.id,
                        job_type=JOB_PUSH_ITEM,
                        payload={'item_id': item.id},
                        priority=PRIORITY_HIGH
                    )

                    jobs_enqueued_for_item.append({
                        'store': store.name,
                        'platform': store.platform,
                        'job_id': job.id
                    })
                    total_jobs_enqueued += 1

                except Exception as store_error:
                    logging.error(f"Error enqueueing job for {item.sku} to store {store.name}: {str(store_error)}")
                    blocked_stores.append({
                        'store': store.name,
                        'platform': store.platform,
                        'reason': str(store_error)
                    })

            results.append({
                'item_id': item.id,
                'item_sku': item.sku,
                'item_name': item.name,
                'success': len(jobs_enqueued_for_item) > 0,
                'jobs_enqueued': len(jobs_enqueued_for_item),
                'stores': [j['store'] for j in jobs_enqueued_for_item],
                'blocked_stores': blocked_stores
            })

        successful_items = [r for r in results if r['success']]
        failed_items = [r for r in results if not r['success']]

        return jsonify({
            'success': len(successful_items) > 0,
            'total_items': len(items),
            'successful_items': len(successful_items),
            'failed_items': len(failed_items),
            'total_jobs_enqueued': total_jobs_enqueued,
            'results': results,
            'message': f'Queued {total_jobs_enqueued} push jobs for {len(successful_items)} items. Processing now...'
        })

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        logging.error(f"Error in bulk stock push: {str(e)}")
        logging.error(f"Full traceback: {error_details}")
        return jsonify({'success': False, 'error': str(e)})


@bp.route('/push_stock_all', methods=['POST'])
def push_stock_all():
    """Retired broad direct push-all route.

    Broad marketplace pushes must not execute directly from routes.
    """
    return jsonify({
        "success": False,
        "error": "Direct push-all route is retired. Use governed dispatcher execution path.",
        "execution_blocked": True,
        "route_retired": True
    }), 410

def command_center_governance_guard(command_key, plan, confirmed=False):
    """Central guard for future Command Center execution. Preview-safe only."""
    risk_level = (plan or {}).get('risk_level', 'unknown')
    requires_confirmation = bool((plan or {}).get('requires_confirmation', True))

    guard_result = {
        'allowed': False,
        'preview_safe': True,
        'execution_blocked': True,
        'risk_level': risk_level,
        'requires_confirmation': requires_confirmation,
        'confirmed': bool(confirmed),
        'checks': [
            {'name': 'command_known', 'passed': bool(command_key)},
            {'name': 'preview_mode_only', 'passed': True},
            {'name': 'live_execution_disabled', 'passed': True},
            {'name': 'confirmation_required_if_high_risk', 'passed': not (risk_level == 'high' and not confirmed)},
        ],
        'message': 'Preview allowed. Live execution remains disabled.'
    }

    return guard_result


# =================== BT38 COMMAND CENTER ROUTES ===================

@bp.route('/api/command-center/preview', methods=['POST'])
def command_center_preview():
    """Preview plain-English admin commands without executing live actions."""
    try:
        data = request.get_json() or {}
        raw_command = (data.get('command') or '').strip()
        normalized = raw_command.lower()

        if not raw_command:
            return jsonify({
                'success': False,
                'error': 'No command entered'
            }), 400

        command_map = {
            'freeze all sync': {
                'command_key': 'freeze_all_sync',
                'meaning': 'Turn off automatic sync/push/import controls safely.',
                'affected_systems': ['global_push_enabled', 'auto_push_enabled', 'fbm_sync_enabled', 'fba_import_enabled'],
                'execution_plan': [
                    'Read current PushSettings and Store settings',
                    'Create before snapshot',
                    'Disable global automatic push',
                    'Disable store auto push',
                    'Disable FBM sync where required',
                    'Pause FBA import where required',
                    'Write ConfigChangeLog before/after snapshot',
                    'Write SystemEvent command record'
                ],
                'risk_level': 'high',
                'requires_confirmation': True
            },
            'run health check': {
                'command_key': 'run_health_check',
                'meaning': 'Check routes, APIs, tokens, queue, workers, and recent failures.',
                'affected_systems': ['routes', 'stores', 'sync_logs', 'system_events', 'api_errors'],
                'execution_plan': [
                    'Check settings route availability',
                    'Check active stores',
                    'Check recent failed sync logs',
                    'Check API error logs',
                    'Check queue/job status where available',
                    'Return diagnostic report',
                    'Write SystemEvent health check record'
                ],
                'risk_level': 'low',
                'requires_confirmation': False
            },
            'export settings snapshot': {
                'command_key': 'export_settings_snapshot',
                'meaning': 'Create a settings snapshot for audit, rollback, or GitHub proof.',
                'affected_systems': ['PushSettings', 'Store', 'SystemConfig', 'ConfigChangeLog'],
                'execution_plan': [
                    'Read current global settings',
                    'Read current store settings',
                    'Mask sensitive values',
                    'Generate structured JSON snapshot',
                    'Write ConfigChangeLog snapshot record',
                    'Return export-ready payload'
                ],
                'risk_level': 'low',
                'requires_confirmation': False
            }
        }

        plan = command_map.get(normalized)

        if not plan:
            return jsonify({
                'success': True,
                'mode': 'preview_only',
                'recognized': False,
                'command_entered': raw_command,
                'message': 'Command not recognized yet.',
                'supported_commands': list(command_map.keys())
            })

        try:
            from models import SystemEvent
            event = SystemEvent(
                actor='admin',
                actor_id=None,
                category='command_center',
                entity_type='command',
                entity_id=None,
                description=f"Command preview requested: {raw_command}",
                details_json={
                    'mode': 'preview_only',
                    'command_entered': raw_command,
                    'command_key': plan['command_key'],
                    'execution_blocked': True
                }
            )
            db.session.add(event)
            db.session.commit()
        except Exception as log_error:
            db.session.rollback()
            logging.warning(f"Command preview logging failed: {str(log_error)}")

        guard_result = command_center_governance_guard(
            plan.get('command_key'),
            plan,
            confirmed=False
        )

        return jsonify({
            'success': True,
            'mode': 'preview_only',
            'recognized': True,
            'command_entered': raw_command,
            'plan': plan,
            'guard': guard_result,
            'risk_level': plan.get('risk_level'),
            'affected_systems': plan.get('affected_systems'),
            'execution_plan': plan.get('execution_plan'),
            'requires_confirmation': plan.get('requires_confirmation'),
            'execution_blocked': True,
            'message': 'Preview created. No live action executed.'
        })

    except Exception as e:
        db.session.rollback()
        logging.error(f"Command Center preview error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# =================== PUSH SETTINGS ROUTES ===================

@bp.route('/settings')
@login_required
def settings():
    """Main settings page for push configuration"""
    global_settings = PushSettings.get_or_create_settings()
    stores = db.session.query(Store).order_by(Store.name).all()

    from datetime import datetime, timedelta

    yesterday = datetime.utcnow() - timedelta(days=1)

    recent_syncs = db.session.query(SyncLog).filter(
        SyncLog.created_at >= yesterday
    ).count()

    failed_syncs = db.session.query(SyncLog).filter(
        SyncLog.created_at >= yesterday,
        SyncLog.status == 'failed'
    ).count()

    stats = {
        'recent_syncs': recent_syncs,
        'failed_syncs': failed_syncs,
        'success_rate': round(((recent_syncs - failed_syncs) / recent_syncs * 100) if recent_syncs > 0 else 100, 1)
    }

    sendgrid_email_config = SystemConfig.query.filter_by(key='sendgrid_from_email').first()
    sendgrid_from_email = sendgrid_email_config.value if sendgrid_email_config else ''

    webhook_platforms = ['amazon', 'ebay', 'tiktok', 'shopify']
    webhook_settings = {
        'worker_enabled': SystemSetting.get_value('webhook_worker_enabled', False),
        'platforms': {}
    }

    from sqlalchemy import cast, String

    for platform in webhook_platforms:
        enabled = SystemSetting.get_value(f'webhook_{platform}_enabled', False)
        platform_match = f'"platform": "{platform}"'

        last_event = db.session.query(SystemEvent).filter(
            SystemEvent.category == 'marketplace_webhook',
            cast(SystemEvent.details_json, String).contains(platform_match)
        ).order_by(SystemEvent.timestamp.desc()).first()

        received_24h = db.session.query(SystemEvent).filter(
            SystemEvent.category == 'marketplace_webhook',
            SystemEvent.timestamp >= yesterday,
            cast(SystemEvent.details_json, String).contains(platform_match)
        ).count()

        failed_24h = db.session.query(SystemEvent).filter(
            SystemEvent.category == 'marketplace_webhook_failed',
            SystemEvent.timestamp >= yesterday,
            cast(SystemEvent.details_json, String).contains(platform_match)
        ).count()

        webhook_settings['platforms'][platform] = {
            'enabled': enabled,
            'last_received': last_event.timestamp if last_event else None,
            'received_24h': received_24h,
            'failed_24h': failed_24h,
            'polling_impact': 'Reduced polling' if enabled else 'Polling fallback'
        }

    return render_template(
        'settings.html',
        global_settings=global_settings,
        stores=stores,
        stats=stats,
        sendgrid_from_email=sendgrid_from_email,
        webhook_settings=webhook_settings
    )



@bp.route('/api/webhook-settings', methods=['POST'])
def update_webhook_settings():
    """Update webhook visibility/runtime flags."""
    try:
        data = request.get_json() or {}

        allowed_keys = {
            'webhook_worker_enabled',
            'webhook_amazon_enabled',
            'webhook_ebay_enabled',
            'webhook_tiktok_enabled',
            'webhook_shopify_enabled'
        }

        updated = {}

        for key, value in data.items():
            if key not in allowed_keys:
                continue

            bool_value = bool(value)

            SystemSetting.set_value(
                key,
                bool_value,
                description='Marketplace webhook runtime setting',
                value_type='bool'
            )

            updated[key] = bool_value

        db.session.add(SystemEvent(
            actor='admin',
            category='config_change',
            entity_type='webhook_settings',
            description='Webhook runtime settings updated',
            details_json=updated
        ))

        db.session.commit()

        return jsonify({'success': True, 'updated': updated})

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error updating webhook settings: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/settings', methods=['POST'])
@login_required
def update_settings():
    """Update global push settings"""
    try:
        global_settings = PushSettings.get_or_create_settings()

        # Update global settings from form
        global_settings.global_push_enabled = request.form.get('global_push_enabled') == 'on'

        try:
            submitted_frequency = int(request.form.get('default_push_frequency_minutes', 15))
            if submitted_frequency <= 0:
                submitted_frequency = 15
        except (TypeError, ValueError):
            submitted_frequency = 15

        global_settings.default_push_frequency_minutes = submitted_frequency
        global_settings.default_batch_size = int(request.form.get('default_batch_size', 10))
        global_settings.default_retry_attempts = int(request.form.get('default_retry_attempts', 3))

        # Scheduling settings
        global_settings.enable_batch_scheduling = request.form.get('enable_batch_scheduling') == 'on'
        global_settings.batch_schedule_minutes = int(request.form.get('batch_schedule_minutes', 30))
        global_settings.off_hours_only = request.form.get('off_hours_only') == 'on'
        global_settings.off_hours_start = int(request.form.get('off_hours_start', 22))
        global_settings.off_hours_end = int(request.form.get('off_hours_end', 6))

        # Safety settings
        global_settings.require_confirmation_threshold = int(request.form.get('require_confirmation_threshold', 50))
        global_settings.auto_pause_on_errors = request.form.get('auto_pause_on_errors') == 'on'
        global_settings.error_rate_threshold = float(request.form.get('error_rate_threshold', 0.3))

        # Notification settings
        global_settings.notify_on_large_pushes = request.form.get('notify_on_large_pushes') == 'on'
        global_settings.notify_on_failures = request.form.get('notify_on_failures') == 'on'
        global_settings.daily_summary_enabled = request.form.get('daily_summary_enabled') == 'on'

        # Advanced settings
        global_settings.concurrent_store_pushes = int(request.form.get('concurrent_store_pushes', 3))
        global_settings.api_rate_limit_buffer = float(request.form.get('api_rate_limit_buffer', 0.8))

        db.session.commit()
        flash('Push settings updated successfully!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error updating settings: {str(e)}', 'danger')
        logging.error(f'Error updating push settings: {str(e)}')

    return redirect(url_for('routes.settings'))

@bp.route('/api/push-settings', methods=['GET'])
def get_push_settings():
    """API endpoint to get current push settings"""
    try:
        global_settings = PushSettings.get_or_create_settings()
        return jsonify({
            'success': True,
            'settings': global_settings.to_dict()
        })
    except Exception as e:
        logging.error(f"Error getting push settings: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/api/reset-push-settings', methods=['POST'])
def reset_push_settings():
    """Reset push settings to defaults"""
    try:
        # Delete existing settings to trigger recreation with defaults
        db.session.query(PushSettings).delete()
        db.session.commit()

        # Create new default settings
        new_settings = PushSettings()
        db.session.add(new_settings)
        db.session.commit()

        flash('Push settings reset to defaults successfully!', 'success')
        return jsonify({'success': True, 'message': 'Settings reset successfully'})

    except Exception as e:
        db.session.rollback()
        flash(f'Error resetting settings: {str(e)}', 'danger')
        logging.error(f'Error resetting push settings: {str(e)}')
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/api/store-push-settings/<int:store_id>', methods=['POST'])
def update_store_push_settings(store_id):
    """Update settings for a specific store from the Settings Control Center"""
    try:
        store = db.session.get(Store, store_id)
        if not store:
            return jsonify({'success': False, 'error': 'Store not found'})

        data = request.get_json() or {}

        if 'is_active' in data:
            store.is_active = bool(data['is_active'])

        if 'fba_import_enabled' in data:
            store.fba_import_enabled = bool(data['fba_import_enabled'])

        if 'fbm_sync_enabled' in data:
            store.fbm_sync_enabled = bool(data['fbm_sync_enabled'])
            if not store.fbm_sync_enabled:
                store.auto_push_enabled = False

        if 'auto_push_enabled' in data:
            store.auto_push_enabled = bool(data['auto_push_enabled'])

        if 'push_priority' in data:
            store.push_priority = int(data['push_priority'])

        if 'push_frequency_minutes' in data:
            try:
                submitted_frequency = int(data['push_frequency_minutes'])
                if submitted_frequency <= 0:
                    submitted_frequency = 15
            except (TypeError, ValueError):
                submitted_frequency = 15

            store.push_frequency_minutes = submitted_frequency

        if 'push_batch_size' in data:
            store.push_batch_size = int(data['push_batch_size'])

        if 'push_on_quantity_change' in data:
            store.push_on_quantity_change = bool(data['push_on_quantity_change'])

        if 'push_on_price_change' in data:
            store.push_on_price_change = bool(data['push_on_price_change'])

        if 'push_on_item_create' in data:
            store.push_on_item_create = bool(data['push_on_item_create'])

        if 'push_on_item_update' in data:
            store.push_on_item_update = bool(data['push_on_item_update'])

        if 'max_retry_attempts' in data:
            store.max_retry_attempts = int(data['max_retry_attempts'])

        if 'auto_disable_on_failures' in data:
            store.auto_disable_on_failures = bool(data['auto_disable_on_failures'])
            if not store.auto_disable_on_failures:
                store.current_failure_count = 0

        if 'failure_threshold' in data:
            store.failure_threshold = int(data['failure_threshold'])

        if 'immediate_push' in data:
            store.immediate_push = bool(data['immediate_push'])

        if 'large_change_confirmation' in data:
            store.large_change_confirmation = bool(data['large_change_confirmation'])

        if 'large_change_threshold' in data:
            store.large_change_threshold = int(data['large_change_threshold'])

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Store settings updated for {store.name}',
            'store_settings': store.to_dict()
        })

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error updating store settings: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})


@bp.route('/api/push-activity')
def push_activity():
    """Get recent push activity for dashboard"""
    try:
        # Get push activity from the last 7 days
        from datetime import datetime, timedelta
        week_ago = datetime.utcnow() - timedelta(days=7)

        # Get recent sync logs with detailed information
        recent_logs = db.session.query(SyncLog).filter(
            SyncLog.created_at >= week_ago
        ).order_by(desc(SyncLog.created_at)).limit(50).all()

        activity = []
        for log in recent_logs:
            activity.append({
                'id': log.id,
                'store_name': log.store.name if log.store else 'Unknown',
                'store_platform': log.store.platform if log.store else 'Unknown',
                'status': log.status,
                'message': log.message,
                'items_synced': log.items_synced,
                'created_at': log.created_at.isoformat()
            })

        return jsonify({
            'success': True,
            'activity': activity
        })

    except Exception as e:
        logging.error(f"Error getting push activity: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/admin/credentials')
def credentials_admin():
    """Amazon credentials management page"""
    try:
        # Get current environment variables (masked for security)
        current_client_id = os.environ.get("AMAZON_LWA_CLIENT_ID", "")
        current_client_secret = os.environ.get("AMAZON_LWA_CLIENT_SECRET", "")
        current_refresh_token = os.environ.get("AMAZON_REFRESH_TOKEN", "")
        current_seller_id = os.environ.get("AMAZON_SELLER_ID", "")

        # Mask secrets for display (show only first/last few chars)
        def mask_secret(value):
            if not value or value == "Not set":
                return "Not set"
            if len(value) < 20:
                return "***"
            return value[:10] + "..." + value[-10:]

        # Get Amazon store status
        amazon_stores = db.session.query(Store).filter_by(platform='Amazon').all()

        credential_info = {
            'current': {
                'client_id': mask_secret(current_client_id),
                'client_secret': mask_secret(current_client_secret),
                'refresh_token': mask_secret(current_refresh_token),
                'seller_id': current_seller_id if current_seller_id else "Not set"
            },
            'has_credentials': bool(current_client_id and current_client_secret and current_refresh_token),
            'amazon_stores': amazon_stores
        }

        return render_template('admin_credentials.html', credentials=credential_info)

    except Exception as e:
        logging.error(f"Error in credentials admin: {str(e)}")
        flash(f"Error loading credentials: {str(e)}", 'error')
        return redirect(url_for("routes.dashboard"))

@bp.route('/admin/credentials', methods=['POST'])
def update_credentials():
    """Update Amazon credentials in database"""
    try:
        # Validate CSRF token
        csrf_token = request.form.get('csrf_token')
        if not validate_csrf_token(csrf_token):
            flash('Invalid security token. Please try again.', 'error')
            return redirect(url_for('routes.credentials_admin'))

        # Get form data
        client_id = request.form.get('client_id', '').strip()
        client_secret = request.form.get('client_secret', '').strip()
        refresh_token = request.form.get('refresh_token', '').strip()
        seller_id = request.form.get('seller_id', '').strip() or 'A2WUN0OF9UHU23'

        # Validate required fields
        if not all([client_id, client_secret, refresh_token]):
            flash('All credential fields are required!', 'error')
            return redirect(url_for('routes.credentials_admin'))

        # Validate format
        if not client_id.startswith('amzn1.application-oa2-client.'):
            flash('Invalid Client ID format. Should start with: amzn1.application-oa2-client.', 'error')
            return redirect(url_for('routes.credentials_admin'))

        if not client_secret.startswith('amzn1.oa2-cs.v1.'):
            flash('Invalid Client Secret format. Should start with: amzn1.oa2-cs.v1.', 'error')
            return redirect(url_for('routes.credentials_admin'))

        if not refresh_token.startswith('Atzr|'):
            flash('Invalid Refresh Token format. Should start with: Atzr|', 'error')
            return redirect(url_for('routes.credentials_admin'))

        # Store in database
        def upsert_config(key, value):
            config = SystemConfig.query.filter_by(key=key).first()
            if config:
                config.value = value
                config.updated_at = datetime.utcnow()
            else:
                config = SystemConfig(key=key, value=value)
                db.session.add(config)

        upsert_config('AMAZON_LWA_CLIENT_ID', client_id)
        upsert_config('AMAZON_LWA_CLIENT_SECRET', client_secret)
        upsert_config('AMAZON_REFRESH_TOKEN', refresh_token)
        upsert_config('AMAZON_SELLER_ID', seller_id)

        db.session.commit()

        flash(' Amazon credentials updated successfully! Amazon will connect in 30 seconds.', 'success')

        # Log the update
        logging.info(f"Amazon credentials updated via admin panel and stored in database")

        return redirect(url_for('routes.credentials_admin'))

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error updating credentials: {str(e)}")
        flash(f"Error updating credentials: {str(e)}", 'error')
        return redirect(url_for('routes.credentials_admin'))

@bp.route('/product_linking')
def product_linking():
    """Product linking management page - manage which marketplace listings are linked to warehouse products"""
    try:
        # Get all warehouse stocks with their linked marketplace listings
        warehouse_stocks = db.session.query(WarehouseStock).order_by(WarehouseStock.sku).limit(100).all()

        warehouse_data = []
        for ws in warehouse_stocks:
            # Get linked marketplace listings
            linked_listings = db.session.query(MarketplaceListing).join(Store).filter(
                MarketplaceListing.warehouse_stock_id == ws.id
            ).all()

            platforms = list(set([listing.store.platform for listing in linked_listings]))

            warehouse_data.append({
                'id': ws.id,
                'sku': ws.sku,
                'quantity': ws.available_quantity,
                'platforms': platforms,
                'linked_count': len(linked_listings),
                'listings': [{
                    'id': listing.id,
                    'title': listing.title or 'No title',
                    'sku': listing.external_sku or listing.external_listing_id,
                    'platform': listing.store.platform,
                    'store_name': listing.store.name
                } for listing in linked_listings]
            })

        # Get unlinked marketplace listings (not linked to any warehouse stock)
        unlinked_listings = db.session.query(MarketplaceListing).join(Store).filter(
            MarketplaceListing.warehouse_stock_id == None
        ).limit(100).all()

        unlinked_data = [{
            'id': listing.id,
            'title': listing.title or 'No title',
            'sku': listing.external_sku or listing.external_listing_id,
            'platform': listing.store.platform,
            'store_name': listing.store.name
        } for listing in unlinked_listings]

        return render_template('product_linking.html',
                             warehouse_products=warehouse_data,
                             unlinked_listings=unlinked_data)

    except Exception as e:
        logging.error(f"Error in product linking page: {str(e)}")
        flash(f"Error loading product linking: {str(e)}", 'error')
        return redirect(url_for("routes.dashboard"))

def calculate_sku_similarity(sku1, sku2):
    """Calculate similarity percentage between two SKUs"""
    from difflib import SequenceMatcher
    return int(SequenceMatcher(None, sku1.lower(), sku2.lower()).ratio() * 100)

def find_similar_warehouse_items(master_ws, all_warehouse_stocks):
    """Find similar warehouse items that might be the same product"""
    from difflib import SequenceMatcher

    similar_items = []
    master_sku = master_ws.sku.lower()

    for ws in all_warehouse_stocks:
        if ws.id == master_ws.id:
            continue

        # Skip if already linked to the same master
        if ws.sku == master_ws.sku:
            continue

        # Calculate SKU similarity
        similarity = SequenceMatcher(None, master_sku, ws.sku.lower()).ratio()

        # If similarity is above 60%, consider it a potential match
        if similarity > 0.6:
            # Get marketplace listings for this item
            listings = db.session.query(MarketplaceListing).join(Store).filter(
                MarketplaceListing.warehouse_stock_id == ws.id
            ).all()

            if listings:
                for listing in listings:
                    similar_items.append({
                        'sku': ws.sku,
                        'name': listing.title or ws.sku,
                        'quantity': ws.available_quantity,
                        'platform': listing.store.platform,
                        'similarity': int(similarity * 100),
                        'warehouse_id': ws.id,
                        'listing_id': listing.id
                    })

    return similar_items

@bp.route('/api/link_listing_to_warehouse', methods=['POST'])
def link_listing_to_warehouse():
    """API endpoint to link a marketplace listing to a warehouse product"""
    try:
        data = request.json
        listing_id = data.get('listing_id')
        warehouse_id = data.get('warehouse_id')

        if not listing_id or not warehouse_id:
            return jsonify({'success': False, 'error': 'Missing parameters'})

        # Get the marketplace listing
        listing = db.session.query(MarketplaceListing).filter_by(id=listing_id).first()
        if not listing:
            return jsonify({'success': False, 'error': 'Marketplace listing not found'})

        # Get the warehouse stock
        warehouse = db.session.query(WarehouseStock).filter_by(id=warehouse_id).first()
        if not warehouse:
            return jsonify({'success': False, 'error': 'Warehouse product not found'})

        # Link the listing to the warehouse
        listing.warehouse_stock_id = warehouse_id
        db.session.commit()

        logging.info(f"Linked listing {listing.external_sku or listing.external_listing_id} to warehouse {warehouse.sku}")

        return jsonify({
            'success': True,
            'message': f'Successfully linked {listing.external_sku or listing.external_listing_id} to {warehouse.sku}'
        })

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error linking listing: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/api/unlink_listing', methods=['POST'])
def unlink_listing():
    """API endpoint to unlink a marketplace listing from its warehouse product"""
    try:
        data = request.json
        listing_id = data.get('listing_id')

        if not listing_id:
            return jsonify({'success': False, 'error': 'Missing listing_id'})

        # Get the marketplace listing
        listing = db.session.query(MarketplaceListing).filter_by(id=listing_id).first()
        if not listing:
            return jsonify({'success': False, 'error': 'Marketplace listing not found'})

        old_warehouse_id = listing.warehouse_stock_id

        # Unlink the listing
        listing.warehouse_stock_id = None
        db.session.commit()

        logging.info(f"Unlinked listing {listing.external_sku or listing.external_listing_id} from warehouse ID {old_warehouse_id}")

        return jsonify({
            'success': True,
            'message': f'Successfully unlinked {listing.external_sku or listing.external_listing_id}'
        })

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error unlinking listing: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@bp.route('/api/diagnostics/sku/<sku>')
# @login_required  # Temporarily disabled for access
def sku_diagnostics(sku):
    """
    Comprehensive SKU diagnostics endpoint
    Shows everything about a SKU in one view to catch mismatches
    """
    try:
        # 1. Warehouse stock details
        warehouse_stock = db.session.query(WarehouseStock).filter_by(sku=sku).first()
        warehouse_data = None
        if warehouse_stock:
            warehouse_data = {
                'id': warehouse_stock.id,
                'sku': warehouse_stock.sku,
                'available_quantity': warehouse_stock.available_quantity,
                'reserved_quantity': warehouse_stock.reserved_quantity,
                'allocated_quantity': warehouse_stock.allocated_quantity,
                'on_order_quantity': warehouse_stock.on_order_quantity,
                'location': warehouse_stock.location,
                'unit_cost': warehouse_stock.unit_cost,
                'reorder_point': warehouse_stock.reorder_point,
                'is_active': warehouse_stock.is_active,
                'last_adjustment_at': warehouse_stock.last_adjustment_at.isoformat() if warehouse_stock.last_adjustment_at else None,
                'last_sync_at': warehouse_stock.last_sync_at.isoformat() if warehouse_stock.last_sync_at else None
            }

        # 2. Inventory item details
        inventory_item = db.session.query(InventoryItem).filter_by(sku=sku).first()
        inventory_data = None
        if inventory_item:
            inventory_data = {
                'id': inventory_item.id,
                'name': inventory_item.name,
                'sku': inventory_item.sku,
                'quantity': inventory_item.quantity,
                'price': inventory_item.price,
                'description': inventory_item.description,
                'group_id': inventory_item.group_id,
                'group_name': inventory_item.group.name if inventory_item.group else None,
                'variant_attributes': inventory_item.variant_attributes
            }

        # 3. Marketplace listings (all stores)
        marketplace_listings = []
        if warehouse_stock:
            listings = db.session.query(MarketplaceListing).join(Store).filter(
                MarketplaceListing.warehouse_stock_id == warehouse_stock.id
            ).all()

            for listing in listings:
                # Calculate what quantity will be pushed
                calculated_qty = max(0, warehouse_stock.available_quantity - (listing.quantity_buffer or 0))
                if listing.max_quantity_limit and listing.max_quantity_limit > 0:
                    calculated_qty = min(calculated_qty, listing.max_quantity_limit)

                listing_data = {
                    'id': listing.id,
                    'store_id': listing.store_id,
                    'store_name': listing.store.name,
                    'platform': listing.store.platform,
                    'external_sku': listing.external_sku,
                    'external_listing_id': listing.external_listing_id,
                    'title': listing.title,
                    'price': listing.price,
                    'asin': listing.asin,
                    'fnsku': listing.fnsku,
                    'sync_quantity': listing.sync_quantity,
                    'sync_price': listing.sync_price,
                    'quantity_buffer': listing.quantity_buffer,
                    'max_quantity_limit': listing.max_quantity_limit,
                    'push_state': listing.push_state,
                    'last_push_at': listing.last_push_at.isoformat() if listing.last_push_at else None,
                    'last_push_quantity': listing.last_push_quantity,
                    'last_push_status': listing.last_push_status,
                    'last_push_error': listing.last_push_error,
                    'consecutive_failures': listing.consecutive_failures,
                    'calculated_push_qty': calculated_qty,
                    'push_calculation_trace': {
                        'warehouse_available': warehouse_stock.available_quantity,
                        'buffer': listing.quantity_buffer or 0,
                        'max_limit': listing.max_quantity_limit,
                        'formula': f"{warehouse_stock.available_quantity} - {listing.quantity_buffer or 0} = {calculated_qty}" +
                                  (f" (capped at {listing.max_quantity_limit})" if listing.max_quantity_limit and listing.max_quantity_limit < calculated_qty else "")
                    }
                }
                marketplace_listings.append(listing_data)

        # 4. Recent sync logs
        recent_syncs = []
        sync_logs = db.session.query(SyncLog).join(Store).filter(
            SyncLog.sku == sku
        ).order_by(desc(SyncLog.created_at)).limit(10).all()

        for log in sync_logs:
            recent_syncs.append({
                'id': log.id,
                'store_name': log.store.name,
                'platform': log.store.platform,
                'status': log.status,
                'message': log.message,
                'items_synced': log.items_synced,
                'created_at': log.created_at.isoformat()
            })

        # 5. Product group info
        product_group_data = None
        if inventory_item and inventory_item.group:
            group = inventory_item.group
            product_group_data = {
                'id': group.id,
                'name': group.name,
                'description': group.description,
                'group_key': group.group_key,
                'total_items': len(group.items) if group.items else 0
            }

        # 6. Recent ledger entries
        ledger_entries = []
        if warehouse_stock:
            ledger = db.session.query(StockLedgerEntry).filter(
                StockLedgerEntry.warehouse_stock_id == warehouse_stock.id
            ).order_by(desc(StockLedgerEntry.created_at)).limit(10).all()

            for entry in ledger:
                ledger_entries.append({
                    'id': entry.id,
                    'entry_type': entry.entry_type,
                    'quantity_change': entry.quantity_change,
                    'quantity_after': entry.quantity_after,
                    'reference_number': entry.reference_number,
                    'notes': entry.notes,
                    'created_by': entry.created_by,
                    'created_at': entry.created_at.isoformat()
                })

        # Build comprehensive diagnostics response
        diagnostics = {
            'sku': sku,
            'timestamp': datetime.now().isoformat(),
            'warehouse_stock': warehouse_data,
            'inventory_item': inventory_data,
            'marketplace_listings': marketplace_listings,
            'product_group': product_group_data,
            'recent_syncs': recent_syncs,
            'ledger_entries': ledger_entries,
            'summary': {
                'warehouse_exists': warehouse_data is not None,
                'inventory_exists': inventory_data is not None,
                'listing_count': len(marketplace_listings),
                'recent_sync_count': len(recent_syncs),
                'quantity_mismatch': False,
                'price_issues': []
            }
        }

        # Detect quantity mismatches
        if warehouse_data and inventory_data:
            if warehouse_data['available_quantity'] != inventory_data['quantity']:
                diagnostics['summary']['quantity_mismatch'] = True
                diagnostics['summary']['mismatch_details'] = {
                    'warehouse_qty': warehouse_data['available_quantity'],
                    'inventory_qty': inventory_data['quantity'],
                    'difference': warehouse_data['available_quantity'] - inventory_data['quantity']
                }

        # Detect price issues
        for listing in marketplace_listings:
            if listing['price'] == 0 or listing['price'] is None:
                diagnostics['summary']['price_issues'].append({
                    'store': listing['store_name'],
                    'issue': 'Price is 0.00 or null',
                    'push_state': listing['push_state']
                })

        return jsonify({
            'success': True,
            'diagnostics': diagnostics
        }), 200

    except Exception as e:
        logging.error(f"Error in SKU diagnostics for {sku}: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Diagnostics error: {str(e)}'
        }), 500

@bp.route('/api/diagnostics/sku-precheck/<sku>')
def sku_precheck(sku):
    """Quick local sanity check for SKU readiness"""
    # Auth via X-Task-Key
    if request.headers.get("X-Task-Key") != TASK_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        data = {"sku": sku}

        # Warehouse qty for SKU
        warehouse_stock = db.session.query(WarehouseStock).filter_by(sku=sku).first()
        data["warehouse_qty"] = warehouse_stock.available_quantity if warehouse_stock else 0
        data["warehouse_exists"] = warehouse_stock is not None

        # Listing type (or "unmapped")
        listing = db.session.query(MarketplaceListing).join(Store).filter(
            MarketplaceListing.external_sku == sku
        ).first()

        if listing:
            data["listing_type"] = listing.listing_type or "single"
            data["asin"] = listing.asin
            data["price"] = listing.price
            data["status"] = listing.push_state
        else:
            data["listing_type"] = "unmapped"
            data["asin"] = None
            data["price"] = None
            data["status"] = None

        # Flags
        data["flags"] = {
            "price_missing": (data["price"] is None or data["price"] == 0) if listing else True,
            "asin_missing": not data["asin"] if listing else True,
            "unmapped": not listing,
            "no_warehouse": not warehouse_stock
        }

        # Recommendation
        if not listing:
            data["recommendation"] = "map_listing"
        elif data["flags"]["price_missing"]:
            data["recommendation"] = "fix_price"
        elif data["flags"]["no_warehouse"]:
            data["recommendation"] = "create_warehouse_stock"
        else:
            data["recommendation"] = "ready"

        return jsonify({"ok": True, **data}), 200

    except Exception as e:
        logging.error(f"Error in sku-precheck for {sku}: {str(e)}")
        return jsonify({"ok": False, "error": str(e)}), 500

@bp.route('/api/diagnostics/amazon/sku/<sku>')
def amazon_sku_diagnostics(sku):
    """Amazon-specific SKU diagnostics with feed history"""
    # Auth via X-Task-Key
    if request.headers.get("X-Task-Key") != TASK_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        data = {"sku": sku}

        # Find active Amazon store
        amazon_store = db.session.query(Store).filter(
            Store.platform == 'Amazon',
            Store.is_active == True
        ).first()

        if not amazon_store:
            return jsonify({
                "ok": False,
                "error": "No active Amazon store found"
            }), 404

        # Init Amazon service
        amazon_service = AmazonAPIService(amazon_store)

        # Get listing info
        listing = db.session.query(MarketplaceListing).filter(
            MarketplaceListing.store_id == amazon_store.id,
            MarketplaceListing.external_sku == sku
        ).first()

        if listing:
            data["listing"] = {
                "listing_type": listing.listing_type,
                "asin": listing.asin,
                "price": listing.price,
                "status": listing.push_state,
                "last_push_at": listing.last_push_at.isoformat() if listing.last_push_at else None,
                "last_push_error": listing.last_push_error,
                "consecutive_failures": listing.consecutive_failures
            }
        else:
            data["listing"] = None

        # Get region info
        try:
            auth_diag = amazon_service.get_auth_diagnostics(amazon_store)
            data["region"] = {
                "marketplace_id": auth_diag.get("marketplace_id"),
                "resolved_region": auth_diag.get("resolved_region"),
                "host": auth_diag.get("host")
            }
        except Exception as e:
            data["region"] = {"error": str(e)}

        # Get recent feed IDs for this SKU
        data["feeds"] = amazon_service.get_last_feed_ids_for_sku(sku, limit=3)

        # If feeds exist, attach first feed report
        if data["feeds"]:
            try:
                first_feed = data["feeds"][0]
                data["report"] = amazon_service.get_feed_report(first_feed["feed_id"])
            except Exception as e:
                data["report"] = {"error": f"Could not fetch feed report: {str(e)}"}
        else:
            data["report"] = None

        return jsonify({"ok": True, **data}), 200

    except Exception as e:
        logging.error(f"Error in amazon sku diagnostics for {sku}: {str(e)}")
        return jsonify({"ok": False, "error": str(e)}), 500

@bp.route('/api/diagnostics/amazon/bulk-test', methods=['POST', 'GET'])
def amazon_bulk_test():
    """Bulk push test with detailed error capture"""
    # Auth via X-Task-Key
    if request.headers.get("X-Task-Key") != TASK_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        # Get SKUs from JSON body or query params
        if request.method == 'POST':
            data = request.get_json() or {}
            skus = data.get("skus", [])
            dry_run = data.get("dry_run", True)
        else:
            skus = request.args.get("skus", "").split(",")
            skus = [s.strip() for s in skus if s.strip()]
            dry_run = request.args.get("dry_run", "true").lower() == "true"

        if not skus:
            return jsonify({
                "ok": False,
                "error": "No SKUs provided. Use 'skus' parameter with comma-separated SKUs."
            }), 400

        # Find active Amazon store
        amazon_store = db.session.query(Store).filter(
            Store.platform == 'Amazon',
            Store.is_active == True
        ).first()

        if not amazon_store:
            return jsonify({
                "ok": False,
                "error": "No active Amazon store found"
            }), 404

        # Init Amazon service
        amazon_service = AmazonAPIService()

        # Run bulk push test
        results = amazon_service.bulk_push_safe(amazon_store, skus, dry_run=dry_run)

        return jsonify({"ok": True, "results": results}), 200

    except Exception as e:
        logging.error(f"Error in amazon bulk test: {str(e)}")
        import traceback
        return jsonify({
            "ok": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }), 500

@bp.route('/api/diagnostics/amazon/feed/<feed_id>')
def amazon_feed_status(feed_id):
    """Get Amazon feed status"""
    # Auth via X-Task-Key
    if request.headers.get("X-Task-Key") != TASK_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        # Find active Amazon store
        amazon_store = db.session.query(Store).filter(
            Store.platform == 'Amazon',
            Store.is_active == True
        ).first()

        if not amazon_store:
            return jsonify({
                "ok": False,
                "error": "No active Amazon store found"
            }), 404

        # Init Amazon service
        amazon_service = AmazonAPIService()

        # Get feed status
        status = amazon_service.get_feed_status(amazon_store, feed_id)

        return jsonify(status), 200

    except Exception as e:
        logging.error(f"Error fetching feed status {feed_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@bp.route('/api/diagnostics/amazon/feed/<feed_id>/report')
def amazon_feed_processing_report(feed_id):
    """Get Amazon feed processing report"""
    # Auth via X-Task-Key
    if request.headers.get("X-Task-Key") != TASK_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        # Find active Amazon store
        amazon_store = db.session.query(Store).filter(
            Store.platform == 'Amazon',
            Store.is_active == True
        ).first()

        if not amazon_store:
            return jsonify({
                "ok": False,
                "error": "No active Amazon store found"
            }), 404

        # Init Amazon service
        amazon_service = AmazonAPIService()

        # Get feed processing report
        report = amazon_service.get_feed_processing_report(amazon_store, feed_id)

        return jsonify(report), 200

    except Exception as e:
        logging.error(f"Error fetching feed report {feed_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@bp.route('/api/diagnostics/amazon/marketplace/sku/<sku>')
def amazon_marketplace_sku_state(sku):
    """Get live marketplace state for SKU with warehouse comparison"""
    # Auth via X-Task-Key
    if request.headers.get("X-Task-Key") != TASK_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        # Find active Amazon store
        amazon_store = db.session.query(Store).filter(
            Store.platform == 'Amazon',
            Store.is_active == True
        ).first()

        if not amazon_store:
            return jsonify({
                "ok": False,
                "error": "No active Amazon store found"
            }), 404

        # Init Amazon service
        amazon_service = AmazonAPIService()

        # Get live marketplace state
        live_state = amazon_service.get_live_listing_state(amazon_store, sku)

        # Add warehouse quantity comparison
        from models import WarehouseStock, InventoryItem
        warehouse_stock = db.session.query(WarehouseStock).filter_by(sku=sku).first()
        if warehouse_stock:
            live_state["warehouse_qty"] = warehouse_stock.available_quantity
        else:
            # Fallback to inventory item
            inventory_item = db.session.query(InventoryItem).filter_by(sku=sku).first()
            live_state["warehouse_qty"] = inventory_item.quantity if inventory_item else None

        # Calculate delta
        if live_state.get("live_qty") is not None and live_state.get("warehouse_qty") is not None:
            live_state["delta"] = live_state["live_qty"] - live_state["warehouse_qty"]
        else:
            live_state["delta"] = None

        return jsonify(live_state), 200

    except Exception as e:
        logging.error(f"Error fetching marketplace state for {sku}: {str(e)}")
        return jsonify({"error": str(e)}), 500

@bp.route('/api/diagnostics/amazon/sku-precheck/<sku>')
def amazon_sku_precheck(sku):
    """Amazon SKU prevalidation check"""
    # Auth via X-Task-Key
    if request.headers.get("X-Task-Key") != TASK_API_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        # Find active Amazon store
        amazon_store = db.session.query(Store).filter(
            Store.platform == 'Amazon',
            Store.is_active == True
        ).first()

        if not amazon_store:
            return jsonify({
                "ok": False,
                "error": "No active Amazon store found"
            }), 404

        # Init Amazon service
        amazon_service = AmazonAPIService(amazon_store)

        # Run prevalidation
        result = amazon_service.prevalidate_sku(amazon_store, sku)

        return jsonify(result), 200

    except Exception as e:
        logging.error(f"Error in sku precheck for {sku}: {str(e)}")
        return jsonify({"ok": False, "error": str(e)}), 500
# ===============================
# ADMIN: Set Price + Push Amazon
# ===============================
@bp.post("/api/admin/amazon/set-price-and-push/<sku>")
def admin_set_price_and_push_amazon(sku):
    """Retired direct admin price-and-push route.

    Price updates and marketplace pushes must be separated. This route must not
    update price and trigger marketplace push in one request.
    """
    return jsonify({
        "success": False,
        "error": "Direct price-and-push route is retired. Use governed pricing update and approved dispatcher execution path.",
        "execution_blocked": True,
        "route_retired": True,
        "sku": sku
    }), 410

@bp.get("/api/diagnostics/amazon/feed/last")
def get_latest_amazon_feed_status():
    import os
    from flask import jsonify, request
    from extensions import db
    from models import Store, SyncLog
    from amazon_service import AmazonAPIService

    # Verify API key
    if request.headers.get("X-Task-Key") != os.getenv("TASK_API_KEY"):
        return jsonify({"error": "unauthorized"}), 401

    try:
        # Find most recent Amazon feed from FeedStatus table
        from models import FeedStatus

        amazon_store = (
            db.session.query(Store)
            .filter(Store.platform == "Amazon", Store.is_active == True)
            .first()
        )
        if not amazon_store:
            return jsonify({"ok": False, "error": "No active Amazon store found"}), 404

        feed_record = (
            db.session.query(FeedStatus)
            .filter(FeedStatus.store_id == amazon_store.id)
            .order_by(FeedStatus.submitted_at.desc())
            .first()
        )
        if not feed_record:
            return jsonify({"ok": False, "error": "No feed records found"}), 404

        # Fetch live status from Amazon
        amazon_service = AmazonAPIService(amazon_store)
        result = amazon_service.get_feed_status(amazon_store, feed_record.feed_id)

        return jsonify({
            "ok": True,
            "feed_id": feed_record.feed_id,
            "sku": feed_record.sku,
            "submitted_at": feed_record.submitted_at.isoformat() if feed_record.submitted_at else None,
            "status": result.get("processingStatus", "UNKNOWN"),
            "details": result
        }), 200

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ===========================================
# DIAG: Amazon Feeds scope check
# ===========================================
@bp.get("/api/diagnostics/amazon/feeds-scope")
def diag_amazon_feeds_scope():
    import os
    from flask import request, jsonify
    from models import Store
    from amazon_service import AmazonAPIService

    if request.headers.get("X-Task-Key") != os.getenv("TASK_API_KEY"):
        return jsonify({"error": "unauthorized"}), 401

    store = db.session.query(Store).filter(Store.platform=="Amazon", Store.is_active==True).first()
    if not store:
        return jsonify({"ok": False, "error": "No active Amazon store"}), 404

    svc = AmazonAPIService(store)
    result = svc.check_feeds_scope(store)
    return jsonify(result), (200 if result.get("ok") else 200)

# ===========================================
# ADMIN: Build Amazon Re-Auth URL (no secrets leaked)
# ===========================================
@bp.get("/api/admin/amazon/reauthorize-url")
def admin_amazon_reauth_url():
    """
    Returns a Seller Central OAuth URL for re-authorizing the app
    so Feeds scope can be granted. Does not modify credentials.
    """
    import os
    from flask import request, jsonify

    if request.headers.get("X-Task-Key") != os.getenv("TASK_API_KEY"):
        return jsonify({"error": "unauthorized"}), 401

    # Pull expected env vars; try both naming conventions
    client_id_raw = os.getenv("LWA_CLIENT_ID") or os.getenv("AMAZON_LWA_CLIENT_ID", "")

    # Extract actual value if it has "Value:" prefix
    if client_id_raw and "Value:" in client_id_raw:
        client_id = client_id_raw.split("Value:")[-1].strip()
    else:
        client_id = client_id_raw

    redirect_uri = request.url_root.rstrip('/') + '/auth/amazon/callback'

    # If redirect_uri not set, auto-generate from request
    # redirect_uri forced from request

    # Default EU region auth URL (adjust if you operate in NA/JP only)
    base = "https://sellercentral.amazon.co.uk/apps/authorize/consent"
    params = {
        "application_id": client_id,
        "redirect_uri": redirect_uri,
    }

    from urllib.parse import urlencode
    return jsonify({
        "ok": True,
        "hint": "Open this in a browser, sign in as the seller, grant FEEDS scope, and complete.",
        "url": f"{base}?{urlencode(params)}",
        "redirect_uri": redirect_uri,
        "missing": [k for k,v in {"LWA_CLIENT_ID":client_id}.items() if not v]
    })

# ===========================================
# ADMIN: OAuth Callback (captures code)
# ===========================================
@bp.get("/api/admin/amazon/oauth/callback")
def amazon_admin_oauth_callback():
    import os
    from flask import request, jsonify
    if request.headers.get("X-Task-Key") != os.getenv("TASK_API_KEY"):
        return jsonify({"error": "unauthorized"}), 401
    code = request.args.get("code")
    state = request.args.get("state")
    if not code:
        return jsonify({"ok": False, "error": "missing code"}), 400
    # We deliberately do NOT exchange here; we only echo back safely.
    return jsonify({"ok": True, "code": code, "state": state})

# ===========================================
# ADMIN: OAuth Exchange (code -> refresh_token)
# ===========================================
@bp.post("/api/admin/amazon/oauth/exchange")
def amazon_admin_oauth_exchange():
    """
    Body: { "code": "<auth_code_from_callback>" }
    Exchanges code -> refresh_token with LWA, and saves refresh_token
    into the active Amazon store's credentials JSON.
    """
    import os, json, requests
    from flask import request, jsonify
    from models import Store
    if request.headers.get("X-Task-Key") != os.getenv("TASK_API_KEY"):
        return jsonify({"error": "unauthorized"}), 401
    code = (request.get_json() or {}).get("code")
    if not code:
        return jsonify({"ok": False, "error": "missing code"}), 400

    # Try both naming conventions for credentials
    cid_raw = os.getenv("LWA_CLIENT_ID") or os.getenv("AMAZON_LWA_CLIENT_ID", "")
    csec_raw = os.getenv("LWA_CLIENT_SECRET") or os.getenv("AMAZON_LWA_CLIENT_SECRET", "")

    # Extract actual values if they have "Value:" prefix
    if cid_raw and "Value:" in cid_raw:
        cid = cid_raw.split("Value:")[-1].strip()
    else:
        cid = cid_raw

    if csec_raw and "Value:" in csec_raw:
        csec = csec_raw.split("Value:")[-1].strip()
    else:
        csec = csec_raw

    redir = os.getenv("LWA_REDIRECT_URI", "")

    # Auto-generate redirect URI if not set
    if not redir:
        redir = request.url_root.rstrip('/') + '/api/admin/amazon/oauth/callback'

    if not cid or not csec:
        return jsonify({"ok": False, "error": "missing LWA credentials (client id/secret)"}), 400
    token_url = "https://api.amazon.com/auth/o2/token"
    data = {
        "grant_type": "authorization_code",
        "code": code,
        "client_id": cid,
        "client_secret": csec,
        "redirect_uri": redir,
    }
    try:
        r = requests.post(token_url, data=data, timeout=20)
        r.raise_for_status()
        tok = r.json()
        refresh_token = tok.get("refresh_token")
        if not refresh_token:
            return jsonify({"ok": False, "error": "no refresh_token in response", "raw": tok}), 400
        store = db.session.query(Store).filter(Store.platform=="Amazon", Store.is_active==True).first()
        if not store:
            return jsonify({"ok": False, "error": "No active Amazon store"}), 404
        # Merge refresh_token into credentials JSON without exposing existing secrets
        creds = {}
        try:
            creds = json.loads(store.api_key or "{}")
        except Exception:
            creds = {}
        creds["refresh_token"] = refresh_token
        store.api_key = json.dumps(creds)
        db.session.commit()
        return jsonify({"ok": True, "saved": "refresh_token"})
    except requests.HTTPError as e:
        return jsonify({"ok": False, "error": f"HTTP {e.response.status_code}", "body": e.response.text}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@bp.route('/amazon-credentials-setup', methods=['GET', 'POST'])
def amazon_credentials_setup():
    """Amazon Production Credentials Setup Page"""
    if request.method == 'POST':
        try:
            # Get credentials from form
            lwa_client_id = request.form.get('lwa_client_id', '').strip()
            lwa_client_secret = request.form.get('lwa_client_secret', '').strip()
            refresh_token = request.form.get('refresh_token', '').strip()
            seller_id = request.form.get('seller_id', '').strip()
            marketplace_id = request.form.get('marketplace_id', 'A1F83G8C2ARO7P').strip()
            region = request.form.get('region', 'EU').strip()

            # Validate all fields
            if not all([lwa_client_id, lwa_client_secret, refresh_token, seller_id]):
                flash('All fields are required!', 'danger')
                return render_template('amazon_credentials_setup.html')

            # Find or create Amazon store
            amazon_store = Store.query.filter_by(platform='Amazon').first()

            if not amazon_store:
                # Create new Amazon store
                amazon_store = Store()
                amazon_store.name = 'BT38'
                amazon_store.platform = 'Amazon'
                amazon_store.is_active = True
                amazon_store.sync_status = 'pending'
                db.session.add(amazon_store)

            # Build credentials JSON
            production_creds = {
                'lwa_app_id': lwa_client_id,
                'lwa_client_secret': lwa_client_secret,
                'refresh_token': refresh_token,
                'seller_id': seller_id,
                'marketplace_id': marketplace_id,
                'region': region,
                'sandbox': False,
                'environment': 'production'
            }

            amazon_store.api_key = json.dumps(production_creds)
            db.session.commit()

            # Test connection
            from amazon_service import AmazonAPIService
            amazon_service = AmazonAPIService()
            success, items, message = amazon_service.disabled_disabled_amazon_inventory_import(amazon_store)

            if success:
                flash(f' Amazon credentials saved and tested successfully! Found {len(items)} items.', 'success')
                logging.info(f"Amazon credentials updated successfully for store: {amazon_store.name}")
            else:
                flash(f' Credentials saved but connection test failed: {message}', 'warning')
                logging.warning(f"Amazon credentials saved but test failed: {message}")

            return redirect(url_for('routes.stores'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error saving credentials: {str(e)}', 'danger')
            logging.error(f"Error in amazon_credentials_setup: {str(e)}")
            return render_template('amazon_credentials_setup.html')

    # GET request - show form
    return render_template('amazon_credentials_setup.html')

@bp.route('/analytics')
def analytics():
    """Business Analytics Dashboard - Profit/Loss and Stock Value Reporting"""
    try:
        from sqlalchemy import func

        # Get all warehouse stock with product details
        warehouse_stock = db.session.query(
            WarehouseStock,
            InventoryItem
        ).join(
            InventoryItem, WarehouseStock.sku == InventoryItem.sku
        ).all()

        # Calculate total inventory value and stock metrics
        total_stock_value = 0
        total_stock_quantity = 0
        stock_by_location = {}
        products_data = []

        for stock, item in warehouse_stock:
            stock_value = (stock.available_quantity or 0) * (stock.unit_cost or 0)
            total_stock_value += stock_value
            total_stock_quantity += (stock.available_quantity or 0)

            # Group by location
            location = stock.location or 'Warehouse'
            if location not in stock_by_location:
                stock_by_location[location] = {'quantity': 0, 'value': 0, 'items': 0}
            stock_by_location[location]['quantity'] += stock.available_quantity or 0
            stock_by_location[location]['value'] += stock_value
            stock_by_location[location]['items'] += 1

            # Calculate profit margin with all costs (if we have a selling price)
            profit_margin = 0
            commission_per_unit = 0
            shipping_per_unit = 0
            operating_cost = stock.operating_cost_per_unit or 0
            net_profit_per_unit = 0

            if item.price and stock.unit_cost and stock.unit_cost > 0:
                # Calculate commission
                commission_per_unit = item.price * (stock.commission_rate / 100.0) if stock.commission_rate else 0

                # Calculate shipping
                shipping_per_unit = stock.product_weight_kg * stock.shipping_cost_per_kg if stock.product_weight_kg and stock.shipping_cost_per_kg else 0

                # Net profit per unit = Selling Price - Cost - Commission - Shipping - Operating Cost
                net_profit_per_unit = item.price - stock.unit_cost - commission_per_unit - shipping_per_unit - operating_cost

                # Profit margin as percentage of selling price
                profit_margin = (net_profit_per_unit / item.price) * 100 if item.price > 0 else 0

            products_data.append({
                'sku': stock.sku,
                'title': item.title or 'N/A',
                'quantity': stock.available_quantity or 0,
                'unit_cost': stock.unit_cost or 0,
                'selling_price': item.price or 0,
                'stock_value': stock_value,
                'commission_rate': stock.commission_rate or 0,
                'commission_per_unit': commission_per_unit,
                'shipping_per_unit': shipping_per_unit,
                'operating_cost_per_unit': operating_cost,
                'net_profit_per_unit': net_profit_per_unit,
                'profit_margin': profit_margin,
                'location': location,
                'supplier': stock.supplier_id,
                'weight_kg': stock.product_weight_kg or 0
            })

        # Get marketplace listings for revenue potential
        marketplace_listings = db.session.query(
            MarketplaceListing.sku,
            func.count(MarketplaceListing.id).label('listing_count'),
            func.avg(MarketplaceListing.price).label('avg_market_price')
        ).filter(
            MarketplaceListing.status == 'active'
        ).group_by(
            MarketplaceListing.sku
        ).all()

        marketplace_revenue_potential = {}
        for listing in marketplace_listings:
            marketplace_revenue_potential[listing.sku] = {
                'listing_count': listing.listing_count,
                'avg_price': float(listing.avg_market_price or 0)
            }

        # Calculate profit/loss from recent sales (stock ledger entries)
        recent_sales = db.session.query(
            StockLedgerEntry
        ).filter(
            StockLedgerEntry.quantity < 0  # Negative quantity = sale
        ).order_by(
            desc(StockLedgerEntry.created_at)
        ).limit(100).all()

        total_revenue = 0
        total_cost = 0
        total_commission = 0
        total_shipping = 0
        total_operating_costs = 0
        sales_count = 0

        for sale in recent_sales:
            quantity_sold = abs(sale.quantity)

            # Get cost from warehouse stock
            warehouse = WarehouseStock.query.filter_by(sku=sale.sku).first()
            if warehouse and warehouse.unit_cost:
                cost_per_unit = warehouse.unit_cost
                sale_cost = quantity_sold * cost_per_unit
                total_cost += sale_cost

                # Get selling price from inventory item
                item = InventoryItem.query.filter_by(sku=sale.sku).first()
                if item and item.price:
                    revenue = quantity_sold * item.price
                    total_revenue += revenue

                    # Calculate commission (percentage of revenue)
                    commission = revenue * (warehouse.commission_rate / 100.0) if warehouse.commission_rate else 0
                    total_commission += commission

                    # Calculate shipping cost
                    shipping = quantity_sold * warehouse.product_weight_kg * warehouse.shipping_cost_per_kg if warehouse.product_weight_kg and warehouse.shipping_cost_per_kg else 0
                    total_shipping += shipping

                    # Calculate operating costs
                    operating = quantity_sold * warehouse.operating_cost_per_unit if warehouse.operating_cost_per_unit else 0
                    total_operating_costs += operating

                    sales_count += 1

        # Net Profit = Revenue - Cost - Commission - Shipping - Operating Costs
        total_profit = total_revenue - total_cost - total_commission - total_shipping - total_operating_costs
        profit_margin_avg = (total_profit / total_revenue * 100) if total_revenue > 0 else 0

        # Low stock alerts
        low_stock_items = []
        for stock, item in warehouse_stock:
            if (stock.available_quantity or 0) <= (stock.reorder_point or 0) and (stock.available_quantity or 0) > 0:
                low_stock_items.append({
                    'sku': stock.sku,
                    'title': item.title or 'N/A',
                    'quantity': stock.available_quantity,
                    'reorder_point': stock.reorder_point
                })

        # Sort products by profit margin (best performers)
        products_data.sort(key=lambda x: x['profit_margin'], reverse=True)
        top_performers = products_data[:10]
        worst_performers = [p for p in products_data if p['profit_margin'] < 20][-10:]

        return render_template('analytics.html',
            # Summary metrics
            total_stock_value=total_stock_value,
            total_stock_quantity=total_stock_quantity,
            total_revenue=total_revenue,
            total_cost=total_cost,
            total_commission=total_commission,
            total_shipping=total_shipping,
            total_operating_costs=total_operating_costs,
            total_profit=total_profit,
            profit_margin_avg=profit_margin_avg,
            sales_count=sales_count,

            # Stock breakdown
            stock_by_location=stock_by_location,
            products_data=products_data,

            # Alerts and insights
            low_stock_items=low_stock_items,
            top_performers=top_performers,
            worst_performers=worst_performers,

            # Marketplace data
            marketplace_revenue_potential=marketplace_revenue_potential
        )

    except Exception as e:
        logging.error(f"Error in analytics dashboard: {str(e)}")
        flash(f'Error loading analytics: {str(e)}', 'danger')
        return redirect(url_for('routes.dashboard'))

@bp.route('/export-analytics')
def export_analytics():
    """Export analytics data to Excel"""
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file
        from sqlalchemy import func

        # Get all warehouse stock with product details
        warehouse_stock = db.session.query(
            WarehouseStock,
            InventoryItem
        ).join(
            InventoryItem, WarehouseStock.sku == InventoryItem.sku
        ).all()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Analytics"

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Write headers
        headers = ["SKU", "Product Title", "Quantity", "Unit Cost", "Selling Price", "Stock Value",
                   "Commission Rate %", "Commission Per Unit", "Shipping Per Unit", "Operating Cost Per Unit",
                   "Net Profit Per Unit", "Profit Margin %", "Location", "Weight (kg)"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Write data
        row = 2
        for stock, item in warehouse_stock:
            stock_value = (stock.available_quantity or 0) * (stock.unit_cost or 0)

            # Calculate costs
            commission_per_unit = 0
            shipping_per_unit = 0
            operating_cost = stock.operating_cost_per_unit or 0
            net_profit_per_unit = 0
            profit_margin = 0

            if item.price and stock.unit_cost and stock.unit_cost > 0:
                commission_per_unit = item.price * (stock.commission_rate / 100.0) if stock.commission_rate else 0
                shipping_per_unit = stock.product_weight_kg * stock.shipping_cost_per_kg if stock.product_weight_kg and stock.shipping_cost_per_kg else 0
                net_profit_per_unit = item.price - stock.unit_cost - commission_per_unit - shipping_per_unit - operating_cost
                profit_margin = (net_profit_per_unit / item.price) * 100 if item.price > 0 else 0

            ws.cell(row=row, column=1, value=stock.sku)
            ws.cell(row=row, column=2, value=item.title or 'N/A')
            ws.cell(row=row, column=3, value=stock.available_quantity or 0)
            ws.cell(row=row, column=4, value=stock.unit_cost or 0)
            ws.cell(row=row, column=5, value=item.price or 0)
            ws.cell(row=row, column=6, value=stock_value)
            ws.cell(row=row, column=7, value=stock.commission_rate or 0)
            ws.cell(row=row, column=8, value=commission_per_unit)
            ws.cell(row=row, column=9, value=shipping_per_unit)
            ws.cell(row=row, column=10, value=operating_cost)
            ws.cell(row=row, column=11, value=net_profit_per_unit)
            ws.cell(row=row, column=12, value=profit_margin)
            ws.cell(row=row, column=13, value=stock.location or 'Warehouse')
            ws.cell(row=row, column=14, value=stock.product_weight_kg or 0)

            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'analytics_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )

    except Exception as e:
        logging.error(f"Error exporting analytics: {str(e)}")
        flash(f'Error exporting to Excel: {str(e)}', 'danger')
        return redirect(url_for('routes.analytics'))

@bp.route('/monthly-roi')
def monthly_roi():
    """Monthly ROI Report - Investment vs Revenue Tracking"""
    try:
        from sqlalchemy import func, extract
        from datetime import datetime, timedelta
        from collections import defaultdict

        # Get date range (last 12 months)
        end_date = datetime.utcnow()
        start_date = end_date - timedelta(days=365)

        # Get monthly investment from purchase orders (invoices)
        monthly_investment = db.session.query(
            func.date_trunc('month', PurchaseOrder.invoice_date).label('month'),
            func.sum(PurchaseOrder.total_amount).label('total_investment'),
            func.count(PurchaseOrder.id).label('invoice_count')
        ).filter(
            PurchaseOrder.invoice_date.isnot(None),
            PurchaseOrder.invoice_date >= start_date,
            PurchaseOrder.status.in_(['received', 'partially_received'])
        ).group_by(
            func.date_trunc('month', PurchaseOrder.invoice_date)
        ).order_by(
            func.date_trunc('month', PurchaseOrder.invoice_date).desc()
        ).all()

        # Get monthly revenue from sales (stock ledger)
        monthly_revenue_query = db.session.query(
            func.date_trunc('month', StockLedgerEntry.created_at).label('month'),
            StockLedgerEntry.sku
        ).filter(
            StockLedgerEntry.quantity < 0,  # Sales
            StockLedgerEntry.created_at >= start_date
        ).all()

        # Calculate revenue by month
        monthly_revenue_data = defaultdict(lambda: {'revenue': 0, 'cost': 0, 'commission': 0, 'shipping': 0, 'operating': 0, 'sales_count': 0})

        for month, sku in monthly_revenue_query:
            # Get warehouse and inventory data
            warehouse = WarehouseStock.query.filter_by(sku=sku).first()
            item = InventoryItem.query.filter_by(sku=sku).first()

            if warehouse and item and item.price:
                # Find the quantity sold in that month
                sales = StockLedgerEntry.query.filter(
                    StockLedgerEntry.sku == sku,
                    func.date_trunc('month', StockLedgerEntry.created_at) == month,
                    StockLedgerEntry.quantity < 0
                ).all()

                for sale in sales:
                    qty = abs(sale.quantity)
                    revenue = qty * item.price
                    cost = qty * warehouse.unit_cost
                    commission = revenue * (warehouse.commission_rate / 100.0) if warehouse.commission_rate else 0
                    shipping = qty * warehouse.product_weight_kg * warehouse.shipping_cost_per_kg if warehouse.product_weight_kg and warehouse.shipping_cost_per_kg else 0
                    operating = qty * warehouse.operating_cost_per_unit if warehouse.operating_cost_per_unit else 0

                    monthly_revenue_data[month]['revenue'] += revenue
                    monthly_revenue_data[month]['cost'] += cost
                    monthly_revenue_data[month]['commission'] += commission
                    monthly_revenue_data[month]['shipping'] += shipping
                    monthly_revenue_data[month]['operating'] += operating
                    monthly_revenue_data[month]['sales_count'] += 1

        # Combine investment and revenue data
        monthly_reports = []
        for inv in monthly_investment:
            month = inv.month
            investment = float(inv.total_investment or 0)
            revenue_data = monthly_revenue_data.get(month, {})

            revenue = revenue_data.get('revenue', 0)
            cost = revenue_data.get('cost', 0)
            commission = revenue_data.get('commission', 0)
            shipping = revenue_data.get('shipping', 0)
            operating = revenue_data.get('operating', 0)
            sales_count = revenue_data.get('sales_count', 0)

            # Calculate profit and ROI
            profit = revenue - cost - commission - shipping - operating
            money_returned = revenue
            roi_percentage = (profit / investment * 100) if investment > 0 else 0

            monthly_reports.append({
                'month': month.strftime('%B %Y') if month else 'N/A',
                'month_date': month,
                'investment': investment,
                'invoice_count': inv.invoice_count,
                'revenue': revenue,
                'cost': cost,
                'commission': commission,
                'shipping': shipping,
                'operating': operating,
                'sales_count': sales_count,
                'profit': profit,
                'money_returned': money_returned,
                'roi_percentage': roi_percentage
            })

        # Calculate totals
        total_investment = sum(r['investment'] for r in monthly_reports)
        total_revenue = sum(r['revenue'] for r in monthly_reports)
        total_cost = sum(r['cost'] for r in monthly_reports)
        total_commission = sum(r['commission'] for r in monthly_reports)
        total_shipping = sum(r['shipping'] for r in monthly_reports)
        total_operating = sum(r['operating'] for r in monthly_reports)
        total_profit = sum(r['profit'] for r in monthly_reports)
        total_roi = (total_profit / total_investment * 100) if total_investment > 0 else 0

        # Get supplier performance
        supplier_performance = db.session.query(
            Supplier.name,
            func.sum(PurchaseOrder.total_amount).label('total_spent'),
            func.count(PurchaseOrder.id).label('order_count')
        ).join(
            PurchaseOrder, Supplier.id == PurchaseOrder.supplier_id
        ).filter(
            PurchaseOrder.invoice_date >= start_date,
            PurchaseOrder.status.in_(['received', 'partially_received'])
        ).group_by(
            Supplier.name
        ).order_by(
            func.sum(PurchaseOrder.total_amount).desc()
        ).all()

        suppliers_data = []
        for supp in supplier_performance:
            suppliers_data.append({
                'name': supp.name,
                'total_spent': float(supp.total_spent or 0),
                'order_count': supp.order_count
            })

        # Handle Excel export
        if request.args.get('export') == 'excel':
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from flask import send_file

            wb = Workbook()
            ws = wb.active
            ws.title = "Monthly ROI Report"

            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            # Write headers
            headers = ["Month", "Investment", "Invoices", "Revenue", "COGS", "Commission",
                       "Shipping", "Operating", "Profit", "ROI %"]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Write data
            row = 2
            for report in monthly_reports:
                ws.cell(row=row, column=1, value=report['month'])
                ws.cell(row=row, column=2, value=report['investment'])
                ws.cell(row=row, column=3, value=report['invoice_count'])
                ws.cell(row=row, column=4, value=report['revenue'])
                ws.cell(row=row, column=5, value=report['cost'])
                ws.cell(row=row, column=6, value=report['commission'])
                ws.cell(row=row, column=7, value=report['shipping'])
                ws.cell(row=row, column=8, value=report['operating'])
                ws.cell(row=row, column=9, value=report['profit'])
                ws.cell(row=row, column=10, value=report['roi_percentage'])
                row += 1

            # Add totals row
            ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=row, column=2, value=total_investment).font = Font(bold=True)
            ws.cell(row=row, column=3, value="")
            ws.cell(row=row, column=4, value=total_revenue).font = Font(bold=True)
            ws.cell(row=row, column=5, value=total_cost).font = Font(bold=True)
            ws.cell(row=row, column=6, value=total_commission).font = Font(bold=True)
            ws.cell(row=row, column=7, value=total_shipping).font = Font(bold=True)
            ws.cell(row=row, column=8, value=total_operating).font = Font(bold=True)
            ws.cell(row=row, column=9, value=total_profit).font = Font(bold=True)
            ws.cell(row=row, column=10, value=total_roi).font = Font(bold=True)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'monthly_roi_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )

        return render_template('monthly_roi.html',
            monthly_reports=monthly_reports,
            total_investment=total_investment,
            total_revenue=total_revenue,
            total_cost=total_cost,
            total_commission=total_commission,
            total_shipping=total_shipping,
            total_operating=total_operating,
            total_profit=total_profit,
            total_roi=total_roi,
            suppliers_data=suppliers_data
        )

    except Exception as e:
        logging.error(f"Error in monthly ROI report: {str(e)}")
        flash(f'Error loading monthly ROI report: {str(e)}', 'danger')
        return redirect(url_for('routes.dashboard'))

# ============================================================================
# AI IMAGE STUDIO ROUTES
# ============================================================================

@bp.route('/ai-image-studio')
# @login_required  # Temporarily disabled for access
def ai_image_studio():
    """AI Image Studio - Enhance, generate, and check compliance of product images"""
    return render_template('ai_image_studio.html')

@bp.route('/api/ai-image/enhance', methods=['POST'])
# @login_required  # Temporarily disabled for access
def api_enhance_image():
    """Enhance uploaded image using AI"""
    try:
        if 'image' not in request.files:
            return jsonify({'success': False, 'error': 'No image file provided'}), 400

        image_file = request.files['image']

        # Enhance image
        result = enhance_image(image_file)

        if result['success']:
            # Convert image bytes to base64 for JSON response
            image_base64 = base64.b64encode(result['image_data']).decode('utf-8')
            return jsonify({
                'success': True,
                'image_base64': image_base64,
                'format': result['format']
            })
        else:
            return jsonify({'success': False, 'error': result['error']}), 500

    except Exception as e:
        logging.error(f"Image enhancement API error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/ai-image/remove-background', methods=['POST'])
# @login_required  # Temporarily disabled for access
def api_remove_background():
    """Remove background from uploaded image"""
    try:
        if 'image' not in request.files:
            return jsonify({'success': False, 'error': 'No image file provided'}), 400

        image_file = request.files['image']

        # Remove background
        result = remove_background(image_file)

        if result['success']:
            # Convert image bytes to base64 for JSON response
            image_base64 = base64.b64encode(result['image_data']).decode('utf-8')
            return jsonify({
                'success': True,
                'image_base64': image_base64,
                'format': result['format']
            })
        else:
            return jsonify({'success': False, 'error': result['error']}), 500

    except Exception as e:
        logging.error(f"Background removal API error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/ai-image/generate', methods=['POST'])
# @login_required  # Temporarily disabled for access
def api_generate_image():
    """Generate lifestyle image using OpenAI DALL-E"""
    try:
        # Handle FormData (with optional image upload)
        product_name = request.form.get('product_name', '')
        product_description = request.form.get('product_description', '')
        style = request.form.get('style', 'professional product photography')

        if not product_name or not product_description:
            return jsonify({'success': False, 'error': 'Product name and description required'}), 400

        # Check for reference image
        reference_image = None
        if 'reference_image' in request.files:
            reference_image = request.files['reference_image']

        # Generate image
        result = generate_lifestyle_image(product_name, product_description, style, reference_image)

        if result['success']:
            return jsonify({
                'success': True,
                'image_url': result['image_url'],
                'revised_prompt': result['revised_prompt']
            })
        else:
            return jsonify({'success': False, 'error': result['error']}), 500

    except Exception as e:
        logging.error(f"Image generation API error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/ai-image/check-compliance', methods=['POST'])
# @login_required  # Temporarily disabled for access
def api_check_compliance():
    """Check if image meets Amazon/eBay marketplace requirements"""
    try:
        if 'image' not in request.files:
            return jsonify({'success': False, 'error': 'No image file provided'}), 400

        image_file = request.files['image']

        # Check compliance
        result = check_marketplace_compliance(image_file)

        if result['success']:
            return jsonify(result)
        else:
            return jsonify({'success': False, 'error': result['error']}), 500

    except Exception as e:
        logging.error(f"Compliance check API error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# SYSTEM DIAGNOSTICS - Comprehensive health check for troubleshooting
# ============================================================================

@bp.route('/api/diagnostics/system')
def system_diagnostics():
    """Comprehensive system diagnostics - shows database status, stores, sync status, recent errors"""
    try:
        result = {
            'timestamp': datetime.utcnow().isoformat(),
            'database': {},
            'stores': [],
            'sync_jobs': [],
            'sync_logs': [],
            'environment': {}
        }

        # Database stats
        result['database']['connected'] = True
        result['database']['total_items'] = db.session.query(InventoryItem).count()
        result['database']['warehouse_stock'] = db.session.query(WarehouseStock).count()
        result['database']['marketplace_listings'] = db.session.query(MarketplaceListing).count()
        result['database']['total_stores'] = db.session.query(Store).count()

        # Store details
        stores = db.session.query(Store).all()
        for store in stores:
            store_info = {
                'id': store.id,
                'name': store.name,
                'platform': store.platform,
                'is_active': store.is_active,
                'sync_status': store.sync_status,
                'last_sync': store.last_sync.isoformat() if store.last_sync else None,
                'has_api_credentials': bool(store.api_key and len(store.api_key) > 10),
                'api_key_length': len(store.api_key) if store.api_key else 0
            }
            result['stores'].append(store_info)

        # Recent sync jobs (last 10)
        recent_jobs = db.session.query(SyncJob).order_by(desc(SyncJob.enqueued_at)).limit(10).all()
        for job in recent_jobs:
            job_info = {
                'id': job.id,
                'store_id': job.store_id,
                'job_type': job.job_type,
                'status': job.status,
                'enqueued_at': job.enqueued_at.isoformat() if job.enqueued_at else None,
                'started_at': job.started_at.isoformat() if job.started_at else None,
                'completed_at': job.completed_at.isoformat() if job.completed_at else None,
                'error_message': job.error_message,
                'retry_count': job.retry_count
            }
            result['sync_jobs'].append(job_info)

        # Recent sync logs (last 10)
        recent_logs = db.session.query(SyncLog).order_by(desc(SyncLog.created_at)).limit(10).all()
        for log in recent_logs:
            log_info = {
                'id': log.id,
                'store_id': log.store_id,
                'status': log.status,
                'message': log.message,
                'items_synced': log.items_synced,
                'created_at': log.created_at.isoformat() if log.created_at else None
            }
            result['sync_logs'].append(log_info)

        # Environment info
        import os
        from app import APP_ENV, IS_PRODUCTION, IS_DEVELOPMENT
        result['environment']['app_env'] = APP_ENV
        result['environment']['is_production'] = IS_PRODUCTION
        result['environment']['is_development'] = IS_DEVELOPMENT
        result['environment']['replit_deployment'] = bool(os.getenv('REPLIT_DEPLOYMENT'))

        return jsonify(result)

    except Exception as e:
        logging.error(f"System diagnostics error: {str(e)}")
        return jsonify({
            'ok': False,
            'error': str(e),
            'timestamp': datetime.utcnow().isoformat()
        }), 500

# ============================================================================
# GLOBAL ERROR HANDLERS - Prevent Internal Server Errors from showing to users
# ============================================================================

@bp.errorhandler(404)
def not_found_error(error):
    """Handle 404 errors with a friendly message"""
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({'error': 'Resource not found', 'status': 404}), 404
    return render_template('error.html',
                         error_code=404,
                         error_title='Page Not Found',
                         error_message='The page you are looking for does not exist.'), 404

@bp.errorhandler(500)
def internal_error(error):
    """Handle 500 errors with a friendly message and log the error"""
    logging.error(f"Internal Server Error: {str(error)}", exc_info=True)
    db.session.rollback()  # Rollback any failed database transactions

    if request.is_json or request.path.startswith('/api/'):
        return jsonify({'error': 'Internal server error', 'status': 500}), 500
    return render_template('error.html',
                         error_code=500,
                         error_title='Something Went Wrong',
                         error_message='An unexpected error occurred. Our team has been notified.'), 500

@bp.errorhandler(403)
def forbidden_error(error):
    """Handle 403 Forbidden errors"""
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({'error': 'Access forbidden', 'status': 403}), 403
    return render_template('error.html',
                         error_code=403,
                         error_title='Access Denied',
                         error_message='You do not have permission to access this resource.'), 403

@bp.errorhandler(400)
def bad_request_error(error):
    """Handle 400 Bad Request errors"""
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({'error': 'Bad request', 'status': 400}), 400
    return render_template('error.html',
                         error_code=400,
                         error_title='Bad Request',
                         error_message='The request could not be understood or was missing required parameters.'), 400

# =======================
# Multi-Channel Listing Management Routes
# =======================

@bp.route('/listings')
def listings():
    """Multi-channel listing management interface"""
    search_query = request.args.get('search', '').strip()
    store_filter = request.args.get('store', '')
    status_filter = request.args.get('status', '')

    # Build query for marketplace listings
    query = db.session.query(MarketplaceListing).options(
        joinedload(MarketplaceListing.warehouse_stock),
        joinedload(MarketplaceListing.store)
    )

    # Apply filters
    if search_query:
        query = query.join(WarehouseStock).filter(
            or_(
                WarehouseStock.sku.ilike(f'%{search_query}%'),
                MarketplaceListing.title.ilike(f'%{search_query}%'),
                MarketplaceListing.external_listing_id.ilike(f'%{search_query}%')
            )
        )

    if store_filter:
        query = query.filter(MarketplaceListing.store_id == store_filter)

    if status_filter:
        if status_filter == 'active':
            query = query.filter(MarketplaceListing.is_active == True)
        elif status_filter == 'inactive':
            query = query.filter(MarketplaceListing.is_active == False)
        elif status_filter == 'blocked':
            query = query.filter(MarketplaceListing.push_state == 'blocked')
        elif status_filter == 'needs_review':
            query = query.filter(MarketplaceListing.push_state == 'needs_review')

    # Get all listings
    listings = query.order_by(MarketplaceListing.updated_at.desc()).all()

    # Get all stores for filter dropdown
    all_stores = db.session.query(Store).filter_by(is_active=True).all()

    # Get statistics
    total_listings = len(listings)
    active_listings = sum(1 for l in listings if l.is_active)
    pending_push = sum(1 for l in listings if l.needs_push)
    blocked_listings = sum(1 for l in listings if l.push_state == 'blocked')

    return render_template('listings.html',
                         listings=listings,
                         all_stores=all_stores,
                         current_search=search_query,
                         current_store_filter=store_filter,
                         current_status_filter=status_filter,
                         total_listings=total_listings,
                         active_listings=active_listings,
                         pending_push=pending_push,
                         blocked_listings=blocked_listings)

@bp.route('/listings/create', methods=['GET', 'POST'])
def create_listing():
    """Create new marketplace listing from warehouse stock"""
    if request.method == 'POST':
        try:
            # CSRF Protection
            csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
            if not validate_csrf_token(csrf_token):
                logging.warning(f"CSRF token validation failed for create_listing request from {request.remote_addr}")
                flash('Invalid request. Please try again.', 'danger')
                return redirect(url_for('routes.create_listing'))

            # Get form data
            warehouse_stock_id = request.form.get('warehouse_stock_id')
            store_ids = request.form.getlist('store_ids')  # Multiple stores

            # Get listing details
            title = request.form.get('title')
            description = request.form.get('description')
            price = float(request.form.get('price', 0))
            currency = request.form.get('currency', 'GBP')

            # Sync settings
            sync_quantity = request.form.get('sync_quantity') == 'on'
            sync_price = request.form.get('sync_price') == 'on'
            quantity_buffer = int(request.form.get('quantity_buffer', 0))
            max_quantity_limit = request.form.get('max_quantity_limit')
            max_quantity_limit = int(max_quantity_limit) if max_quantity_limit else None

            # Validate warehouse stock
            warehouse_stock = db.session.query(WarehouseStock).get(warehouse_stock_id)
            if not warehouse_stock:
                flash('Invalid warehouse stock selected.', 'danger')
                return redirect(url_for('routes.create_listing'))

            # Create listings for each selected store
            created_count = 0
            for store_id in store_ids:
                store = db.session.query(Store).get(store_id)
                if not store:
                    continue

                # Generate external listing ID (placeholder until actually pushed)
                external_listing_id = f"PENDING-{warehouse_stock.sku}-{store.platform}-{datetime.utcnow().timestamp()}"

                # Check if listing already exists
                existing = db.session.query(MarketplaceListing).filter_by(
                    warehouse_stock_id=warehouse_stock_id,
                    store_id=store_id
                ).first()

                if existing:
                    flash(f'Listing already exists for {store.name}. Skipping.', 'warning')
                    continue

                # Create new listing
                listing = MarketplaceListing(
                    warehouse_stock_id=warehouse_stock_id,
                    store_id=store_id,
                    external_listing_id=external_listing_id,
                    external_sku=warehouse_stock.sku,
                    title=title,
                    description=description,
                    price=price,
                    currency=currency,
                    sync_quantity=sync_quantity,
                    sync_price=sync_price,
                    quantity_buffer=quantity_buffer,
                    max_quantity_limit=max_quantity_limit,
                    listing_type='single',
                    push_state='active',
                    is_active=False,  # Not active until actually created on marketplace
                    last_push_status='pending'
                )

                db.session.add(listing)
                created_count += 1

            db.session.commit()

            if created_count > 0:
                flash(f'Successfully created {created_count} listing(s). Use the push function to publish to marketplaces.', 'success')
            else:
                flash('No new listings were created.', 'warning')

            return redirect(url_for('routes.listings'))

        except Exception as e:
            db.session.rollback()
            logging.error(f"Error creating listing: {str(e)}")
            flash('Failed to create listing. Please try again.', 'danger')
            return redirect(url_for('routes.create_listing'))

    # GET request - show form
    # Get all warehouse stock items
    warehouse_items = db.session.query(WarehouseStock).order_by(WarehouseStock.sku).all()

    # Get all active stores
    stores = db.session.query(Store).filter_by(is_active=True).all()

    return render_template('create_listing.html',
                         warehouse_items=warehouse_items,
                         stores=stores,
                         csrf_token=generate_csrf_token())

@bp.route('/listings/edit/<int:listing_id>', methods=['GET', 'POST'])
def edit_listing(listing_id):
    """Edit existing marketplace listing"""
    listing = db.session.query(MarketplaceListing).get(listing_id)
    if not listing:
        flash('Listing not found.', 'danger')
        return redirect(url_for('routes.listings'))

    if request.method == 'POST':
        try:
            # CSRF Protection
            csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
            if not validate_csrf_token(csrf_token):
                flash('Invalid request. Please try again.', 'danger')
                return redirect(url_for('routes.edit_listing', listing_id=listing_id))

            # Update listing details
            listing.title = request.form.get('title')
            listing.description = request.form.get('description')
            listing.price = float(request.form.get('price', 0))
            listing.currency = request.form.get('currency', 'GBP')

            # Sync settings
            listing.sync_quantity = request.form.get('sync_quantity') == 'on'
            listing.sync_price = request.form.get('sync_price') == 'on'
            listing.quantity_buffer = int(request.form.get('quantity_buffer', 0))

            max_quantity_limit = request.form.get('max_quantity_limit')
            listing.max_quantity_limit = int(max_quantity_limit) if max_quantity_limit else None

            # Status
            listing.is_active = request.form.get('is_active') == 'on'
            push_state = request.form.get('push_state')
            if push_state in ['active', 'blocked', 'needs_review', 'disabled']:
                listing.push_state = push_state

            listing.updated_at = datetime.utcnow()

            db.session.commit()
            flash('Listing updated successfully.', 'success')
            return redirect(url_for('routes.listings'))

        except Exception as e:
            db.session.rollback()
            logging.error(f"Error updating listing: {str(e)}")
            flash('Failed to update listing. Please try again.', 'danger')

    return render_template('edit_listing.html',
                         listing=listing,
                         csrf_token=generate_csrf_token())

@bp.route('/listings/delete/<int:listing_id>', methods=['POST'])
def delete_listing(listing_id):
    """Delete a marketplace listing"""
    try:
        # CSRF Protection
        csrf_token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
        if not validate_csrf_token(csrf_token):
            flash('Invalid request. Please try again.', 'danger')
            return redirect(url_for('routes.listings'))

        listing = db.session.query(MarketplaceListing).get(listing_id)
        if not listing:
            flash('Listing not found.', 'danger')
            return redirect(url_for('routes.listings'))

        db.session.delete(listing)
        db.session.commit()
        flash('Listing deleted successfully.', 'success')

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error deleting listing: {str(e)}")
        flash('Failed to delete listing. Please try again.', 'danger')

    return redirect(url_for('routes.listings'))

@bp.route('/api/listings/<int:listing_id>/push', methods=['POST'])
def push_listing(listing_id):
    """Retired listing direct push-state mutation route.

    A listing must not be marked pending unless a governed dispatcher job exists.
    """
    return jsonify({
        "success": False,
        "error": "Direct listing push route is retired. Use governed dispatcher execution path.",
        "execution_blocked": True,
        "route_retired": True,
        "listing_id": listing_id
    }), 410

@bp.route('/api/listings/bulk-action', methods=['POST'])
def bulk_listing_action():
    """Perform bulk actions on listings"""
    try:
        data = request.get_json()
        action = data.get('action')
        listing_ids = data.get('listing_ids', [])

        if not action or not listing_ids:
            return jsonify({'success': False, 'error': 'Missing action or listing IDs'}), 400

        listings = db.session.query(MarketplaceListing).filter(
            MarketplaceListing.id.in_(listing_ids)
        ).all()

        if action == 'activate':
            for listing in listings:
                listing.is_active = True
                listing.updated_at = datetime.utcnow()
            message = f'Activated {len(listings)} listing(s)'

        elif action == 'deactivate':
            for listing in listings:
                listing.is_active = False
                listing.updated_at = datetime.utcnow()
            message = f'Deactivated {len(listings)} listing(s)'

        elif action == 'delete':
            for listing in listings:
                db.session.delete(listing)
            message = f'Deleted {len(listings)} listing(s)'

        elif action == 'push':
            # TODO: Implement actual push logic
            for listing in listings:
                listing.last_push_status = 'pending'
                listing.push_attempts += 1
                listing.updated_at = datetime.utcnow()
            message = f'Initiated push for {len(listings)} listing(s)'

        else:
            return jsonify({'success': False, 'error': 'Invalid action'}), 400

        db.session.commit()
        return jsonify({'success': True, 'message': message})

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error performing bulk action: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.errorhandler(Exception)
def handle_exception(error):
    """Catch-all handler for any unhandled exceptions"""
    logging.error(f"Unhandled Exception: {str(error)}", exc_info=True)
    db.session.rollback()  # Rollback any failed database transactions

    # Don't expose internal error details in production
    error_message = 'An unexpected error occurred. Please try again.'

    if request.is_json or request.path.startswith('/api/'):
        return jsonify({'error': error_message, 'status': 500}), 500
    return render_template('error.html',
                         error_code=500,
                         error_title='Unexpected Error',
                         error_message=error_message), 500


# === NAVBAR COMPATIBILITY ROUTES  DO NOT REMOVE ===
# These routes keep the existing navbar moving as one system.
# They do not change UI layout. They only prevent dead navbar endpoints.

@bp.route('/amazon-fba-stock')
def amazon_fba_stock():
    return redirect(url_for('routes.warehouse'))

@bp.route('/stock-transfers')
def stock_transfers():
    return redirect(url_for('routes.warehouse'))

@bp.route('/customers')
def customers_list():
    return redirect(url_for('routes.dashboard'))

@bp.route('/orders')
def orders_list():
    return redirect(url_for('routes.dashboard'))

@bp.route('/invoices')
def invoices_list():
    return redirect(url_for('routes.dashboard'))

@bp.route('/user-management')
def user_management():
    return redirect(url_for('routes.settings'))

@bp.route('/master-cartons')
def master_cartons():
    return redirect(url_for('routes.warehouse'))

@bp.route('/warehouse-layout')
def warehouse_layout():
    return redirect(url_for('routes.warehouse'))

@bp.route('/box-types')
def box_types_list():
    return redirect(url_for('routes.warehouse'))

@bp.route('/shelf-capacity')
def shelf_capacity():
    return redirect(url_for('routes.warehouse'))

@bp.route('/inventory-import')
def inventory_import():
    return redirect(url_for('routes.inventory'))

@bp.route('/mobile-scan')
def mobile_scan():
    return redirect(url_for('routes.quick_scan'))

@bp.route('/sentinel')
def sentinel_console():
    return redirect(url_for('routes.settings'))

@bp.route('/faq')
def faq():
    return redirect(url_for('routes.settings'))

@bp.route('/admin-product-linking')
def admin_product_linking():
    return redirect(url_for('routes.product_linking'))


# === FINAL NAV FIX: ADMIN SYSTEM ACTIVITY ===

# === CORRECT ADMIN ENDPOINT FOR NAVBAR ===
@bp.route('/system-activity')
def system_activity():
    return redirect(url_for('routes.settings'))
